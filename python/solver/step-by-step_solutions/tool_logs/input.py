
import itertools
from operator import itemgetter
from openpyxl.reader.excel import load_workbook

from generator.model import Instance, EMPTY_CHAR, DUMMY_CHAR
from generator.algo_human import solve_using_human_techniques
from generator.boxes import preprocess_layout_boxes, create_default_layout_boxes


# TODO Make this variable
LENGTH = 3  # TODO Remove
SIZE = 9


# One does not create a function for single used - one shall reuse or remove
def _read_char(sheet, coords):
    # Make sure the characters are in string format
    char = (lambda value: str(value) if value else EMPTY_CHAR)(
        sheet.cell(*coords).value
    )
    return char


# TODO This logic can't read instances produced by tool_patterns
def read_instance(file_name):

    print("Reading instance..")

    workbook = load_workbook(file_name)
    sheet = workbook.worksheets[0]

    # The tool should be able to handle output files produces by both tool_create and tool_patterns, which are
    #  currently significantly different; First we identify the type of output file, locate the required data, and use
    #  common functionality to read it;
    # TODO Create a more consistent/coherent data structure for the output files of both tools

    print(" reading size/layout..")

    size = SIZE

    print(f"  size: {size}")

    if sheet.cell(1, 1).value == "Size":
        assert sheet.cell(1, 2).value == str(SIZE), "Can only use tool for instances of size 9"
        # Make determining the idxs a little versatile, in case whitespaces are modified
        idx_solution = None
        idx_instance = None
        for _idx_row in range(sheet.max_row):
            _idx = (1 + _idx_row, 1)
            _value = sheet.cell(*_idx).value
            if _value == "Solution":
                idx_solution = (1 + _idx_row + 1, 1)
            elif _value == "Instance":
                idx_instance = (1 + _idx_row + 1, 1)
        assert idx_solution is not None, "Could not locate the solution in the input file!"
        assert idx_instance is not None, "Could not locate the instance in the input file!"
        idx_custom_layout = (idx_solution[0], idx_solution[1] + SIZE + 1)
    else:
        idx_instance = (1, 1)
        idx_custom_layout = (1, size + 2)

    print(" reading grid..")

    instance = [
        [
            _read_char(sheet, (idx_instance[0] + i1, idx_instance[1] + i2))
            for i2 in range(size)
        ]
        for i1 in range(size)
    ]

    # TODO Aggregate logic of reading instances with tool_create/logs.py
    print(" reading custom layout..")

    # Optionally: Read custom boxes layout
    if sheet.cell(*idx_custom_layout).value:

        raw_data = [
            [
                sheet.cell(idx_custom_layout[0] + i1, + idx_custom_layout[1] + i2).value
                for i2 in range(size)
            ]
            for i1 in range(size)
        ]

        layout_boxes = preprocess_layout_boxes(raw_data, size)

    else:
        layout_boxes = None

        print("  not specified")

    # Initialise layout if not defined in file
    layout_boxes_for_validation = layout_boxes or create_default_layout_boxes(LENGTH, LENGTH)

    # Validate instance
    _validate_instance(instance, layout_boxes_for_validation)

    instance = Instance(instance, layout_boxes=layout_boxes)

    return instance


# TODO Does this work for a list of instances produced both by tool_create and tool_patterns?
def read_list_of_instances(file_name, should_screen_duplicate_solutions=True):
    """
    Steps
      1 Read data
      2 Verify instances
      3 Ask for continuation confirmation
    """

    print("Reading list of instances..")

    workbook = load_workbook(file_name)
    sheet = workbook.worksheets[0]

    # Approach: First read all rows, then group and process; Sometimes the lines are not ordered correctly

    SKIP_NUMBER_ROWS = 1

    # Reading
    print("Reading data..")
    rows = []
    # This is an expensive operation (not a constant!) and should not be done in the loop; Figured this out as this is
    #  done here {SIZE} times more often as in the old code, which made it much much slower
    max_row = sheet.max_row
    print(" Number of rows:", max_row)
    for idx_row in range(SKIP_NUMBER_ROWS, max_row):
        instance_id = sheet.cell(1 + idx_row, 1).value
        line_id = int(sheet.cell(1 + idx_row, 2).value)
        row_chars = [_read_char(sheet, (1 + idx_row, 3 + i2)) for i2 in range(SIZE)]
        row_layout = [sheet.cell(1 + idx_row, 3 + SIZE + i2).value for i2 in range(SIZE)]
        rows.append((instance_id, line_id, (row_chars, row_layout)))
        if idx_row % 1_000 == 0:
            print(f"Progress: {idx_row} / {max_row}")

    # Processing (this is basically a more complicated piece of code for the same that can be achieved by using pandas)
    print("Processing data..")
    grouped_lines = []
    grouper = itertools.groupby(rows, itemgetter(0))
    for key, items in grouper:
        instance_id = key
        grouped_rows = list(items)  # Unpack
        assert {row[1] for row in grouped_rows} == set(range(1, SIZE + 1)), \
            f"Instance {instance_id} not correctly specified"
        lines_chars = [[] for _ in range(SIZE)]
        lines_layout = [[] for _ in range(SIZE)]
        for _, line_id, (row_chars, row_layout) in grouped_rows:
            row_idx = line_id - 1
            lines_chars[row_idx] = row_chars
            lines_layout[row_idx] = row_layout
        grouped_lines.append((instance_id, (lines_chars, lines_layout)))

    # TODO Read from file
    layout_boxes_for_validation = create_default_layout_boxes(LENGTH, LENGTH)

    # Validate
    print("Validate instances..")
    instances = []
    for idx, (instance_id, lines) in enumerate(grouped_lines):
        try:
            lines_chars, lines_layout = lines
            # Check whether a layout was specified
            if all(e is None for row in lines_layout for e in row):
                layout_boxes = None
            else:
                layout_boxes = preprocess_layout_boxes(lines_layout, SIZE)
            _validate_instance(lines_chars, layout_boxes or layout_boxes_for_validation)
            # Validated instances can be cast
            instance = Instance(lines_chars, layout_boxes=layout_boxes)
            instances.append((instance_id, instance))
        except AssertionError as e:
            print(f"WARNING - Instance {instance_id} not valid: {str(e)}")
            for line in zip(*lines):
                print(line)

    # Identify duplicate solutions
    if len(instances) > 1 and should_screen_duplicate_solutions:
        _screen_for_duplicate_solutions(instances)

    return instances


# TODO One should not repeat the same logic in different forms throught the code - one removes and reuses
def _validate_instance(instance, layout_boxes):

    # Correct number of characters
    chars = {e for e in set(itertools.chain(*instance)) if e != EMPTY_CHAR}

    # Allow for one dummy character
    if len(chars) == SIZE - 1:
        chars.update({DUMMY_CHAR})

    assert len(chars) == SIZE, f"Number of chars present in the instance not equal to {SIZE}: {chars}"

    # No base constraints violated
    for i1 in range(SIZE):
        chars_row = instance[i1]
        for char in chars:
            assert chars_row.count(char) <= 1, f"Char '{char}' present more than once in row {i1}"

    for i2 in range(SIZE):
        chars_col = [instance[i1][i2] for i1 in range(SIZE)]
        for char in chars:
            assert chars_col.count(char) <= 1, f"Char '{char}' present more than once in col {i2}"

    for idx_box in range(SIZE):
        chars_box = [instance[i1][i2] for (i1, i2) in layout_boxes[idx_box]]
        for char in chars:
            assert chars_box.count(char) <= 1, f"Char '{char}' present more than once in box {idx_box}"


def _screen_for_duplicate_solutions(instances):

    print("Screening for duplicate solutions..")

    print("Solving..")

    solutions = []
    for instance_id, instance in instances:
        try:
            solution, _ = solve_using_human_techniques(instance)
            solutions.append((instance_id, solution))
        except Exception as e:
            print(f"WARNING - Instance {instance_id} not solvable!")
            print("", str(e))

        if len(solutions) % 100 == 0:
            print(f"Progress: {len(solutions)} / {len(instances)}")

    print("Comparing solutions..")

    groups = []
    for idx, (instance_id_1, solution_1) in enumerate(solutions):
        for instance_id_2, solution_2 in solutions[idx + 1:]:
            assert solution_1.size == solution_2.size
            size = solution_1.size

            try:
                mapping = {}
                for i1 in range(size):
                    for i2 in range(size):
                        e1 = solution_1[i1][i2]
                        e2 = solution_2[i1][i2]
                        if e1 in mapping:
                            assert mapping[e1] == e2
                        else:
                            mapping[e1] = e2
                is_exact_mapping = True
            except AssertionError:
                is_exact_mapping = False

            if is_exact_mapping:

                # Identify whether the instance is already in a group
                idx_group = -1
                for i, group in enumerate(groups):
                    if instance_id_1 in group:
                        idx_group = i
                        break

                if idx_group != -1:
                    groups[idx_group].add(instance_id_2)
                else:
                    groups.append({instance_id_1, instance_id_2})

                # print("Found new exact mapping:", instance_id_1, instance_id_2)

        if idx > 0 and idx % 100 == 0:
            print(f"Progress: {idx} / {len(solutions)}")

    if len(groups) > 0:
        print("WARNING - Identified instances with duplicate solutions:")
        for group in groups:
            print(len(group), sorted(group))
