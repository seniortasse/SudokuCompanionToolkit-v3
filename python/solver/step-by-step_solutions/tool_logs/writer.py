
from collections import defaultdict
from collections.abc import Iterable
from copy import copy
import itertools
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Color
from openpyxl.styles.borders import Border
import re

from generator.model import EMPTY_CHAR, DIMENSIONS
from generator.algo_human import BASE_TECHNIQUES, ADVANCED_TECHNIQUES, TECHNIQUES

from tool_logs.template import ROW_START, SIZE, SHEET_NAME_STEPS, read_formatting, COL_START
from tool_logs.messages import generate_message
from tool_logs.layers import identify_layers
from tool_logs.formatting import process_formatting
import tool_logs.messages_templates as messages_templates
from tool_logs.template_messages import read_template


# Template formatting

FILE_NAME_TEMPLATE = "Template.xlsx"

SHEET_NAME_INSTANCE = "Problem"
SHEET_NAME_OUTPUT = "Steps"


# Template messages

FILE_NAME_TEMPLATE_MESSAGES = "Template_Messages.xlsx"


# TODO Read from template
DEFAULT_COLOR_VALUES = "FF000000"
# DEFAULT_FONT_SIZE_MESSAGE = 9
# DEFAULT_COLOR_BG_DIM_BASE = "EA" * 3
# DEFAULT_COLOR_BG_DIM_HELP = "dce6f2"


NO_STEPS_PER_ROW = 2
WRITE_MAIN_HEADER_EVERY_NO_STEPS = 4


NUMBER_ROWS_PER_STEP = (1 + SIZE + 3)


# TODO Heavily refactor
def write_steps(instance, logs_steps, instance_id, data_folder):

    # The data_only mode is needed to read values instead of formulas, which gave a problem for reading message font sizes
    workbook = load_workbook(FILE_NAME_TEMPLATE, data_only=True)

    sheet_format = workbook.get_sheet_by_name(SHEET_NAME_STEPS)

    sheet_output = workbook.copy_worksheet(workbook[SHEET_NAME_STEPS])
    sheet_output.title = SHEET_NAME_OUTPUT

    formatting = read_formatting(workbook)

    # Read templates for messages
    templates_messages = read_template(FILE_NAME_TEMPLATE_MESSAGES)

    ### CLEAN SHEET

    # Apparently merged cells are not automatically unmerged, which gives some problems when writing data
    for cell_range in sheet_output.merged_cell_ranges:
        _cell_from, _cell_to = str(cell_range).split(':')
        _row_from = int(re.findall(r"\d+", _cell_from)[0])
        # Only unmerge for the rows that we want to delete
        # print(str(cell_range))
        if _row_from >= ROW_START:
            # print("Unmerge", str(cell_range))
            sheet_output.unmerge_cells(range_string=str(cell_range))

    # Remove template steps
    sheet_output.delete_rows(ROW_START, sheet_output.max_row)

    ### UPDATE TEMPLATE

    _update_template_header(sheet_output, formatting, instance, instance_id=instance_id)

    ### GENERATE STEPS

    for _i, step_logs in enumerate(logs_steps):
        step_no = _i + 1
        _write_step(sheet_output, step_no, step_logs, formatting, instance_id=instance_id)

    # Remove temporary data
    # sheet.delete_cols(COL_TO, sheet.max_column)


    # Read instance values -> Should already be available in the code, from the instance file (template only used for formatting)
    # instance = workbook.get_sheet_by_name(SHEET_NAME_INSTANCE)

    ### WRITE INSTANCE
    sheet_instance = workbook[SHEET_NAME_INSTANCE]

    for i1 in range(SIZE):
        for i2 in range(SIZE):
            cell = sheet_instance.cell(i1 + 1, i2 + 1)
            cell.value = instance[i1][i2]


    # Rename sheet
    # sheet.title = SHEET_NAME_OUTPUT

    # Copy zoom size
    sheet_output.sheet_view.zoomScale = sheet_format.sheet_view.zoomScale

    # Set focus on cell
    # sheet_output.sheet_view.selection[0].activeCell = f"A{50}"

    # Remove all other sheets before writing to file
    for sheet_name in workbook.get_sheet_names():
        if sheet_name not in [SHEET_NAME_INSTANCE, SHEET_NAME_OUTPUT]:
            workbook.remove(workbook[sheet_name])

    # Activate sheet with steps
    workbook.active = sheet_output

    # workbook.copy_worksheet(workbook[SHEET_NAME_OUTPUT])

    file_name_output = data_folder + instance_id + "_user_logs.xlsx"
    print("Writing output to file:", file_name_output)
    workbook.save(filename=file_name_output)

    return file_name_output


def _update_template_header(sheet, formatting, instance, instance_id="<Problem ID>"):
    """
    Update main header and boxes layouts
    """

    (
        style_header_main, style_header_step, style_cells, style_message, style_borders, colors_boxes,
        colors_new_values, colors_bg, colors_key_cells, colors_bg_cleanup, fonts_cleanup,
        font_sizes,
    ) = formatting

    # Modify the main header on the first row
    _write_main_header(sheet, 1, instance_id, style_header_main)

    # Update borders for template showing boxes
    coords_cells_template_boxes = (17, 2)
    _draw_borders(sheet, coords_cells_template_boxes, instance, style_borders)

    # Update box IDs and add colors
    for (i1, i2) in itertools.product(range(SIZE), repeat=2):
        idx_box = instance.get_idx_box(i1, i2)
        box_id = str(idx_box + 1)  # Convert internal 0-based to user-readable 1-based

        cell = sheet.cell(*tuple(map(sum, zip(coords_cells_template_boxes, (i1, i2)))))
        cell.value = box_id

        fill_box = colors_boxes[idx_box % 2]
        cell.fill = copy(fill_box)


def _write_main_header(sheet, idx_row, instance_id, style_header_main):

    cell = sheet.cell(idx_row, COL_START)
    cell.value = messages_templates.template_header_main.format(instance_id=instance_id)
    cell._style = copy(style_header_main)

    # Merge header cells
    sheet.merge_cells(
        start_row=idx_row, start_column=COL_START,
        end_row=idx_row, end_column=COL_START + NO_STEPS_PER_ROW * SIZE
    )


def _write_step(sheet, step_no, step_logs, formatting, instance_id="<Problem ID>"):

    (
        style_header_main, style_header_step, style_cells, style_message, style_borders, colors_boxes,
        colors_new_values, colors_bg, colors_key_cells, colors_bg_cleanup, fonts_cleanup,
        font_sizes,
    ) = formatting

    # row_step, col_step = coords
    assert step_no >= 1

    # Write a main header for every 4 steps
    write_main_header = (step_no - 1) % WRITE_MAIN_HEADER_EVERY_NO_STEPS == 0
    if write_main_header:
        row_main_header = ROW_START + (step_no - 1) // NO_STEPS_PER_ROW * NUMBER_ROWS_PER_STEP + NO_STEPS_PER_ROW * (step_no - 1) // WRITE_MAIN_HEADER_EVERY_NO_STEPS
        _write_main_header(sheet, row_main_header, instance_id, style_header_main)

    row_start = ROW_START + (step_no - 1) // NO_STEPS_PER_ROW * NUMBER_ROWS_PER_STEP + NO_STEPS_PER_ROW * (1 + (step_no - 1) // WRITE_MAIN_HEADER_EVERY_NO_STEPS)
    col_start = COL_START + (SIZE + 1) * ((step_no - 1) % NO_STEPS_PER_ROW)

    coords_header = (row_start, col_start)
    coords_cells = (row_start + 1, col_start)
    coords_message = (row_start + 1 + SIZE, col_start)

    # Write header
    cell = sheet.cell(*coords_header)
    cell.value = messages_templates.template_header_step.format(step_no=step_no)
    cell._style = copy(style_header_step)

    # Merge header cells
    sheet.merge_cells(
        start_row=coords_header[0], start_column=coords_header[1],
        end_row=coords_header[0], end_column=coords_header[1] + SIZE - 1
    )

    step_instance_before, step_instance_after, step_new_values, step_technique_used, step_is_cleanup_issue, step_cleanup_steps = step_logs

    print()
    print(f"Step {step_no}")
    print(" Technique used:", step_technique_used)
    print(" New values:", len(step_new_values))
    for new_value in step_new_values:
        print("", new_value)

    print("==")
    print("CLEANUP STEPS")
    for idx, (
            cleanup_step_latest_technique, cleanup_step_current_technique, cleanup_step_iteration,
            cleanup_step_options_before_application, cleanup_step_details) in enumerate(step_cleanup_steps):
        if not cleanup_step_details:
            continue
        print(f" Cleanup step {idx + 1}")
        print("  latest technique: ", cleanup_step_latest_technique)
        print("  current technique:", cleanup_step_current_technique)
        print("  iteration:", cleanup_step_iteration)
        print("  --- ")
        for detail in cleanup_step_details:
            print(" ", detail[0][1])
            print("   -> removing", detail[-1])
        print("  --- ")
    print("==")

    # Write grid
    idxs_new_values = [t[0] for t in step_new_values]

    for i1 in range(SIZE):
        for i2 in range(SIZE):
            cell = sheet.cell(coords_cells[0] + i1, coords_cells[1] + i2)
            # if sheet.cell(row, col_from).has_style:
            cell._style = copy(style_cells[i1][i2])

            # Fill custom values
            cell.value = step_instance_after[i1][i2]

            # Modify colors
            # cell_to.font.color.theme = None
            # cell.font.color.rgb = "FF0000"

            if (i1, i2) in idxs_new_values:
                # color_text = "FFFF0000"
                color_text = colors_new_values[step_technique_used]
                # print(i1, i2, color_text)
                # print(cell.font)
            else:
                color_text = DEFAULT_COLOR_VALUES

            # Make sure the rgb vaue is used and not the theme value
            # cell.font = copy(cell.font)
            # if i1 == i2 == 0:
            #     cell.font.color.type = "rgb"
            #     cell.font.color.rgb = color_text
            #     print((i1 + 1, i2 + 1), cell.font.color)
            # It is not possible to change the color attribute of the font, a new font has to be created
            cell.font = Font(name=cell.font.name, size=cell.font.size, color=color_text)
            # cell.font.color = Color(rgb=color_text)

    # Optionally: Draw borders (only when a custom layout for boxes is used)
    if step_instance_before.uses_custom_boxes_layout:
        _draw_borders(sheet, coords_cells, step_instance_before, style_borders)

    # color_bg = "FFFF0000"
    # color_bg = colors_bg[step_technique_used]
    color_bg_singles = colors_bg["singles-1"]  # Used for advanced techniques
    # color_bg_cleanup = colors_bg_cleanup[step_technique_used]

    # fill_bg = PatternFill("solid", start_color=color_bg)
    fill_bg_singles = PatternFill("solid", start_color=color_bg_singles)
    # fill_bg_cleanup = PatternFill("solid", start_color="00FFFF00")
    # fill_bg_cleanup = PatternFill("solid", start_color=color_bg_cleanup)

    fills_bg_key = {
        name_technique: PatternFill("solid", start_color=colors_bg[name_technique])
        for name_technique in TECHNIQUES
    }

    fills_bg_cln = {
        name_technique: PatternFill("solid", start_color=colors_bg_cleanup[name_technique])
        for name_technique in TECHNIQUES
    }

    # Instead of only the color we can use the entire font
    # font_key_cells = colors_key_cells[step_technique_used]
    # font_cleanup = fonts_cleanup[step_technique_used]

    # Background coloring
    # TODO Somehow combine with generating messages, ie which dims are used
    if step_technique_used == "singles-1":
        for new_value in step_new_values:
            coords = new_value[0]
            char = new_value[1]
            dim_base = new_value[2].split(' & ')[0]
            idxs = step_instance_before.get_idxs_in_dim(dim_base, *coords)
            for (i1, i2) in idxs:
                if (i1, i2) != coords:  # Do not fill the new value cell
                    cell = sheet.cell(coords_cells[0] + i1, coords_cells[1] + i2)
                    fill_bg = fills_bg_key[step_technique_used]
                    cell.fill = fill_bg

    elif step_technique_used == "singles-2":

        # Only color helper cells when only one value was found (note: both the empty cells in the same dimension as
        #  well as the cells with the same value in the helper dimensions are considered value cells)
        if len(step_new_values) == 1:

            for new_value in step_new_values:
                coords = new_value[0]
                char = new_value[1]
                dim_base, dim_help = new_value[2].split(' & ')[0].split(' + ')
                idxs_base = step_instance_before.get_idxs_in_dim(dim_base, *coords)

                for (i1, i2) in idxs_base:
                    if step_instance_after[i1][i2] == EMPTY_CHAR:
                        cell = sheet.cell(coords_cells[0] + i1, coords_cells[1] + i2)
                        fill_bg = fills_bg_key[step_technique_used]
                        cell.fill = fill_bg

                        # Find values in help dim
                        for (_i1, _i2) in step_instance_before.get_idxs_in_dim(dim_help, i1, i2):
                            if step_instance_after[_i1][_i2] == char and (_i1, _i2) != coords:
                                cell = sheet.cell(coords_cells[0] + _i1, coords_cells[1] + _i2)
                                fill_bg = fills_bg_key[step_technique_used]
                                cell.fill = fill_bg

    elif step_technique_used == "singles-3":

        # Only color helper cells when only one value was found
        if len(step_new_values) == 1:

            for new_value in step_new_values:
                coords = new_value[0]
                char = new_value[1]
                dim_base = new_value[2].split(' & ')[0]
                dims_help = [dim for dim in DIMENSIONS if dim != dim_base]
                idxs_base = step_instance_before.get_idxs_in_dim(dim_base, *coords)

                # This is cleaner code -- Even better to actually fill the cells after the entire if/else statement, and
                #  define this collection before
                # idxs_to_fill = []
                #
                # for (i1, i2) in idxs_base:
                #     if step_instance[i1][i2] == EMPTY_CHAR:
                #         idxs_to_fill.append((i1, i2))
                #
                # for (_i1, _i2) in idxs_base:
                #     if step_instance[_i1][_i2] == EMPTY_CHAR:
                #         for dim in dims_help:
                #             idxs_help = get_idxs_for_dimension(dim, _i1, _i2)
                #             for (i1, i2) in idxs_help:
                #                 if step_instance[i1][i2] == char and (i1, i2) != coords:
                #                     idxs_to_fill.append((i1, i2))

                idxs_to_fill = find_idxs_color_final_value(coords, dim_base, char, step_instance_after, defaultdict(list))

                for (i1, i2) in idxs_to_fill:
                    cell = sheet.cell(coords_cells[0] + i1, coords_cells[1] + i2)
                    fill_bg = fills_bg_key[step_technique_used]
                    cell.fill = fill_bg

    elif step_technique_used == "singles-naked-2":
        for new_value in step_new_values:
            coords = new_value[0]
            char = new_value[1]
            dims = new_value[2].split(' & ')[0].split(' + ')
            for dim in dims:
                idxs = step_instance_before.get_idxs_in_dim(dim, *coords)
                for (i1, i2) in idxs:
                    if (i1, i2) != coords and step_instance_after[i1][i2] != EMPTY_CHAR:
                        cell = sheet.cell(coords_cells[0] + i1, coords_cells[1] + i2)
                        fill_bg = fills_bg_key[step_technique_used]
                        cell.fill = fill_bg

    elif step_technique_used == "singles-naked-3":
        for new_value in step_new_values:
            coords = new_value[0]
            for dim in DIMENSIONS:
                idxs = step_instance_before.get_idxs_in_dim(dim, *coords)
                for (i1, i2) in idxs:
                    if (i1, i2) != coords and step_instance_after[i1][i2] != EMPTY_CHAR:
                        cell = sheet.cell(coords_cells[0] + i1, coords_cells[1] + i2)
                        fill_bg = fills_bg_key[step_technique_used]
                        cell.fill = fill_bg

    else:
        assert step_technique_used in ADVANCED_TECHNIQUES

        # Reuse same functionality across advanced techniques
        assert len(step_new_values) == 1
        new_value = step_new_values[0]

        assert len(new_value) == 4  # Generalised structure for all techniques
        coords = new_value[0]
        char = new_value[1]
        dim_final_value = new_value[2].split(' & ')[0]
        details = new_value[3]

        # Identify relevant applications
        # UPDATE: Instead of only showing the final technique applied, we now show all applications used for finding
        #  the new value (keys, layer 1), as well as the applications that were needed to make those available
        #  (cleanup, layer 2) - besides the final value cells (layer 0)
        # TODO This should be only done once in writer/messages
        print(" ~~~~~~ Identify layers ")
        applications_layer_1, applications_layer_2, unused_applications = identify_layers(
            step_instance_before, step_cleanup_steps, new_value
        )

        # List to be updated by each technique
        idxs_bg_color_key_cells = []
        idxs_bg_color_cleanup_cells = []

        # Extra customisation highlighting cells
        idxs_bg_color_key_cells_highlighted = []
        idxs_bg_color_cleanup_cells_highlighted = []

        # For singles, we use the claiming options of key cells to determine which cells to color; For key cells, we
        #  should use claiming options of cleanup cells, plus the cell of the relevant options they removed
        idxs_bg_color_key_helper_cells = []

        # For some techniques the cells containing the key values (eg wings, rays) exclude options for the value cell,
        #  and should be colored - since they do not yet have a value in the instance they are not detected through the
        #  generalised logic and should be added separately for each technique
        # idxs_extra_color_singles = []
        idxs_claiming_options_key = defaultdict(list)
        idxs_claiming_options_cleanup = defaultdict(list)

        # First store which values have to be shown for multiple applications, then merge them if there are more than
        #  one in a cell
        show_options_key = defaultdict(list)
        show_options_cleanup = defaultdict(list)

        for application in applications_layer_1:
            name_technique, name_application = application[0]
            print(f"Process formatting for layer 1 application: {name_application}")
            _idxs_color_cells, _show_values, _idxs_claiming_options, _idxs_bg_color_cells_highlighted = process_formatting(application, step_instance_after, coords)
            idxs_bg_color_key_cells.append((name_technique, _idxs_color_cells))
            idxs_bg_color_key_cells_highlighted.append((name_technique, _idxs_bg_color_cells_highlighted))
            for k, v in _show_values.items():
                show_options_key[k].append((name_technique, v))
            # show_key_values.append((name_technique, _show_values))
            for k, v in _idxs_claiming_options.items():
                idxs_claiming_options_key[k].extend(v)

        def format_recursively(earlier_applications):
            for application in earlier_applications:
                (name_technique, name_application), _, removed_chars = application
                print(f"Process formatting for layer 2 application: {name_application}")
                _idxs_color_cells, _show_values, _idxs_claiming_options, _idxs_bg_color_cells_highlighted = process_formatting(application, step_instance_after, coords)
                idxs_bg_color_cleanup_cells.append((name_technique, _idxs_color_cells))
                idxs_bg_color_cleanup_cells_highlighted.append((name_technique, _idxs_bg_color_cells_highlighted))
                for k, v in _show_values.items():
                    show_options_cleanup[k].append((name_technique, v))
                # show_cleanup_values.append((name_technique, _show_values))
                for k, v in _idxs_claiming_options.items():
                    idxs_claiming_options_cleanup[k].extend(v)
                # The helper key cells consist of the claiming options and the removed values of the layer 2 application
                _idxs_bg_color_key_helper_cells = list(_idxs_claiming_options.keys()) + [idx for idx, char in removed_chars]
                # Only add this for techniques using partial information, the ones using options to remove a value are
                #  typically not colored
                # TODO We could return this information from the function above
                # is_group_2_technique = all(chars.startswith('-') for v in _idxs_claiming_options.values() for chars in v)
                is_group_2_technique = name_technique in [
                    "singles-pointing", "singles-boxed",
                    "x-wings", "x-wings-3", "x-wings-4",
                    "boxed-wings", "boxed-ray",
                ]
                if is_group_2_technique:
                    idxs_bg_color_key_helper_cells.append((name_technique, _idxs_bg_color_key_helper_cells))
                print(f"Add key helper cells for application {name_application}:", _idxs_bg_color_key_helper_cells)
                format_recursively(applications_layer_2[name_application])

        for application in applications_layer_1:
            name_application = application[0][1]
            format_recursively(applications_layer_2[name_application])

        # POST-PROCESSING

        # Priority of showing values: value, key, cleanup; group 1 over group 2
        # With this priority, group 1 cleanup values get priority over group 2 key values

        def _wrap_chars(chars, max_chars_per_line):
            return '\n'.join(
                chars[i * max_chars_per_line:(i + 1) * max_chars_per_line]
                for i in range((len(chars) + max_chars_per_line - 1) // max_chars_per_line)
            )

        def combine_char_list(char_list):
            # Two types of values: group 1 (all options) and group 2 (partial options)
            chars_group_1 = [chars for chars in char_list if not isinstance(chars, str)]
            chars_group_2 = [chars for chars in char_list if isinstance(chars, str)]
            # TODO Temporary checks, now that options are not preprocessed (joined and sorted) in formatting
            assert all(isinstance(chars, Iterable) for chars in chars_group_1)
            assert all(chars.startswith('-') for chars in chars_group_2)
            assert len(chars_group_1) + len(chars_group_2) > 0
            # Group 1 gets priority over group 2
            if len(chars_group_1) > 0:
                # Take the options of the application applied last in the process
                #  -> Instead of introducing some custom logic identifying the latest application (which information we
                #   currently don't have here), simply take the one with the least number of values, which by definition
                #   should be the options seen by the last application as options only get less after more applications
                chars_sorted = sorted(chars_group_1, key=len)
                # Options should always get less and should not have any new options added
                assert all(set(chars_sorted[i]).issubset(chars_sorted[i + 1]) for i in range(len(chars_sorted) - 1))
                chars_selected = chars_sorted[0]
                # Centralised processing which before we did in formatting for each technique separately
                # Note: We assume that the chars are of type str by default
                show_value = ''.join(sorted(chars_selected))
                # Extra processing step: When the number of options is large, wrap onto multiple lines
                show_value = _wrap_chars(show_value, 4)
                group = 1
            else:
                # Concatenate unique chars
                show_value = '-' + ''.join(sorted(set(chars.lstrip('-') for chars in char_list)))
                group = 2
            return show_value, group

        # Keep track of group 1 values shown for cleanup, which get priority over group 2 values for key
        idxs_cleanup_values_group_1 = []

        # Show cleanup values
        for (i1, i2), _v in show_options_cleanup.items():
            assert step_instance_before[i1][i2] == EMPTY_CHAR  # We should only show options for empty cells
            name_technique = [name_technique for name_technique, _ in _v][0]  # TODO Which one to take?
            char_list = list(itertools.chain.from_iterable(char_list for _, char_list in _v))
            if (i1, i2) not in idxs_new_values:
                cell = sheet.cell(coords_cells[0] + i1, coords_cells[1] + i2)
                # TODONE Reuse processing from key values
                # Some extra processing to avoid the result becoming too long
                show_value, group = combine_char_list(char_list)
                if group == 1:
                    idxs_cleanup_values_group_1.append((i1, i2))
                cell.value = show_value
                font_cleanup = fonts_cleanup[name_technique]
                cell.font = copy(font_cleanup)

        # Show key values
        for (i1, i2), _v in show_options_key.items():
            assert step_instance_before[i1][i2] == EMPTY_CHAR  # We should only show options for empty cells
            name_technique = [name_technique for name_technique, _ in _v][0]  # TODO Which one to take?
            char_list = list(itertools.chain.from_iterable(char_list for _, char_list in _v))
            # Note: For singles-boxed, we found a case where one of the key cells is also where a new value is found,
            #  we do not want to overwrite the new value cell
            if (i1, i2) not in idxs_new_values:
                cell = sheet.cell(coords_cells[0] + i1, coords_cells[1] + i2)
                # Some extra processing to avoid the result becoming too long
                show_value, group = combine_char_list(char_list)
                if group == 2 and (i1, i2) in idxs_cleanup_values_group_1:
                    # Group 1 cleanup values get priority over group 2 key values
                    pass
                else:
                    cell.value = show_value
                    # TODO The font size could better be based on the number of options, which for the new technique
                    #  leftovers can be any amount, but also when merging options the original intended size could not
                    #  be suitable anymore
                    font_key_cells = colors_key_cells[name_technique]
                    cell.font = copy(font_key_cells)

        # Priority of background coloring: cleanup, key, value
        # By ordering the coloring logic the latest block gets the highest priority

        # Background coloring for final value
        # TODONE Make this already generalised functionality available for all techniques
        # TODONE Add this as a final code block for all advanced techniques
        # TODO Be able to add idxs (eg for pointing singles all options in the ray should be colored)
        idxs_color_singles = find_idxs_color_final_value(coords, dim_final_value, char, step_instance_after, idxs_claiming_options_key)
        for (i1, i2) in idxs_color_singles:
            cell = sheet.cell(coords_cells[0] + i1, coords_cells[1] + i2)
            cell.fill = fill_bg_singles

        # Background coloring for key cells
        for name_technique, idxs in idxs_bg_color_key_cells + idxs_bg_color_key_helper_cells:
            for (i1, i2) in idxs:
                cell = sheet.cell(coords_cells[0] + i1, coords_cells[1] + i2)
                fill_bg_key = fills_bg_key[name_technique]
                cell.fill = fill_bg_key
                # print(f"Set bg for key cell {(i1, i2)}:", fill_bg_key.start_color)

        # Highlighted cells overwrite the custom ones, some cells might be included in both
        # TODO What about the priority between key and key helper cells?
        for name_technique, idxs in idxs_bg_color_key_cells_highlighted:
            for (i1, i2) in idxs:
                cell = sheet.cell(coords_cells[0] + i1, coords_cells[1] + i2)
                fill_bg_key = fills_bg_key[name_technique]
                # Convert to darker shade
                fill_bg_key_highlighted = copy(fill_bg_key)
                fill_bg_key_highlighted.start_color.rgb = _darken_color(fill_bg_key_highlighted.start_color.rgb)
                cell.fill = fill_bg_key_highlighted

        # Background coloring for cleanup cells
        for name_technique, idxs in idxs_bg_color_cleanup_cells:
            for (i1, i2) in idxs:
                cell = sheet.cell(coords_cells[0] + i1, coords_cells[1] + i2)
                fill_bg_cln = fills_bg_cln[name_technique]
                cell.fill = fill_bg_cln
                # print(f"Set bg for cln cell {(i1, i2)}:", fill_bg_cln.start_color)

        # Also highlight cleanup cells
        for name_technique, idxs in idxs_bg_color_cleanup_cells_highlighted:
            for (i1, i2) in idxs:
                cell = sheet.cell(coords_cells[0] + i1, coords_cells[1] + i2)
                fill_bg_cln = fills_bg_cln[name_technique]
                # Convert to darker shade
                fill_bg_cln_highlighted = copy(fill_bg_cln)
                fill_bg_cln_highlighted.start_color.rgb = _darken_color(fill_bg_cln_highlighted.start_color.rgb)
                cell.fill = fill_bg_cln_highlighted

        # print("Show cln:", show_cleanup_values)
        # print("Show key:", show_key_values)
        # print("Col Sin:", idxs_color_singles)
        # print("Col key:", idxs_color_key_cells)
        # print("Col key help:", idxs_color_key_helper_cells)
        # print("Col cln:", idxs_color_cleanup_cells)
        # print("Claim key:", sorted(idxs_claiming_options_key))
        # print("Claim cln:", sorted(idxs_claiming_options_cleanup))

    # Write message
    message = generate_message(step_logs)

    cell = sheet.cell(*coords_message)
    cell.value = message
    cell._style = copy(style_message)
    # Font size based on number of new values found
    # if step_technique_used in BASE_TECHNIQUES:
    #     font_size_counter = len(step_new_values)
    # else:
    #     font_size_counter = len(relevant_applications)
    # font_size = font_sizes.get(step_technique_used, {}).get(font_size_counter, DEFAULT_FONT_SIZE_MESSAGE)
    # Font size based on message length
    message_length = len(message)
    # The font sizes are defined for message length "up to" the value indicated
    max_message_length = max(font_sizes.keys())
    if message_length > max_message_length:
        font_size = font_sizes[max_message_length]
    else:
        filtered_font_sizes = {k: v for k, v in font_sizes.items() if message_length <= k}
        font_size = font_sizes[min(filtered_font_sizes.keys())]
    print(f"Message font size: {font_size}, based on message length: {message_length}")
    cell.font = Font(name=cell.font.name, size=font_size)

    # Merge message cells
    sheet.merge_cells(
        start_row=coords_message[0], start_column=coords_message[1],
        end_row=coords_message[0] + 2, end_column=coords_message[1] + SIZE - 1
    )


def _rgb_to_hex(value):
    assert 0 <= value < 256
    value_rounded = int(value)
    hex_value = "{:0>2}".format(hex(value_rounded).lstrip("0x"))
    assert int(hex_value, 16) == value_rounded
    return hex_value


def _darken_color(hex_color_code):
    assert len(hex_color_code) == 8
    a, rgb = hex_color_code[:2], hex_color_code[2:]
    # The opaque value seems irrelevant
    rgb_tuple = tuple(int(rgb[2 * i:2 * (i + 1)], 16) for i in range(3))
    assert all(0 <= e < 256 for e in rgb_tuple)
    # By reducing the rgb values proportionally, the color gets darker
    factor = 0.816
    darkened_rgb_tuple = tuple(e * factor for e in rgb_tuple)
    # Convert back to hex
    darkened_hex_color_code = a + "".join(_rgb_to_hex(e) for e in darkened_rgb_tuple)
    return darkened_hex_color_code


def _draw_borders(sheet, coords, step_instance_before, style_borders):

    # Preprocess the border styles from the formatting
    border_style_boundary = style_borders.left
    border_style_regular = style_borders.right

    # Draw borders
    for (i1, i2) in itertools.product(range(SIZE), repeat=2):
        idx_box = step_instance_before.get_idx_box(i1, i2)

        draw_left_border = (step_instance_before.get_idx_box(i1, i2 - 1) != idx_box) if i2 - 1 >= 0 else True
        draw_right_border = (step_instance_before.get_idx_box(i1, i2 + 1) != idx_box) if i2 + 1 < SIZE else True
        draw_top_border = (step_instance_before.get_idx_box(i1 - 1, i2) != idx_box) if i1 - 1 >= 0 else True
        draw_bottom_border = (step_instance_before.get_idx_box(i1 + 1, i2) != idx_box) if i1 + 1 < SIZE else True

        cell = sheet.cell(coords[0] + i1, coords[1] + i2)
        cell.border = Border(
            left=border_style_boundary if draw_left_border else border_style_regular,
            right=border_style_boundary if draw_right_border else border_style_regular,
            top=border_style_boundary if draw_top_border else border_style_regular,
            bottom=border_style_boundary if draw_bottom_border else border_style_regular,
        )


def find_idxs_color_final_value(coords, dim_final_value, char, step_instance_after, idxs_claiming_options):
    """
    idxs_claiming_options : contains an additional collection (map of idx -> list of chars) with characters enforced by
      the technique, which should be taken into account for the coloring

    """

    # If the final value is found based on only occurrence in cell, different background coloring is needed
    assert dim_final_value in DIMENSIONS + ["cell"]

    idxs = []

    if dim_final_value == "cell":
        for dim in DIMENSIONS:
            for (i1, i2) in step_instance_after.get_idxs_in_dim(dim, *coords):
                if step_instance_after[i1][i2] != EMPTY_CHAR and (i1, i2) != coords:
                    idxs.append((i1, i2))
                # Note: The char does not need to be equal to the one in the key cell - in fact it cannot in this case
                elif (i1, i2) in idxs_claiming_options and (i1, i2) != coords:
                    # print(f"~~ Coloring of {(i1, i2)} based on claiming option!")
                    idxs.append((i1, i2))

    else:
        dims_help = [dim for dim in DIMENSIONS if dim != dim_final_value]

        # We only want to show occurrence in one dimension, with priority defined below
        # dims_help = sorted(dims_help, key=lambda dim: ["box", "row", "col"].index(dim))

        for (i1, i2) in step_instance_after.get_idxs_in_dim(dim_final_value, *coords):
            if step_instance_after[i1][i2] == EMPTY_CHAR:  # Note: The final value is already filled in and not empty
                # Only apply background coloring if there actually is a value present in one of the help dims
                should_color = False
                for dim_help in dims_help:
                    # This check makes sure we only identify a helper cell in one dimension -> for now disabled
                    # if (i1, i2) in idxs:
                    #     print(f"~~ Skipping checking in dim {dim_help} as {(i1, i2)} is already covered by another helper dim")
                    #     continue
                    for (_i1, _i2) in step_instance_after.get_idxs_in_dim(dim_help, i1, i2):
                        if (step_instance_after[_i1][_i2] == char) and (_i1, _i2) != coords:
                            should_color = True
                            idxs.append((_i1, _i2))
                        # TODO The char needs not be present in the claiming options (eg for doubles, it excludes an
                        #  option which is not present in the remaining options (see pairs-6 step 7)
                        elif (char in idxs_claiming_options[(_i1, _i2)]) and (_i1, _i2) != coords:
                            # print(f"~~ Coloring of {(_i1, _i2)} based on claiming option!")
                            should_color = True
                            idxs.append((_i1, _i2))
                # We have to add it in the for-loop to make sure the check at the start of the for-loop works
                if should_color:
                    idxs.append((i1, i2))

        # TODO We could do an extra check that if the value is already covered by the box (easier), we do
        #  not format the ones found in row/col

    return idxs
