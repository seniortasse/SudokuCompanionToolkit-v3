from __future__ import annotations

import ast
import json
import subprocess
import sys
from dataclasses import dataclass
from itertools import cycle, islice
from pathlib import Path
from typing import Dict, List, Mapping, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook

from python.publishing.pattern_library.pattern_filters import filter_patterns
from python.publishing.pattern_library.pattern_registry import PatternRegistry
from python.publishing.schemas.models import PatternRecord


LEGACY_GENERATOR_VERSION = "legacy_pattern_tool_v1"


@dataclass
class PatternProductionRequest:
    request_id: str
    pattern: PatternRecord


@dataclass
class PatternProductionRunResult:
    candidates: List[dict]
    rejected: List[dict]
    stdout: str
    stderr: str
    input_workbook: Path
    output_workbook: Path


def select_patterns_from_catalog(
    *,
    registry: PatternRegistry,
    pattern_ids: Optional[Sequence[str]] = None,
    family_ids: Optional[Sequence[str]] = None,
    tags_any: Optional[Sequence[str]] = None,
    min_clue_count: Optional[int] = None,
    max_clue_count: Optional[int] = None,
    min_aesthetic_score: Optional[float] = None,
    min_print_score: Optional[float] = None,
    min_legibility_score: Optional[float] = None,
    max_patterns: Optional[int] = None,
) -> List[PatternRecord]:
    patterns = filter_patterns(
        registry.patterns,
        status="active",
        min_clue_count=min_clue_count,
        max_clue_count=max_clue_count,
        min_aesthetic_score=min_aesthetic_score,
        min_print_score=min_print_score,
        min_legibility_score=min_legibility_score,
    )

    if pattern_ids:
        allowed = {str(x).strip() for x in pattern_ids if str(x).strip()}
        patterns = [p for p in patterns if str(p.pattern_id) in allowed]

    if family_ids:
        allowed = {str(x).strip() for x in family_ids if str(x).strip()}
        patterns = [p for p in patterns if str(p.family_id or "") in allowed]

    if tags_any:
        allowed_tags = {str(x).strip() for x in tags_any if str(x).strip()}
        patterns = [
            p for p in patterns
            if allowed_tags.intersection(set(p.tags or []))
        ]

    patterns = sorted(
        patterns,
        key=lambda p: (
            float(p.aesthetic_score) if p.aesthetic_score is not None else -1.0,
            float(p.print_score) if p.print_score is not None else -1.0,
            float(p.legibility_score) if p.legibility_score is not None else -1.0,
            str(p.pattern_id),
        ),
        reverse=True,
    )

    if max_patterns is not None:
        patterns = patterns[: int(max_patterns)]

    return patterns


def build_production_requests(
    *,
    patterns: Sequence[PatternRecord],
    count: int,
) -> List[PatternProductionRequest]:
    if count <= 0:
        raise ValueError("count must be >= 1")
    if not patterns:
        raise ValueError("No patterns available for production")

    requests: List[PatternProductionRequest] = []
    for idx, pattern in enumerate(islice(cycle(patterns), int(count)), start=1):
        request_id = f"REQ-{idx:05d}__{pattern.pattern_id}"
        requests.append(
            PatternProductionRequest(
                request_id=request_id,
                pattern=pattern,
            )
        )
    return requests


def write_pattern_requests_workbook(
    *,
    requests: Sequence[PatternProductionRequest],
    workbook_path: Path,
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Patterns"

    # Match the legacy list-input format expectations:
    # row 1 and row 2 are reserved / skipped by the legacy reader.
    sheet.cell(1, 1).value = "Pattern ID"
    for i in range(9):
        sheet.cell(1, 2 + i).value = f"Col {i + 1}"
    for i in range(9):
        sheet.cell(1, 11 + i).value = f"Layout {i + 1}"

    current_row = 3
    for request in requests:
        mask81 = str(request.pattern.mask81)
        rows = [mask81[i:i + 9] for i in range(0, 81, 9)]

        for row_bits in rows:
            sheet.cell(current_row, 1).value = request.request_id
            for c, bit in enumerate(row_bits, start=2):
                sheet.cell(current_row, c).value = 1 if bit == "1" else None
            # Leave custom layout columns blank for classic9x9.
            current_row += 1

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_path)
    return workbook_path


def _build_steering_args(
    *,
    weight_min: Optional[int],
    weight_max: Optional[int],
    technique_count_min: Optional[int],
    technique_count_max: Optional[int],
    required_techniques: Sequence[str],
    excluded_techniques: Sequence[str],
) -> List[str]:
    args: List[str] = []

    if technique_count_min is not None or technique_count_max is not None:
        if technique_count_min is None or technique_count_max is None:
            raise ValueError("Both technique_count_min and technique_count_max must be supplied together")
        args.extend(["-n", f"{int(technique_count_min)}-{int(technique_count_max)}"])

    if weight_min is not None or weight_max is not None:
        if weight_min is None or weight_max is None:
            raise ValueError("Both weight_min and weight_max must be supplied together")
        args.extend(["-w", f"{int(weight_min)}-{int(weight_max)}"])

    if required_techniques:
        args.extend(["-t", ",".join(str(x).strip() for x in required_techniques if str(x).strip())])

    if excluded_techniques:
        args.extend(["-e", ",".join(str(x).strip() for x in excluded_techniques if str(x).strip())])

    return args


def run_legacy_pattern_generator(
    *,
    generator_root: Path,
    input_workbook: Path,
    charset: str,
    weight_min: Optional[int] = None,
    weight_max: Optional[int] = None,
    technique_count_min: Optional[int] = None,
    technique_count_max: Optional[int] = None,
    required_techniques: Optional[Sequence[str]] = None,
    excluded_techniques: Optional[Sequence[str]] = None,
) -> Tuple[Path, str, str]:
    generator_root = Path(generator_root)
    main_script = generator_root / "tool_patterns" / "main.py"

    if not generator_root.exists():
        raise FileNotFoundError(f"Legacy generator root not found: {generator_root}")
    if not main_script.exists():
        raise FileNotFoundError(f"Legacy generator main.py not found: {main_script}")
    if not input_workbook.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_workbook}")

    existing_outputs = {p.resolve() for p in generator_root.glob("output_patterns_list_*.xlsx")}

    cmd = [
        sys.executable,
        "-u",
        str(main_script.relative_to(generator_root)),
        "9",
        str(charset),
        "--list",
        str(input_workbook.resolve()),
    ]
    cmd.extend(
        _build_steering_args(
            weight_min=weight_min,
            weight_max=weight_max,
            technique_count_min=technique_count_min,
            technique_count_max=technique_count_max,
            required_techniques=list(required_techniques or []),
            excluded_techniques=list(excluded_techniques or []),
        )
    )

    process = subprocess.Popen(
        cmd,
        cwd=str(generator_root),
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        bufsize=1,
    )

    assert process.stdin is not None
    assert process.stdout is not None

    # Auto-confirm the legacy prompt once, same behavior as before.
    process.stdin.write("y\n")
    process.stdin.flush()
    process.stdin.close()

    stdout_lines: List[str] = []

    for line in process.stdout:
        stdout_lines.append(line)
        print(line, end="", flush=True)

    return_code = process.wait()
    stdout_text = "".join(stdout_lines)
    stderr_text = ""

    current_outputs = {p.resolve() for p in generator_root.glob("output_patterns_list_*.xlsx")}
    new_outputs = [p for p in current_outputs if p not in existing_outputs]

    if return_code != 0:
        dependency_hint = ""

        if "No module named 'colorama'" in stdout_text:
            dependency_hint = (
                "\nLikely cause: the legacy generator subprocess is using a different Python interpreter\n"
                "than the current workflow environment. Ensure the subprocess is launched with sys.executable.\n"
                "If needed, also verify colorama is installed in that interpreter.\n"
            )

        raise RuntimeError(
            "Legacy pattern generator failed.\n"
            f"Return code: {return_code}\n"
            f"{dependency_hint}"
            f"COMBINED OUTPUT:\n{stdout_text}"
        )

    if not new_outputs:
        raise RuntimeError(
            "Legacy pattern generator finished but no new output_patterns_list_*.xlsx file was created.\n"
            f"COMBINED OUTPUT:\n{stdout_text}"
        )

    output_workbook = sorted(new_outputs, key=lambda p: p.stat().st_mtime, reverse=True)[0]
    return output_workbook, stdout_text, stderr_text


def _read_grid81(sheet, start_row: int, start_col: int, *, zero_for_blanks: bool) -> str:
    chars: List[str] = []
    for r in range(start_row, start_row + 9):
        for c in range(start_col, start_col + 9):
            value = sheet.cell(r, c).value
            if value is None or str(value).strip() == "":
                chars.append("0" if zero_for_blanks else "")
            else:
                chars.append(str(value).strip())
    return "".join(chars)


def _parse_bool(value: object) -> bool:
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def _parse_histogram(value: object) -> Dict[str, int]:
    if value is None:
        return {}

    raw = str(value).strip()
    if not raw:
        return {}

    parsed = ast.literal_eval(raw)
    if not isinstance(parsed, Mapping):
        return {}

    return {str(k): int(v) for k, v in parsed.items()}


def parse_legacy_output_workbook(
    *,
    output_workbook: Path,
    request_lookup: Mapping[str, PatternRecord],
) -> Tuple[List[dict], List[dict]]:
    workbook = load_workbook(filename=str(output_workbook), data_only=True)
    try:
        sheet = workbook.worksheets[0]

        candidates: List[dict] = []
        rejected: List[dict] = []

        row = 3
        max_row = int(sheet.max_row or 0)

        while row <= max_row:
            request_id = sheet.cell(row, 1).value
            if request_id is None or str(request_id).strip() == "":
                row += 1
                continue

            request_id = str(request_id).strip()
            pattern = request_lookup.get(request_id)
            if pattern is None:
                rejected.append(
                    {
                        "request_id": request_id,
                        "reason": "request_id_not_found_in_lookup",
                    }
                )
                row += 9
                continue

            no_instance_marker = sheet.cell(row, 12).value
            if str(no_instance_marker).strip() == "NO INSTANCE":
                rejected.append(
                    {
                        "request_id": request_id,
                        "pattern_id": pattern.pattern_id,
                        "pattern_name": pattern.name,
                        "reason": "no_instance_generated",
                    }
                )
                row += 9
                continue

            solution81 = _read_grid81(sheet, row, 12, zero_for_blanks=False)
            givens81 = _read_grid81(sheet, row, 32, zero_for_blanks=True)

            solvable_value = sheet.cell(row + 0, 43).value
            hints_value = sheet.cell(row + 1, 43).value
            techniques_value = sheet.cell(row + 2, 43).value
            weight_value = sheet.cell(row + 3, 43).value

            technique_histogram = _parse_histogram(techniques_value)
            techniques_used = sorted([name for name, count in technique_histogram.items() if int(count) > 0])

            candidate = {
                "givens81": givens81,
                "solution81": solution81,
                "weight": int(weight_value) if weight_value not in (None, "") else 0,
                "techniques_used": techniques_used,
                "technique_histogram": technique_histogram,
                "pattern_id": pattern.pattern_id,
                "pattern_name": pattern.name,
                "pattern_family_id": pattern.family_id,
                "pattern_family_name": pattern.family_name,
                "pattern_mask81": pattern.mask81,
                "generation_seed": None,
                "generator_version": LEGACY_GENERATOR_VERSION,
                "generation_method": "pattern_fill",
                "is_unique": True,
                "is_human_solvable": _parse_bool(solvable_value),
                "title": f"{pattern.name} / {request_id}",
                "subtitle": "Classic 9x9",
                "request_id": request_id,
                "hint_count": int(hints_value) if hints_value not in (None, "") else None,
            }
            candidates.append(candidate)

            row += 9

        return candidates, rejected
    finally:
        workbook.close()


def write_candidates_jsonl(
    *,
    candidates: Sequence[dict],
    output_jsonl: Path,
    append: bool = True,
) -> Path:
    output_jsonl.parent.mkdir(parents=True, exist_ok=True)
    mode = "a" if append else "w"
    with output_jsonl.open(mode, encoding="utf-8", newline="\n") as handle:
        for candidate in candidates:
            handle.write(json.dumps(candidate, ensure_ascii=False))
            handle.write("\n")
    return output_jsonl