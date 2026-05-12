from __future__ import annotations

import json
import shutil
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from python.publishing.step_solutions.locale_templates import (
    DEFAULT_SOLUTION_TEMPLATES_ROOT,
    message_template_path,
    normalize_step_solution_locale,
)
from python.publishing.step_solutions.localization_master import (
    DEFAULT_LOCALIZATION_MASTER_XLSX,
)
from python.publishing.step_solutions.template_localization_contract import (
    extract_placeholders,
)


DEFAULT_GENERATED_TEMPLATE_REPORT = (
    Path("datasets/sudoku_books/classic9/solution_templates/localization_master")
    / "generated_template_report.json"
)


def generate_step_solution_templates_from_master(
    master_xlsx: Path = DEFAULT_LOCALIZATION_MASTER_XLSX,
    templates_root: Path = DEFAULT_SOLUTION_TEMPLATES_ROOT,
    locales: Sequence[str] = ("fr", "de", "it", "es"),
    english_template_path: Optional[Path] = None,
    backup: bool = True,
    overwrite: bool = True,
    report_json: Path = DEFAULT_GENERATED_TEMPLATE_REPORT,
    fail_on_placeholder_mismatch: bool = True,
    generate_english_copy: bool = False,
) -> Dict[str, Any]:
    """
    Generate localized Template_Messages.xlsx files from the localization master.

    The English workbook remains the schema/format master. For each locale, this
    workflow copies the English workbook and writes translated cells into the
    same source row/column positions.

    This creates production files like:
        solution_templates/messages/fr/Template_Messages.xlsx
        solution_templates/messages/de/Template_Messages.xlsx
        solution_templates/messages/it/Template_Messages.xlsx
        solution_templates/messages/es/Template_Messages.xlsx
    """

    master_xlsx = Path(master_xlsx)
    templates_root = Path(templates_root)
    report_json = Path(report_json)

    if not master_xlsx.exists():
        raise FileNotFoundError(f"Localization master not found: {master_xlsx}")

    normalized_locales = [
        normalize_step_solution_locale(locale)
        for locale in locales
    ]

    if generate_english_copy and "en" not in normalized_locales:
        normalized_locales = ["en"] + normalized_locales

    english_path = (
        Path(english_template_path)
        if english_template_path
        else message_template_path("en", templates_root)
    )

    if not english_path.exists():
        raise FileNotFoundError(f"English template not found: {english_path}")

    master_wb = load_workbook(master_xlsx, data_only=False)

    generation_report: Dict[str, Any] = {
        "schema_version": "step_solution_template_generation_report.v1",
        "created_at": _now_iso(),
        "master_xlsx": str(master_xlsx),
        "templates_root": str(templates_root),
        "english_template_path": str(english_path),
        "locales": normalized_locales,
        "backup": backup,
        "overwrite": overwrite,
        "generated": [],
        "warnings": [],
        "errors": [],
    }

    for locale in normalized_locales:
        try:
            locale_report = _generate_one_locale_template(
                locale=locale,
                master_wb=master_wb,
                english_template_path=english_path,
                templates_root=templates_root,
                backup=backup,
                overwrite=overwrite,
                fail_on_placeholder_mismatch=fail_on_placeholder_mismatch,
                generate_english_copy=generate_english_copy,
            )
            generation_report["generated"].append(locale_report)
        except Exception as exc:
            generation_report["errors"].append(
                {
                    "locale": locale,
                    "message": str(exc),
                }
            )

    report_json.parent.mkdir(parents=True, exist_ok=True)
    report_json.write_text(
        json.dumps(generation_report, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    if generation_report["errors"]:
        raise RuntimeError(
            f"Template generation failed for {len(generation_report['errors'])} locale(s). "
            f"See report: {report_json}"
        )

    return generation_report


def _generate_one_locale_template(
    locale: str,
    master_wb,
    english_template_path: Path,
    templates_root: Path,
    backup: bool,
    overwrite: bool,
    fail_on_placeholder_mismatch: bool,
    generate_english_copy: bool,
) -> Dict[str, Any]:
    if locale == "en":
        if not generate_english_copy:
            return {
                "locale": "en",
                "status": "skipped",
                "reason": "English source template is not overwritten by default.",
            }
        output_path = message_template_path("en", templates_root).with_name(
            "Template_Messages.generated.xlsx"
        )
    else:
        output_path = message_template_path(locale, templates_root)

    if output_path.exists() and not overwrite:
        raise FileExistsError(
            f"Target template already exists: {output_path}. "
            "Pass --overwrite to replace it."
        )

    backup_path = None
    if backup and output_path.exists():
        backup_path = _backup_existing_template(output_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Copy English workbook first so sheet structure, styles, formulas, and
    # non-translatable machine keys remain exactly aligned to the legacy writer.
    shutil.copy2(english_template_path, output_path)

    wb = load_workbook(output_path)

    updates: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    _apply_headers_from_master(
        wb=wb,
        master_wb=master_wb,
        locale=locale,
        updates=updates,
        warnings=warnings,
        errors=errors,
        fail_on_placeholder_mismatch=fail_on_placeholder_mismatch,
    )
    _apply_names_from_master(
        wb=wb,
        master_wb=master_wb,
        locale=locale,
        updates=updates,
        warnings=warnings,
        errors=errors,
        fail_on_placeholder_mismatch=fail_on_placeholder_mismatch,
    )
    _apply_keywords_from_master(
        wb=wb,
        master_wb=master_wb,
        locale=locale,
        updates=updates,
        warnings=warnings,
        errors=errors,
        fail_on_placeholder_mismatch=fail_on_placeholder_mismatch,
    )
    _apply_messages_from_master(
        wb=wb,
        master_wb=master_wb,
        locale=locale,
        updates=updates,
        warnings=warnings,
        errors=errors,
        fail_on_placeholder_mismatch=fail_on_placeholder_mismatch,
    )

    if errors:
        raise ValueError(
            f"Template generation for locale {locale!r} has {len(errors)} error(s): "
            f"{errors[:3]}"
        )

    wb.save(output_path)

    return {
        "locale": locale,
        "status": "ok",
        "output_path": str(output_path),
        "backup_path": str(backup_path) if backup_path else None,
        "update_count": len(updates),
        "warning_count": len(warnings),
        "error_count": len(errors),
        "updates": updates,
        "warnings": warnings,
        "errors": errors,
    }


def _apply_headers_from_master(
    wb,
    master_wb,
    locale: str,
    updates: List[Dict[str, Any]],
    warnings: List[Dict[str, Any]],
    errors: List[Dict[str, Any]],
    fail_on_placeholder_mismatch: bool,
) -> None:
    _apply_direct_cell_sheet(
        wb=wb,
        master_wb=master_wb,
        master_sheet_name="Headers",
        locale=locale,
        updates=updates,
        warnings=warnings,
        errors=errors,
        fail_on_placeholder_mismatch=fail_on_placeholder_mismatch,
    )


def _apply_names_from_master(
    wb,
    master_wb,
    locale: str,
    updates: List[Dict[str, Any]],
    warnings: List[Dict[str, Any]],
    errors: List[Dict[str, Any]],
    fail_on_placeholder_mismatch: bool,
) -> None:
    _apply_direct_cell_sheet(
        wb=wb,
        master_wb=master_wb,
        master_sheet_name="Names",
        locale=locale,
        updates=updates,
        warnings=warnings,
        errors=errors,
        fail_on_placeholder_mismatch=fail_on_placeholder_mismatch,
    )


def _apply_keywords_from_master(
    wb,
    master_wb,
    locale: str,
    updates: List[Dict[str, Any]],
    warnings: List[Dict[str, Any]],
    errors: List[Dict[str, Any]],
    fail_on_placeholder_mismatch: bool,
) -> None:
    _apply_direct_cell_sheet(
        wb=wb,
        master_wb=master_wb,
        master_sheet_name="Keywords",
        locale=locale,
        updates=updates,
        warnings=warnings,
        errors=errors,
        fail_on_placeholder_mismatch=fail_on_placeholder_mismatch,
    )


def _apply_direct_cell_sheet(
    wb,
    master_wb,
    master_sheet_name: str,
    locale: str,
    updates: List[Dict[str, Any]],
    warnings: List[Dict[str, Any]],
    errors: List[Dict[str, Any]],
    fail_on_placeholder_mismatch: bool,
) -> None:
    if master_sheet_name not in master_wb.sheetnames:
        errors.append(
            {
                "sheet": master_sheet_name,
                "issue": "missing_master_sheet",
            }
        )
        return

    if master_sheet_name not in wb.sheetnames:
        errors.append(
            {
                "sheet": master_sheet_name,
                "issue": "missing_target_sheet",
            }
        )
        return

    ws_master = master_wb[master_sheet_name]
    ws_target = wb[master_sheet_name]
    header_map = _header_map(ws_master)

    if locale not in header_map:
        errors.append(
            {
                "sheet": master_sheet_name,
                "issue": "missing_locale_column",
                "locale": locale,
            }
        )
        return

    for row in range(2, ws_master.max_row + 1):
        source_row = _int_cell(ws_master, row, header_map, "source_row")
        source_column = _int_cell(ws_master, row, header_map, "source_column")
        en_value = _str_cell(ws_master, row, header_map, "en")
        locale_value = _str_cell(ws_master, row, header_map, locale)
        record_id = _str_cell(ws_master, row, header_map, "record_id")
        field_name = _str_cell(ws_master, row, header_map, "field")

        if not source_row or not source_column:
            continue

        # Legacy runtime compatibility:
        # In the Names sheet, the rating column is not display prose. It is a
        # machine key consumed by template_messages.py and must remain one of:
        #   Easy, Medium, Hard
        #
        # The localized difficulty label is resolved later through the Keywords
        # sheet/MAP_RATINGS path. If we write "Facile" or "Leicht" here, the
        # legacy reader crashes with:
        #   Rating for technique "singles-1" not recognised
        if master_sheet_name == "Names" and field_name == "rating":
            locale_value = en_value

        if not locale_value:
            warnings.append(
                {
                    "sheet": master_sheet_name,
                    "record_id": record_id,
                    "row": row,
                    "issue": "missing_locale_value",
                    "locale": locale,
                }
            )
            locale_value = en_value

        _validate_placeholders(
            sheet=master_sheet_name,
            record_id=record_id,
            locale=locale,
            en_value=en_value,
            locale_value=locale_value,
            errors=errors,
            fail_on_placeholder_mismatch=fail_on_placeholder_mismatch,
        )

        old_value = ws_target.cell(row=source_row, column=source_column).value
        if old_value != locale_value:
            ws_target.cell(row=source_row, column=source_column).value = locale_value
            updates.append(
                {
                    "sheet": master_sheet_name,
                    "record_id": record_id,
                    "source_row": source_row,
                    "source_column": source_column,
                    "locale": locale,
                    "old": old_value,
                    "new": locale_value,
                }
            )


def _apply_messages_from_master(
    wb,
    master_wb,
    locale: str,
    updates: List[Dict[str, Any]],
    warnings: List[Dict[str, Any]],
    errors: List[Dict[str, Any]],
    fail_on_placeholder_mismatch: bool,
) -> None:
    master_sheet_name = "Messages"

    if master_sheet_name not in master_wb.sheetnames:
        errors.append(
            {
                "sheet": master_sheet_name,
                "issue": "missing_master_sheet",
            }
        )
        return

    if master_sheet_name not in wb.sheetnames:
        errors.append(
            {
                "sheet": master_sheet_name,
                "issue": "missing_target_sheet",
            }
        )
        return

    ws_master = master_wb[master_sheet_name]
    ws_target = wb[master_sheet_name]
    header_map = _header_map(ws_master)

    if locale not in header_map:
        errors.append(
            {
                "sheet": master_sheet_name,
                "issue": "missing_locale_column",
                "locale": locale,
            }
        )
        return

    grouped: Dict[Tuple[int, int], List[Dict[str, Any]]] = defaultdict(list)

    for row in range(2, ws_master.max_row + 1):
        source_row = _int_cell(ws_master, row, header_map, "source_row")
        source_column = _int_cell(ws_master, row, header_map, "source_column")
        if not source_row or not source_column:
            continue

        record = {
            "master_row": row,
            "source_row": source_row,
            "source_column": source_column,
            "fragment_index": _int_cell(ws_master, row, header_map, "fragment_index") or 0,
            "record_id": _str_cell(ws_master, row, header_map, "record_id"),
            "en": _str_cell(ws_master, row, header_map, "en"),
            "locale_value": _str_cell(ws_master, row, header_map, locale),
        }
        grouped[(source_row, source_column)].append(record)

    for (source_row, source_column), records in grouped.items():
        records.sort(key=lambda record: record["fragment_index"])

        rendered_fragments = []
        for record in records:
            en_value = record["en"]
            locale_value = record["locale_value"]

            if not locale_value:
                warnings.append(
                    {
                        "sheet": master_sheet_name,
                        "record_id": record["record_id"],
                        "row": record["master_row"],
                        "issue": "missing_locale_value",
                        "locale": locale,
                    }
                )
                locale_value = en_value

            _validate_placeholders(
                sheet=master_sheet_name,
                record_id=record["record_id"],
                locale=locale,
                en_value=en_value,
                locale_value=locale_value,
                errors=errors,
                fail_on_placeholder_mismatch=fail_on_placeholder_mismatch,
            )

            rendered_fragments.append(_quote_message_fragment(locale_value))

        new_value = "\n".join(rendered_fragments)
        old_value = ws_target.cell(row=source_row, column=source_column).value

        if old_value != new_value:
            ws_target.cell(row=source_row, column=source_column).value = new_value
            updates.append(
                {
                    "sheet": master_sheet_name,
                    "source_row": source_row,
                    "source_column": source_column,
                    "locale": locale,
                    "fragment_count": len(records),
                    "old": old_value,
                    "new": new_value,
                }
            )


def _validate_placeholders(
    sheet: str,
    record_id: str,
    locale: str,
    en_value: str,
    locale_value: str,
    errors: List[Dict[str, Any]],
    fail_on_placeholder_mismatch: bool,
) -> None:
    en_placeholders = extract_placeholders(en_value)
    locale_placeholders = extract_placeholders(locale_value)

    if en_placeholders == locale_placeholders:
        return

    issue = {
        "sheet": sheet,
        "record_id": record_id,
        "locale": locale,
        "issue": "placeholder_mismatch",
        "missing": sorted(en_placeholders - locale_placeholders),
        "extra": sorted(locale_placeholders - en_placeholders),
        "en": en_value,
        "localized": locale_value,
    }

    if fail_on_placeholder_mismatch:
        errors.append(issue)


def _quote_message_fragment(value: str) -> str:
    """
    Legacy writer expects message fragments in quoted form.

    Single-fragment cell:
        "Sentence with {placeholder}. "

    Multi-fragment cell:
        "First fragment. "
        "Second fragment. "
    """

    text = str(value or "")
    escaped = text.replace('"', '""')
    return f'"{escaped}"'


def _header_map(ws) -> Dict[str, int]:
    result: Dict[str, int] = {}
    for cell in ws[1]:
        if cell.value:
            result[str(cell.value).strip()] = int(cell.column)
    return result


def _str_cell(ws, row: int, header_map: Dict[str, int], key: str) -> str:
    col = header_map.get(key)
    if not col:
        return ""
    value = ws.cell(row=row, column=col).value
    return "" if value is None else str(value)


def _int_cell(ws, row: int, header_map: Dict[str, int], key: str) -> int:
    value = _str_cell(ws, row, header_map, key)
    try:
        return int(value)
    except Exception:
        return 0


def _backup_existing_template(path: Path) -> Path:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup_dir = path.parents[2] / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)

    locale = path.parent.name
    backup_path = backup_dir / f"{locale}_Template_Messages.before_master_{timestamp}.xlsx"

    shutil.copy2(path, backup_path)
    return backup_path


def _now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()