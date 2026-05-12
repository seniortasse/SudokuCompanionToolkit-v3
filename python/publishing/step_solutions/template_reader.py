from __future__ import annotations

from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook

from python.publishing.step_solutions.template_localization_contract import (
    CANONICAL_TEMPLATE_SHEETS,
    extract_placeholders,
    normalize_template_text,
)


@dataclass(frozen=True)
class TemplateCellRecord:
    sheet: str
    row: int
    column: int
    value: str

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class TemplateMessageRecord:
    """
    One row from the Messages sheet.

    The workbook format has technique keys in leading columns and one or more
    message fragments in later columns.
    """

    row: int
    technique_keys: List[str]
    fragments: List[str]
    placeholders: List[str]

    def primary_key(self) -> str:
        return "|".join(self.technique_keys)

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class TemplateNameRecord:
    """
    One technique-name row from the Names sheet.
    """

    row: int
    technique_key: str
    name: str
    rating: str
    placeholders: List[str]

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class TemplateKeywordRecord:
    """
    One row from the Keywords sheet.

    The exact keyword sheet is intentionally treated flexibly because the
    legacy writer reads it by position/key. We preserve all non-empty row cells.
    """

    row: int
    key: str
    values: List[str]
    placeholders: List[str]

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class TemplateHeaderRecord:
    """
    One row from the Headers sheet.
    """

    row: int
    key: str
    values: List[str]
    placeholders: List[str]

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass
class StepSolutionTemplateWorkbook:
    path: Path
    locale: str
    sheet_names: List[str]

    headers: List[TemplateHeaderRecord] = field(default_factory=list)
    messages: List[TemplateMessageRecord] = field(default_factory=list)
    names: List[TemplateNameRecord] = field(default_factory=list)
    keywords: List[TemplateKeywordRecord] = field(default_factory=list)

    raw_cells: Dict[str, List[TemplateCellRecord]] = field(default_factory=dict)

    def message_by_primary_key(self) -> Dict[str, TemplateMessageRecord]:
        return {record.primary_key(): record for record in self.messages}

    def name_by_key(self) -> Dict[str, TemplateNameRecord]:
        return {record.technique_key: record for record in self.names}

    def keyword_by_key(self) -> Dict[str, TemplateKeywordRecord]:
        return {record.key: record for record in self.keywords}

    def header_by_key(self) -> Dict[str, TemplateHeaderRecord]:
        return {record.key: record for record in self.headers}

    def to_dict(self) -> Dict[str, Any]:
        return {
            "path": str(self.path),
            "locale": self.locale,
            "sheet_names": self.sheet_names,
            "headers": [record.to_dict() for record in self.headers],
            "messages": [record.to_dict() for record in self.messages],
            "names": [record.to_dict() for record in self.names],
            "keywords": [record.to_dict() for record in self.keywords],
        }


def read_step_solution_template_workbook(
    path: Path,
    locale: str,
    data_only: bool = True,
) -> StepSolutionTemplateWorkbook:
    """
    Read a Template_Messages.xlsx workbook into a normalized structure.

    This reader is intentionally tolerant. It does not assume perfect formatting;
    it captures the workbook as-is so the auditor can report issues.
    """

    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Template workbook not found: {path}")

    wb = load_workbook(path, data_only=data_only)
    workbook = StepSolutionTemplateWorkbook(
        path=path,
        locale=locale,
        sheet_names=list(wb.sheetnames),
    )

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        workbook.raw_cells[sheet_name] = _read_raw_cells(sheet_name, sheet)

    if "Headers" in wb.sheetnames:
        workbook.headers = _read_headers_sheet(wb["Headers"])

    if "Messages" in wb.sheetnames:
        workbook.messages = _read_messages_sheet(wb["Messages"])

    if "Names" in wb.sheetnames:
        workbook.names = _read_names_sheet(wb["Names"])

    if "Keywords" in wb.sheetnames:
        workbook.keywords = _read_keywords_sheet(wb["Keywords"])

    return workbook


def _read_raw_cells(sheet_name: str, sheet) -> List[TemplateCellRecord]:
    records: List[TemplateCellRecord] = []
    for row in sheet.iter_rows():
        for cell in row:
            value = normalize_template_text(cell.value)
            if value:
                records.append(
                    TemplateCellRecord(
                        sheet=sheet_name,
                        row=int(cell.row),
                        column=int(cell.column),
                        value=value,
                    )
                )
    return records


def _read_headers_sheet(sheet) -> List[TemplateHeaderRecord]:
    records: List[TemplateHeaderRecord] = []

    for row in sheet.iter_rows():
        values = [_cell_text(cell.value) for cell in row]
        non_empty = [value for value in values if value]

        if not non_empty:
            continue

        # Use the first non-empty cell as the structural key.
        key = non_empty[0]
        placeholders = sorted(
            set().union(*(extract_placeholders(value) for value in non_empty))
        )

        records.append(
            TemplateHeaderRecord(
                row=int(row[0].row),
                key=key,
                values=non_empty,
                placeholders=placeholders,
            )
        )

    return records


def _read_messages_sheet(sheet) -> List[TemplateMessageRecord]:
    records: List[TemplateMessageRecord] = []

    for row in sheet.iter_rows():
        values = [_cell_text(cell.value) for cell in row]
        if not any(values):
            continue

        technique_keys = []
        fragments = []

        for value in values:
            if not value:
                continue

            if _looks_like_technique_key(value):
                technique_keys.append(value)
            else:
                fragments.append(value)

        if not technique_keys and not fragments:
            continue

        # Skip visual/comment rows that are not real message definitions.
        if not technique_keys:
            continue

        placeholders = sorted(
            set().union(*(extract_placeholders(fragment) for fragment in fragments))
        )

        records.append(
            TemplateMessageRecord(
                row=int(row[0].row),
                technique_keys=technique_keys,
                fragments=fragments,
                placeholders=placeholders,
            )
        )

    return records


def _read_names_sheet(sheet) -> List[TemplateNameRecord]:
    records: List[TemplateNameRecord] = []

    for row in sheet.iter_rows():
        values = [_cell_text(cell.value) for cell in row]
        non_empty = [value for value in values if value]
        if len(non_empty) < 2:
            continue

        technique_key = non_empty[0]
        if not _looks_like_technique_key(technique_key):
            continue

        name = non_empty[1] if len(non_empty) >= 2 else ""
        rating = non_empty[2] if len(non_empty) >= 3 else ""

        placeholders = sorted(
            set().union(*(extract_placeholders(value) for value in non_empty))
        )

        records.append(
            TemplateNameRecord(
                row=int(row[0].row),
                technique_key=technique_key,
                name=name,
                rating=rating,
                placeholders=placeholders,
            )
        )

    return records


def _read_keywords_sheet(sheet) -> List[TemplateKeywordRecord]:
    records: List[TemplateKeywordRecord] = []

    for row in sheet.iter_rows():
        values = [_cell_text(cell.value) for cell in row]
        non_empty = [value for value in values if value]
        if not non_empty:
            continue

        key = non_empty[0]
        values_tail = non_empty[1:]

        placeholders = sorted(
            set().union(*(extract_placeholders(value) for value in non_empty))
        )

        records.append(
            TemplateKeywordRecord(
                row=int(row[0].row),
                key=key,
                values=values_tail,
                placeholders=placeholders,
            )
        )

    return records


def _cell_text(value: object) -> str:
    return normalize_template_text(value)


def _looks_like_technique_key(value: str) -> bool:
    """
    Technique keys are stable English-like machine keys, for example:
        singles-1
        x-wings-3
        doubles-naked
        leftovers-9
    """

    text = str(value or "").strip()
    if not text:
        return False

    if " " in text:
        return False

    if "{" in text or "}" in text:
        return False

    # Most technique keys contain hyphens. Allow a few simple lowercase keys too.
    allowed = set("abcdefghijklmnopqrstuvwxyz0123456789-_")
    return text.lower() == text and all(char in allowed for char in text)