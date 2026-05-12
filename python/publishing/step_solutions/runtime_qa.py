from __future__ import annotations

import csv
import json
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import load_workbook

from python.publishing.step_solutions.locale_templates import (
    DEFAULT_SOLUTION_TEMPLATES_ROOT,
    message_template_path,
    normalize_step_solution_locale,
)
from python.publishing.step_solutions.models import StepSolutionPackageRequest
from python.publishing.step_solutions.paths import (
    resolve_package_paths,
)


DEFAULT_STEP_SOLUTION_PACKAGES_ROOT = Path(
    "datasets/sudoku_books/classic9/step_solution_packages"
)
DEFAULT_BOOKS_ROOT = Path("datasets/sudoku_books/classic9/books")


LOCALIZED_STEP_HEADER_PREFIXES = {
    "en": ("STEP",),
    "fr": ("ÉTAPE", "ETAPE"),
    "de": ("SCHRITT",),
    "it": ("PASSO",),
    "es": ("PASO",),
}


# Strong English leakage signals. These are intentionally phrase-heavy to avoid
# false positives from valid words shared across languages.
ENGLISH_RUNTIME_LEAKAGE_PHRASES = (
    "Looking at",
    "There is a single position",
    "The other cells",
    "cannot contain",
    "because their",
    "Therefore",
    "can be removed",
    "is the only missing value",
    "remains the only candidate",
    "single position left",
)


@dataclass
class RuntimeQaIssue:
    severity: str
    issue_type: str
    message: str
    path: Optional[str] = None
    details: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass
class RuntimeQaPackageReport:
    book_id: str
    locale: str
    package_id: str
    package_root: Path

    status: str = "planned"
    csv_row_count: int = 0
    manifest_asset_count: int = 0
    user_log_count: int = 0
    answer_image_count: int = 0
    step_image_count: int = 0

    checked_workbooks: int = 0
    checked_csv_explanations: int = 0

    issues: List[RuntimeQaIssue] = field(default_factory=list)

    def add_issue(
        self,
        severity: str,
        issue_type: str,
        message: str,
        path: Optional[Path | str] = None,
        details: Optional[Dict[str, Any]] = None,
    ) -> None:
        self.issues.append(
            RuntimeQaIssue(
                severity=severity,
                issue_type=issue_type,
                message=message,
                path=str(path) if path is not None else None,
                details=details or {},
            )
        )

    @property
    def error_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == "error")

    @property
    def warning_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == "warning")

    def finalize(self) -> None:
        self.status = "ok" if self.error_count == 0 else "failed"

    def to_dict(self) -> Dict[str, Any]:
        return {
            "book_id": self.book_id,
            "locale": self.locale,
            "package_id": self.package_id,
            "package_root": str(self.package_root),
            "status": self.status,
            "csv_row_count": self.csv_row_count,
            "manifest_asset_count": self.manifest_asset_count,
            "user_log_count": self.user_log_count,
            "answer_image_count": self.answer_image_count,
            "step_image_count": self.step_image_count,
            "checked_workbooks": self.checked_workbooks,
            "checked_csv_explanations": self.checked_csv_explanations,
            "error_count": self.error_count,
            "warning_count": self.warning_count,
            "issues": [issue.to_dict() for issue in self.issues],
        }


def qa_step_solution_runtime_package(
    book_id: str,
    locale: str,
    output_root: Path = DEFAULT_STEP_SOLUTION_PACKAGES_ROOT,
    books_root: Path = DEFAULT_BOOKS_ROOT,
    templates_root: Path = DEFAULT_SOLUTION_TEMPLATES_ROOT,
    max_workbooks_to_check: int = 3,
    max_csv_rows_to_check: int = 25,
    require_localized_text: bool = True,
) -> RuntimeQaPackageReport:
    """
    Inspect a generated step-solution package after runtime export.

    This verifies package structure and samples generated content. It is meant
    for L6 QA after running export_step_solution_package.
    """

    locale = normalize_step_solution_locale(locale)

    request = StepSolutionPackageRequest(
        book_id=book_id,
        locale=locale,
        output_root=output_root,
        books_root=books_root,
    )
    paths = resolve_package_paths(request)

    report = RuntimeQaPackageReport(
        book_id=book_id,
        locale=locale,
        package_id=request.package_id(),
        package_root=paths.package_root,
    )

    _check_required_package_files(report, paths)
    _check_manifest(report, paths)
    _check_image_files(report, paths)
    _check_csv(report, paths, locale, max_csv_rows_to_check, require_localized_text)
    _check_generated_workbooks(
        report=report,
        user_logs_dir=paths.user_logs_dir,
        locale=locale,
        templates_root=templates_root,
        max_workbooks_to_check=max_workbooks_to_check,
        require_localized_text=require_localized_text,
    )

    report.finalize()
    return report


def qa_multiple_step_solution_runtime_packages(
    book_id: str,
    locales: Sequence[str],
    output_root: Path = DEFAULT_STEP_SOLUTION_PACKAGES_ROOT,
    books_root: Path = DEFAULT_BOOKS_ROOT,
    templates_root: Path = DEFAULT_SOLUTION_TEMPLATES_ROOT,
    max_workbooks_to_check: int = 3,
    max_csv_rows_to_check: int = 25,
    require_localized_text: bool = True,
) -> Dict[str, Any]:
    reports = [
        qa_step_solution_runtime_package(
            book_id=book_id,
            locale=locale,
            output_root=output_root,
            books_root=books_root,
            templates_root=templates_root,
            max_workbooks_to_check=max_workbooks_to_check,
            max_csv_rows_to_check=max_csv_rows_to_check,
            require_localized_text=require_localized_text,
        )
        for locale in locales
    ]

    return {
        "schema_version": "step_solution_runtime_qa_report.v1",
        "book_id": book_id,
        "locales": [normalize_step_solution_locale(locale) for locale in locales],
        "package_count": len(reports),
        "ok": sum(1 for report in reports if report.status == "ok"),
        "failed": sum(1 for report in reports if report.status == "failed"),
        "error_count": sum(report.error_count for report in reports),
        "warning_count": sum(report.warning_count for report in reports),
        "packages": [report.to_dict() for report in reports],
    }


def write_runtime_qa_report(
    payload: Dict[str, Any],
    output_path: Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    return output_path


def _check_required_package_files(report: RuntimeQaPackageReport, paths) -> None:
    required_files = [
        paths.manifest_json_path,
        paths.sudoku_index_csv_path,
        paths.reports_dir / "package_export_report.json",
        paths.reports_dir / "image_export_report.json",
        paths.reports_dir / "csv_export_report.json",
    ]

    if not paths.package_root.exists():
        report.add_issue(
            severity="error",
            issue_type="missing_package_root",
            message="Package root does not exist.",
            path=paths.package_root,
        )
        return

    for path in required_files:
        if not path.exists():
            report.add_issue(
                severity="error",
                issue_type="missing_required_file",
                message="Required package file is missing.",
                path=path,
            )

    for folder in (paths.user_logs_dir, paths.image_files_dir, paths.reports_dir):
        if not folder.exists():
            report.add_issue(
                severity="error",
                issue_type="missing_required_folder",
                message="Required package folder is missing.",
                path=folder,
            )


def _check_manifest(report: RuntimeQaPackageReport, paths) -> None:
    path = paths.manifest_json_path
    if not path.exists():
        return

    try:
        manifest = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        report.add_issue(
            severity="error",
            issue_type="manifest_read_failed",
            message=str(exc),
            path=path,
        )
        return

    assets = list(manifest.get("assets") or [])
    report.manifest_asset_count = len(assets)

    if not assets:
        report.add_issue(
            severity="warning",
            issue_type="manifest_has_no_assets",
            message="Manifest has no assets.",
            path=path,
        )

    for asset in assets:
        status = str(asset.get("status") or "")
        if status and status not in {"ok", "skipped_existing"}:
            report.add_issue(
                severity="warning",
                issue_type="manifest_asset_not_ok",
                message="Manifest asset is not ok.",
                path=path,
                details={
                    "internal_puzzle_code": asset.get("internal_puzzle_code"),
                    "external_puzzle_code": asset.get("external_puzzle_code"),
                    "status": status,
                    "errors": asset.get("errors"),
                },
            )


def _check_image_files(report: RuntimeQaPackageReport, paths) -> None:
    if not paths.user_logs_dir.exists():
        return
    if not paths.image_files_dir.exists():
        return

    user_logs = sorted(paths.user_logs_dir.glob("*_user_logs.xlsx"))
    answers = sorted(paths.image_files_dir.glob("*_answer.png"))
    steps = sorted(paths.image_files_dir.glob("*_step*.png"))

    report.user_log_count = len(user_logs)
    report.answer_image_count = len(answers)
    report.step_image_count = len(steps)

    if not user_logs:
        report.add_issue(
            severity="error",
            issue_type="no_user_logs",
            message="No user log workbooks found.",
            path=paths.user_logs_dir,
        )

    if not answers:
        report.add_issue(
            severity="error",
            issue_type="no_answer_images",
            message="No answer images found.",
            path=paths.image_files_dir,
        )

    if not steps:
        report.add_issue(
            severity="error",
            issue_type="no_step_images",
            message="No step images found.",
            path=paths.image_files_dir,
        )

    if user_logs and answers and len(answers) < len(user_logs):
        report.add_issue(
            severity="warning",
            issue_type="fewer_answers_than_logs",
            message="There are fewer answer images than user log workbooks.",
            details={
                "user_logs": len(user_logs),
                "answers": len(answers),
            },
        )


def _check_csv(
    report: RuntimeQaPackageReport,
    paths,
    locale: str,
    max_csv_rows_to_check: int,
    require_localized_text: bool,
) -> None:
    path = paths.sudoku_index_csv_path
    if not path.exists():
        return

    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)
    except Exception as exc:
        report.add_issue(
            severity="error",
            issue_type="csv_read_failed",
            message=str(exc),
            path=path,
        )
        return

    report.csv_row_count = len(rows)

    if not rows:
        report.add_issue(
            severity="error",
            issue_type="csv_has_no_rows",
            message="sudokuIndexFile.csv has no data rows.",
            path=path,
        )
        return

    required_headers = [
        "Problem ID",
        "Problem Name",
        "Book",
        "Level",
        "Answer",
        "Step 1",
        "Explanation 1",
    ]
    headers = list(rows[0].keys())
    missing_headers = [header for header in required_headers if header not in headers]
    if missing_headers:
        report.add_issue(
            severity="error",
            issue_type="csv_missing_required_headers",
            message="sudokuIndexFile.csv is missing required headers.",
            path=path,
            details={"missing_headers": missing_headers},
        )

    for row_index, row in enumerate(rows[:max_csv_rows_to_check], start=1):
        _check_csv_row_image_paths(report, paths, path, row_index, row)
        _check_csv_row_explanations(
            report=report,
            csv_path=path,
            row_index=row_index,
            row=row,
            locale=locale,
            require_localized_text=require_localized_text,
        )


def _check_csv_row_image_paths(
    report: RuntimeQaPackageReport,
    paths,
    csv_path: Path,
    row_index: int,
    row: Dict[str, str],
) -> None:
    image_fields = ["Answer"]
    image_fields.extend(
        [f"Step {step_number}" for step_number in range(1, 41)]
    )

    for field in image_fields:
        value = str(row.get(field) or "").strip()
        if not value:
            continue

        candidate = paths.package_root / value
        if not candidate.exists():
            report.add_issue(
                severity="error",
                issue_type="csv_image_path_missing",
                message=f"CSV image path does not exist for field {field}.",
                path=csv_path,
                details={
                    "row_index": row_index,
                    "field": field,
                    "value": value,
                    "resolved_path": str(candidate),
                },
            )


def _check_csv_row_explanations(
    report: RuntimeQaPackageReport,
    csv_path: Path,
    row_index: int,
    row: Dict[str, str],
    locale: str,
    require_localized_text: bool,
) -> None:
    explanation_fields = [
        f"Explanation {step_number}" for step_number in range(1, 41)
    ]

    for field in explanation_fields:
        value = str(row.get(field) or "").strip()
        if not value:
            continue

        report.checked_csv_explanations += 1

        if locale != "en":
            leaks = _find_runtime_english_leakage(value)
            if leaks:
                severity = "error" if require_localized_text else "warning"
                report.add_issue(
                    severity=severity,
                    issue_type="csv_explanation_english_leakage",
                    message="CSV explanation appears to contain English narration.",
                    path=csv_path,
                    details={
                        "row_index": row_index,
                        "field": field,
                        "leakage_phrases": leaks,
                        "value_excerpt": value[:500],
                    },
                )


def _check_generated_workbooks(
    report: RuntimeQaPackageReport,
    user_logs_dir: Path,
    locale: str,
    templates_root: Path,
    max_workbooks_to_check: int,
    require_localized_text: bool,
) -> None:
    if not user_logs_dir.exists():
        return

    workbooks = sorted(user_logs_dir.glob("*_user_logs.xlsx"))[:max_workbooks_to_check]
    if not workbooks:
        return

    expected_step_prefixes = LOCALIZED_STEP_HEADER_PREFIXES.get(locale, ())
    template_header_sample = _read_template_header_sample(locale, templates_root)

    for workbook_path in workbooks:
        _check_one_generated_workbook(
            report=report,
            workbook_path=workbook_path,
            locale=locale,
            expected_step_prefixes=expected_step_prefixes,
            template_header_sample=template_header_sample,
            require_localized_text=require_localized_text,
        )


def _check_one_generated_workbook(
    report: RuntimeQaPackageReport,
    workbook_path: Path,
    locale: str,
    expected_step_prefixes: Sequence[str],
    template_header_sample: str,
    require_localized_text: bool,
) -> None:
    try:
        wb = load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception as exc:
        report.add_issue(
            severity="error",
            issue_type="workbook_read_failed",
            message=str(exc),
            path=workbook_path,
        )
        return

    report.checked_workbooks += 1

    if "Steps" not in wb.sheetnames:
        report.add_issue(
            severity="error",
            issue_type="workbook_missing_steps_sheet",
            message="Generated workbook has no Steps sheet.",
            path=workbook_path,
        )
        return

    ws = wb["Steps"]

    step_header_hits = 0
    english_step_header_hits = 0
    english_leakage_hits: List[Dict[str, Any]] = []

    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value).strip()
            if not value:
                continue

            upper_value = value.upper()

            if expected_step_prefixes and any(
                upper_value.startswith(prefix)
                for prefix in expected_step_prefixes
            ):
                step_header_hits += 1

            if locale != "en" and upper_value.startswith("STEP "):
                english_step_header_hits += 1

            if locale != "en":
                leaks = _find_runtime_english_leakage(value)
                if leaks:
                    english_leakage_hits.append(
                        {
                            "cell": cell.coordinate,
                            "leakage_phrases": leaks,
                            "value_excerpt": value[:300],
                        }
                    )

    if expected_step_prefixes and step_header_hits == 0:
        report.add_issue(
            severity="error",
            issue_type="workbook_no_localized_step_headers",
            message="No localized step headers were found in the generated workbook.",
            path=workbook_path,
            details={
                "expected_prefixes": list(expected_step_prefixes),
                "template_header_sample": template_header_sample,
            },
        )

    if english_step_header_hits:
        report.add_issue(
            severity="error" if require_localized_text else "warning",
            issue_type="workbook_english_step_headers",
            message="English STEP headers were found in a localized workbook.",
            path=workbook_path,
            details={"count": english_step_header_hits},
        )

    if english_leakage_hits:
        report.add_issue(
            severity="error" if require_localized_text else "warning",
            issue_type="workbook_english_narration_leakage",
            message="Possible English narration remains in generated workbook.",
            path=workbook_path,
            details={
                "hit_count": len(english_leakage_hits),
                "sample_hits": english_leakage_hits[:10],
            },
        )


def _read_template_header_sample(locale: str, templates_root: Path) -> str:
    try:
        path = message_template_path(locale, templates_root)
        if not path.exists():
            return ""

        wb = load_workbook(path, data_only=True, read_only=True)
        if "Headers" not in wb.sheetnames:
            return ""

        ws = wb["Headers"]
        values = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    values.append(str(cell.value))
        return " | ".join(values[:5])
    except Exception:
        return ""


def _find_runtime_english_leakage(text: str) -> List[str]:
    value = str(text or "")
    if not value:
        return []

    hits: List[str] = []
    for phrase in ENGLISH_RUNTIME_LEAKAGE_PHRASES:
        if re.search(re.escape(phrase), value, flags=re.IGNORECASE):
            hits.append(phrase)

    return sorted(set(hits))