from __future__ import annotations

import json
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import load_workbook

from python.publishing.step_solutions.localization_master import (
    DEFAULT_LOCALIZATION_MASTER_XLSX,
)
from python.publishing.step_solutions.localization_translation_seed import (
    HEADER_TRANSLATIONS_BY_EN,
    KEYWORD_TRANSLATIONS_BY_EN,
    MESSAGE_TRANSLATIONS,
    RATING_TRANSLATIONS,
    TECHNIQUE_NAME_TRANSLATIONS,
)
from python.publishing.step_solutions.template_localization_contract import (
    extract_placeholders,
)


DEFAULT_APPLY_SEED_REPORT = (
    Path("datasets/sudoku_books/classic9/solution_templates/localization_master")
    / "step_solution_localization_master.apply_seed_report.json"
)


def apply_localization_seed_to_master(
    master_xlsx: Path = DEFAULT_LOCALIZATION_MASTER_XLSX,
    report_json: Path = DEFAULT_APPLY_SEED_REPORT,
    locales: Sequence[str] = ("fr", "de", "it", "es"),
    backup: bool = True,
    fail_on_placeholder_mismatch: bool = True,
) -> Dict[str, Any]:
    """
    Apply curated translation seeds to the localization master workbook.

    This updates only the master workbook. It does not generate localized
    Template_Messages.xlsx files yet.
    """

    master_xlsx = Path(master_xlsx)
    report_json = Path(report_json)

    if not master_xlsx.exists():
        raise FileNotFoundError(f"Localization master not found: {master_xlsx}")

    backup_path: Optional[Path] = None
    if backup:
        backup_path = _backup_master(master_xlsx)

    wb = load_workbook(master_xlsx)
    updates: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    for sheet_name in ("Headers", "Messages", "Names", "Keywords"):
        if sheet_name not in wb.sheetnames:
            errors.append(
                {
                    "sheet": sheet_name,
                    "issue": "missing_sheet",
                    "message": f"Master workbook is missing sheet {sheet_name!r}.",
                }
            )
            continue

        ws = wb[sheet_name]
        header_map = _header_map(ws)

        for required in ("en", "fr", "de", "it", "es"):
            if required not in header_map:
                errors.append(
                    {
                        "sheet": sheet_name,
                        "issue": "missing_locale_column",
                        "message": f"Sheet {sheet_name!r} is missing column {required!r}.",
                    }
                )

        if sheet_name == "Messages":
            _apply_message_seed(ws, header_map, locales, updates, warnings, errors)
        elif sheet_name == "Names":
            _apply_name_seed(ws, header_map, locales, updates, warnings, errors)
        elif sheet_name == "Keywords":
            _apply_keyword_seed(ws, header_map, locales, updates, warnings, errors)
        elif sheet_name == "Headers":
            _apply_header_seed(ws, header_map, locales, updates, warnings, errors)

    if fail_on_placeholder_mismatch:
        placeholder_errors = [
            error for error in errors if error.get("issue") == "placeholder_mismatch"
        ]
        if placeholder_errors:
            _write_report(
                report_json=report_json,
                master_xlsx=master_xlsx,
                backup_path=backup_path,
                updates=updates,
                warnings=warnings,
                errors=errors,
            )
            raise ValueError(
                f"Refusing to save because {len(placeholder_errors)} placeholder mismatch(es) were found. "
                f"See report: {report_json}"
            )

    wb.save(master_xlsx)

    return _write_report(
        report_json=report_json,
        master_xlsx=master_xlsx,
        backup_path=backup_path,
        updates=updates,
        warnings=warnings,
        errors=errors,
    )


def _apply_message_seed(ws, header_map, locales, updates, warnings, errors) -> None:
    for row in range(2, ws.max_row + 1):
        key = _cell(ws, row, header_map, "key")
        fragment_index = _cell(ws, row, header_map, "fragment_index")
        en = _cell(ws, row, header_map, "en")

        try:
            fragment_index_int = int(fragment_index or 0)
        except Exception:
            fragment_index_int = 0

        # Safety guard: some older master builds may still contain template
        # maintainer comments such as "Remark: ...". Those are not solver
        # narration messages and must never receive message translations.
        if en.strip().lower().startswith("remark:"):
            continue

        translations = MESSAGE_TRANSLATIONS.get((key, fragment_index_int))
        if not translations:
            continue

        _apply_locale_values(
            ws=ws,
            row=row,
            header_map=header_map,
            locales=locales,
            translations=translations,
            en=en,
            source="MESSAGE_TRANSLATIONS",
            updates=updates,
            warnings=warnings,
            errors=errors,
        )


def _apply_name_seed(ws, header_map, locales, updates, warnings, errors) -> None:
    for row in range(2, ws.max_row + 1):
        technique_key = _cell(ws, row, header_map, "key")
        field = _cell(ws, row, header_map, "field")
        en = _cell(ws, row, header_map, "en")

        translations = None
        source = ""

        if field == "technique_name":
            translations = TECHNIQUE_NAME_TRANSLATIONS.get(technique_key)
            source = "TECHNIQUE_NAME_TRANSLATIONS"
        elif field == "rating":
            translations = RATING_TRANSLATIONS.get(en)
            source = "RATING_TRANSLATIONS"

        if not translations:
            continue

        _apply_locale_values(
            ws=ws,
            row=row,
            header_map=header_map,
            locales=locales,
            translations=translations,
            en=en,
            source=source,
            updates=updates,
            warnings=warnings,
            errors=errors,
        )


def _apply_keyword_seed(ws, header_map, locales, updates, warnings, errors) -> None:
    for row in range(2, ws.max_row + 1):
        en = _cell(ws, row, header_map, "en")
        translations = KEYWORD_TRANSLATIONS_BY_EN.get(en)
        if not translations:
            continue

        _apply_locale_values(
            ws=ws,
            row=row,
            header_map=header_map,
            locales=locales,
            translations=translations,
            en=en,
            source="KEYWORD_TRANSLATIONS_BY_EN",
            updates=updates,
            warnings=warnings,
            errors=errors,
        )


def _apply_header_seed(ws, header_map, locales, updates, warnings, errors) -> None:
    for row in range(2, ws.max_row + 1):
        en = _cell(ws, row, header_map, "en")
        translations = HEADER_TRANSLATIONS_BY_EN.get(en)
        if not translations:
            continue

        _apply_locale_values(
            ws=ws,
            row=row,
            header_map=header_map,
            locales=locales,
            translations=translations,
            en=en,
            source="HEADER_TRANSLATIONS_BY_EN",
            updates=updates,
            warnings=warnings,
            errors=errors,
        )


def _apply_locale_values(
    ws,
    row: int,
    header_map: Dict[str, int],
    locales: Sequence[str],
    translations: Dict[str, str],
    en: str,
    source: str,
    updates: List[Dict[str, Any]],
    warnings: List[Dict[str, Any]],
    errors: List[Dict[str, Any]],
) -> None:
    en_placeholders = extract_placeholders(en)

    for locale in locales:
        if locale not in header_map:
            continue

        translated = translations.get(locale)
        if translated is None:
            warnings.append(
                {
                    "sheet": ws.title,
                    "row": row,
                    "locale": locale,
                    "issue": "missing_seed_locale",
                    "source": source,
                }
            )
            continue

        translated_placeholders = extract_placeholders(translated)
        if en_placeholders != translated_placeholders:
            errors.append(
                {
                    "sheet": ws.title,
                    "row": row,
                    "locale": locale,
                    "issue": "placeholder_mismatch",
                    "source": source,
                    "en": en,
                    "translated": translated,
                    "missing": sorted(en_placeholders - translated_placeholders),
                    "extra": sorted(translated_placeholders - en_placeholders),
                }
            )
            continue

        col = header_map[locale]
        old_value = ws.cell(row=row, column=col).value
        if old_value != translated:
            ws.cell(row=row, column=col).value = translated
            updates.append(
                {
                    "sheet": ws.title,
                    "row": row,
                    "locale": locale,
                    "source": source,
                    "old": old_value,
                    "new": translated,
                }
            )


def _header_map(ws) -> Dict[str, int]:
    result: Dict[str, int] = {}
    for cell in ws[1]:
        if cell.value:
            result[str(cell.value).strip()] = int(cell.column)
    return result


def _cell(ws, row: int, header_map: Dict[str, int], key: str) -> str:
    col = header_map.get(key)
    if not col:
        return ""
    value = ws.cell(row=row, column=col).value
    return "" if value is None else str(value)


def _backup_master(master_xlsx: Path) -> Path:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup_dir = master_xlsx.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"{master_xlsx.stem}.before_seed_{timestamp}{master_xlsx.suffix}"
    shutil.copy2(master_xlsx, backup_path)
    return backup_path


def _write_report(
    report_json: Path,
    master_xlsx: Path,
    backup_path: Optional[Path],
    updates: List[Dict[str, Any]],
    warnings: List[Dict[str, Any]],
    errors: List[Dict[str, Any]],
) -> Dict[str, Any]:
    report = {
        "schema_version": "step_solution_localization_seed_apply_report.v1",
        "created_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "master_xlsx": str(master_xlsx),
        "backup_path": str(backup_path) if backup_path else None,
        "update_count": len(updates),
        "warning_count": len(warnings),
        "error_count": len(errors),
        "updates": updates,
        "warnings": warnings,
        "errors": errors,
    }

    report_json.parent.mkdir(parents=True, exist_ok=True)
    report_json.write_text(
        json.dumps(report, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    return report