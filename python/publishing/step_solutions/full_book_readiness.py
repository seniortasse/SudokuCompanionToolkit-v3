from __future__ import annotations

import json
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import load_workbook

# Intentionally do not import book_loader helper names here.
# The step-solution book loader has changed names across patch phases.
# This readiness checker reads book_manifest.json directly so it remains stable.

from python.publishing.step_solutions.locale_templates import (
    DEFAULT_SOLUTION_TEMPLATES_ROOT,
    message_template_path,
    normalize_step_solution_locale,
)
from python.publishing.step_solutions.models import StepSolutionPackageRequest
from python.publishing.step_solutions.paths import resolve_package_paths


DEFAULT_BOOKS_ROOT = Path("datasets/sudoku_books/classic9/books")
DEFAULT_PACKAGES_ROOT = Path("datasets/sudoku_books/classic9/step_solution_packages")


@dataclass
class FullBookReadinessIssue:
    severity: str
    issue_type: str
    message: str
    path: Optional[str] = None
    details: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass
class FullBookReadinessLocaleReport:
    book_id: str
    locale: str
    package_id: str
    package_root: Path
    template_path: Path

    template_exists: bool = False
    package_exists: bool = False
    existing_user_logs: int = 0
    existing_answer_images: int = 0
    existing_step_images: int = 0
    existing_csv_rows: int = 0

    issues: List[FullBookReadinessIssue] = field(default_factory=list)

    def add_issue(
        self,
        severity: str,
        issue_type: str,
        message: str,
        path: Optional[Path | str] = None,
        details: Optional[Dict[str, Any]] = None,
    ) -> None:
        self.issues.append(
            FullBookReadinessIssue(
                severity=severity,
                issue_type=issue_type,
                message=message,
                path=str(path) if path is not None else None,
                details=details or {},
            )
        )

    @property
    def error_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == "error")

    @property
    def warning_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == "warning")

    @property
    def ok(self) -> bool:
        return self.error_count == 0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "book_id": self.book_id,
            "locale": self.locale,
            "package_id": self.package_id,
            "package_root": str(self.package_root),
            "template_path": str(self.template_path),
            "template_exists": self.template_exists,
            "package_exists": self.package_exists,
            "existing_user_logs": self.existing_user_logs,
            "existing_answer_images": self.existing_answer_images,
            "existing_step_images": self.existing_step_images,
            "existing_csv_rows": self.existing_csv_rows,
            "ok": self.ok,
            "error_count": self.error_count,
            "warning_count": self.warning_count,
            "issues": [issue.to_dict() for issue in self.issues],
        }


def check_full_book_step_solution_readiness(
    book_id: str,
    locales: Sequence[str],
    books_root: Path = DEFAULT_BOOKS_ROOT,
    packages_root: Path = DEFAULT_PACKAGES_ROOT,
    templates_root: Path = DEFAULT_SOLUTION_TEMPLATES_ROOT,
    require_clean_output: bool = False,
) -> Dict[str, Any]:
    """
    Preflight check before running a full-book step-solution package export.

    This does not generate anything. It verifies that the book exists, localized
    templates exist, and the output folders are in a sensible state.
    """

    books_root = Path(books_root)
    packages_root = Path(packages_root)
    templates_root = Path(templates_root)
    normalized_locales = [normalize_step_solution_locale(locale) for locale in locales]

    issues: List[FullBookReadinessIssue] = []

    book_payload: Optional[Dict[str, Any]] = None
    puzzle_count = 0

    try:
        book_payload, puzzle_count = _load_readiness_book_payload(
            book_id=book_id,
            books_root=books_root,
        )
    except Exception as exc:
        issues.append(
            FullBookReadinessIssue(
                severity="error",
                issue_type="book_load_failed",
                message=str(exc),
                path=str(books_root / book_id),
            )
        )

    locale_reports: List[FullBookReadinessLocaleReport] = []

    for locale in normalized_locales:
        request = StepSolutionPackageRequest(
            book_id=book_id,
            locale=locale,
            output_root=packages_root,
            books_root=books_root,
        )
        paths = resolve_package_paths(request)
        template_path = message_template_path(locale, templates_root)

        locale_report = FullBookReadinessLocaleReport(
            book_id=book_id,
            locale=locale,
            package_id=request.package_id(),
            package_root=paths.package_root,
            template_path=template_path,
        )

        _check_locale_template(locale_report, template_path)
        _check_existing_package_state(
            locale_report=locale_report,
            paths=paths,
            expected_puzzle_count=puzzle_count,
            require_clean_output=require_clean_output,
        )

        locale_reports.append(locale_report)

    expected = {
        "book_id": book_id,
        "locale_count": len(normalized_locales),
        "puzzle_count_per_locale": puzzle_count,
        "expected_user_logs_per_locale": puzzle_count,
        "expected_answer_images_per_locale": puzzle_count,
        "expected_csv_rows_per_locale": puzzle_count,
        "expected_total_user_logs": puzzle_count * len(normalized_locales),
        "expected_total_answer_images": puzzle_count * len(normalized_locales),
        "step_images": (
            "Variable. Depends on each puzzle's solved step count. "
            "Use runtime QA/package reports after generation."
        ),
    }

    all_errors = len([issue for issue in issues if issue.severity == "error"])
    all_warnings = len([issue for issue in issues if issue.severity == "warning"])

    all_errors += sum(report.error_count for report in locale_reports)
    all_warnings += sum(report.warning_count for report in locale_reports)

    return {
        "schema_version": "step_solution_full_book_readiness.v1",
        "book": (
            book_payload
            if book_payload is not None
            else {"book_id": book_id, "error": "book not loaded"}
        ),
        "locales": normalized_locales,
        "paths": {
            "books_root": str(books_root),
            "packages_root": str(packages_root),
            "templates_root": str(templates_root),
        },
        "expected_outputs": expected,
        "require_clean_output": require_clean_output,
        "ok": all_errors == 0,
        "error_count": all_errors,
        "warning_count": all_warnings,
        "issues": [issue.to_dict() for issue in issues],
        "locale_reports": [report.to_dict() for report in locale_reports],
        "recommended_commands": _recommended_commands(book_id, normalized_locales),
    }


def write_full_book_readiness_report(
    payload: Dict[str, Any],
    output_path: Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    return output_path



def _load_readiness_book_payload(
    book_id: str,
    books_root: Path,
) -> tuple[Dict[str, Any], int]:
    """
    Load just enough book information for production readiness checks.

    This avoids relying on book_loader helper names, which may change across
    publishing pipeline patches.
    """

    book_dir = Path(books_root) / book_id
    manifest_path = book_dir / "book_manifest.json"

    if not book_dir.exists():
        raise FileNotFoundError(f"Book directory not found: {book_dir}")

    if not manifest_path.exists():
        raise FileNotFoundError(f"Book manifest not found: {manifest_path}")

    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise ValueError(f"Could not read book manifest {manifest_path}: {exc}") from exc

    puzzle_count = _infer_puzzle_count_from_book_manifest(
        book_dir=book_dir,
        manifest=manifest,
    )

    if puzzle_count <= 0:
        raise ValueError(
            f"Could not infer puzzle count from book manifest: {manifest_path}"
        )

    metadata = dict(manifest.get("metadata") or {})

    return (
        {
            "book_id": str(
                manifest.get("book_id")
                or metadata.get("book_id")
                or book_id
            ),
            "title": str(
                manifest.get("title")
                or metadata.get("title")
                or ""
            ),
            "subtitle": str(
                manifest.get("subtitle")
                or metadata.get("subtitle")
                or ""
            ),
            "puzzle_count": puzzle_count,
            "manifest_path": str(manifest_path),
            "book_dir": str(book_dir),
            "raw_manifest_keys": sorted(str(key) for key in manifest.keys()),
        },
        puzzle_count,
    )


def _infer_puzzle_count_from_book_manifest(
    book_dir: Path,
    manifest: Dict[str, Any],
) -> int:
    direct_keys = (
        "puzzle_count",
        "total_puzzles",
        "selected_puzzle_count",
        "manifest_puzzle_count",
        "loaded_puzzle_count",
    )

    for key in direct_keys:
        value = manifest.get(key)
        parsed = _safe_positive_int(value)
        if parsed:
            return parsed

    # Common direct puzzle-list shapes.
    for key in ("puzzles", "puzzle_records", "selected_puzzles", "book_puzzles"):
        value = manifest.get(key)
        if isinstance(value, list) and value:
            return len(value)

    # Common section shapes.
    section_total = _infer_puzzle_count_from_sections(manifest)
    if section_total:
        return section_total

    # Last-resort recursive scan: collect plausible puzzle-count fields/lists
    # and use the largest. This avoids returning only a section count when the
    # manifest contains multiple nested summaries.
    candidates: List[int] = []

    def walk(obj: Any) -> None:
        if isinstance(obj, dict):
            for key, value in obj.items():
                key_str = str(key)
                if key_str in direct_keys:
                    parsed = _safe_positive_int(value)
                    if parsed:
                        candidates.append(parsed)

                if key_str in {"puzzles", "puzzle_records", "selected_puzzles", "book_puzzles"}:
                    if isinstance(value, list) and value:
                        candidates.append(len(value))

                walk(value)

        elif isinstance(obj, list):
            for item in obj:
                walk(item)

    walk(manifest)

    if candidates:
        return max(candidates)

    # Final fallback: count puzzle JSON files if the book stores per-puzzle files.
    puzzle_dirs = [
        book_dir / "puzzles",
        book_dir / "puzzle_records",
    ]
    for puzzle_dir in puzzle_dirs:
        if puzzle_dir.exists():
            files = [
                path for path in puzzle_dir.glob("*.json")
                if not path.name.startswith("_")
            ]
            if files:
                return len(files)

    return 0


def _infer_puzzle_count_from_sections(manifest: Dict[str, Any]) -> int:
    sections = manifest.get("sections")
    if not isinstance(sections, list):
        return 0

    total = 0
    for section in sections:
        if not isinstance(section, dict):
            continue

        for key in (
            "puzzle_count",
            "count",
            "selected_puzzle_count",
            "total_puzzles",
        ):
            parsed = _safe_positive_int(section.get(key))
            if parsed:
                total += parsed
                break
        else:
            puzzles = section.get("puzzles")
            if isinstance(puzzles, list):
                total += len(puzzles)

    return total


def _safe_positive_int(value: Any) -> int:
    try:
        parsed = int(value)
    except Exception:
        return 0
    return parsed if parsed > 0 else 0


def _check_locale_template(
    report: FullBookReadinessLocaleReport,
    template_path: Path,
) -> None:
    report.template_exists = template_path.exists()

    if not template_path.exists():
        report.add_issue(
            severity="error",
            issue_type="missing_locale_template",
            message="Localized Template_Messages.xlsx is missing.",
            path=template_path,
        )
        return

    try:
        wb = load_workbook(template_path, read_only=True, data_only=True)
    except Exception as exc:
        report.add_issue(
            severity="error",
            issue_type="template_read_failed",
            message=str(exc),
            path=template_path,
        )
        return

    required_sheets = ["Headers", "Messages", "Names", "Keywords"]
    missing = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]
    if missing:
        report.add_issue(
            severity="error",
            issue_type="template_missing_required_sheets",
            message="Template is missing required sheet(s).",
            path=template_path,
            details={"missing_sheets": missing},
        )


def _check_existing_package_state(
    locale_report: FullBookReadinessLocaleReport,
    paths,
    expected_puzzle_count: int,
    require_clean_output: bool,
) -> None:
    locale_report.package_exists = paths.package_root.exists()

    if not paths.package_root.exists():
        return

    if paths.user_logs_dir.exists():
        locale_report.existing_user_logs = len(
            list(paths.user_logs_dir.glob("*_user_logs.xlsx"))
        )

    if paths.image_files_dir.exists():
        locale_report.existing_answer_images = len(
            list(paths.image_files_dir.glob("*_answer.png"))
        )
        locale_report.existing_step_images = len(
            list(paths.image_files_dir.glob("*_step*.png"))
        )

    if paths.sudoku_index_csv_path.exists():
        locale_report.existing_csv_rows = _count_csv_rows(paths.sudoku_index_csv_path)

    if require_clean_output:
        existing_counts = {
            "user_logs": locale_report.existing_user_logs,
            "answer_images": locale_report.existing_answer_images,
            "step_images": locale_report.existing_step_images,
            "csv_rows": locale_report.existing_csv_rows,
        }
        if any(count > 0 for count in existing_counts.values()):
            locale_report.add_issue(
                severity="error",
                issue_type="output_not_clean",
                message="Existing outputs were found but clean output was required.",
                path=paths.package_root,
                details=existing_counts,
            )
    else:
        if locale_report.existing_user_logs and locale_report.existing_user_logs < expected_puzzle_count:
            locale_report.add_issue(
                severity="warning",
                issue_type="partial_existing_user_logs",
                message="Existing package appears partial. Use --force for clean rebuild or --skip-existing to resume.",
                path=paths.user_logs_dir,
                details={
                    "existing_user_logs": locale_report.existing_user_logs,
                    "expected_puzzle_count": expected_puzzle_count,
                },
            )


def _count_csv_rows(path: Path) -> int:
    try:
        import csv

        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return len(list(csv.DictReader(handle)))
    except Exception:
        return 0


def _recommended_commands(book_id: str, locales: Sequence[str]) -> Dict[str, str]:
    locale_string = " ".join(locales)
    first_locale = locales[0] if locales else "fr"

    return {
        "prepare_templates": (
            "python -m python.publishing.workflows.prepare_step_solution_localization_templates "
            f"--locales {locale_string}"
        ),
        "one_puzzle_qa": (
            "python -m python.publishing.workflows.export_step_solution_packages_with_qa "
            f"--book-id {book_id} --locales {locale_string} --only-puzzle L1-001 "
            "--force --max-workbooks-to-check 1 --max-csv-rows-to-check 1"
        ),
        "five_puzzle_qa": (
            "python -m python.publishing.workflows.export_step_solution_packages_with_qa "
            f"--book-id {book_id} --locales {locale_string} --limit 5 "
            "--force --max-workbooks-to-check 3 --max-csv-rows-to-check 5"
        ),
        "full_one_locale": (
            "python -m python.publishing.workflows.export_step_solution_packages_with_qa "
            f"--book-id {book_id} --locales {first_locale} "
            "--force --max-workbooks-to-check 10 --max-csv-rows-to-check 25"
        ),
        "full_all_locales": (
            "python -m python.publishing.workflows.export_step_solution_packages_with_qa "
            f"--book-id {book_id} --locales {locale_string} "
            "--force --max-workbooks-to-check 10 --max-csv-rows-to-check 25"
        ),
        "resume_all_locales": (
            "python -m python.publishing.workflows.export_step_solution_packages_with_qa "
            f"--book-id {book_id} --locales {locale_string} "
            "--skip-existing --max-workbooks-to-check 10 --max-csv-rows-to-check 25"
        ),
    }