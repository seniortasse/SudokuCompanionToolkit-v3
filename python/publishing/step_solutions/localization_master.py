from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from python.publishing.step_solutions.locale_templates import (
    DEFAULT_SOLUTION_TEMPLATES_ROOT,
    message_template_path,
    normalize_step_solution_locale,
)
from python.publishing.step_solutions.template_localization_contract import (
    LOCALIZATION_MASTER_SCHEMA_VERSION,
    extract_placeholders,
    normalize_template_text,
)


DEFAULT_LOCALIZATION_MASTER_DIR = Path(
    "datasets/sudoku_books/classic9/solution_templates/localization_master"
)
DEFAULT_LOCALIZATION_MASTER_XLSX = (
    DEFAULT_LOCALIZATION_MASTER_DIR / "step_solution_localization_master.xlsx"
)
DEFAULT_LOCALIZATION_MASTER_IMPORT_REPORT = (
    DEFAULT_LOCALIZATION_MASTER_DIR / "step_solution_localization_master.import_report.json"
)

MASTER_SHEETS = (
    "Metadata",
    "Headers",
    "Messages",
    "Names",
    "Keywords",
    "AuditNotes",
)

MESSAGE_QUOTED_FRAGMENT_RE = re.compile(r'"([^"]*)"', re.DOTALL)


def build_step_solution_localization_master(
    templates_root: Path = DEFAULT_SOLUTION_TEMPLATES_ROOT,
    locales: Sequence[str] = ("en", "fr", "de", "it", "es"),
    output_xlsx: Path = DEFAULT_LOCALIZATION_MASTER_XLSX,
    report_json: Path = DEFAULT_LOCALIZATION_MASTER_IMPORT_REPORT,
    overwrite: bool = False,
) -> Dict[str, Any]:
    """
    Build the long-term localization master workbook.

    English remains the schema master. Localized templates are imported as
    current seed translations.

    This function does not modify any existing Template_Messages.xlsx file.
    """

    templates_root = Path(templates_root)
    output_xlsx = Path(output_xlsx)
    report_json = Path(report_json)

    normalized_locales = _normalize_locales(locales)
    if "en" not in normalized_locales:
        normalized_locales = ["en"] + normalized_locales

    if output_xlsx.exists() and not overwrite:
        raise FileExistsError(
            f"Localization master already exists: {output_xlsx}. "
            "Pass --overwrite to rebuild it."
        )

    source_paths = {
        locale: message_template_path(locale, templates_root)
        for locale in normalized_locales
    }

    source_workbooks = {
        locale: _load_template_workbook(path)
        for locale, path in source_paths.items()
    }

    english_wb = source_workbooks["en"]

    master_wb = Workbook()
    default_sheet = master_wb.active
    master_wb.remove(default_sheet)

    for sheet_name in MASTER_SHEETS:
        master_wb.create_sheet(sheet_name)

    import_notes: List[Dict[str, Any]] = []

    _write_metadata_sheet(
        ws=master_wb["Metadata"],
        templates_root=templates_root,
        source_paths=source_paths,
        output_xlsx=output_xlsx,
        locales=normalized_locales,
    )

    header_records = _build_headers_records(
        english_wb=english_wb,
        source_workbooks=source_workbooks,
        locales=normalized_locales,
        import_notes=import_notes,
    )
    message_records = _build_messages_records(
        english_wb=english_wb,
        source_workbooks=source_workbooks,
        locales=normalized_locales,
        import_notes=import_notes,
    )
    name_records = _build_names_records(
        english_wb=english_wb,
        source_workbooks=source_workbooks,
        locales=normalized_locales,
        import_notes=import_notes,
    )
    keyword_records = _build_keywords_records(
        english_wb=english_wb,
        source_workbooks=source_workbooks,
        locales=normalized_locales,
        import_notes=import_notes,
    )

    _write_records_sheet(
        ws=master_wb["Headers"],
        records=header_records,
        locales=normalized_locales,
    )
    _write_records_sheet(
        ws=master_wb["Messages"],
        records=message_records,
        locales=normalized_locales,
    )
    _write_records_sheet(
        ws=master_wb["Names"],
        records=name_records,
        locales=normalized_locales,
    )
    _write_records_sheet(
        ws=master_wb["Keywords"],
        records=keyword_records,
        locales=normalized_locales,
    )
    _write_audit_notes_sheet(
        ws=master_wb["AuditNotes"],
        notes=import_notes,
    )

    _style_master_workbook(master_wb)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    master_wb.save(output_xlsx)

    report = {
        "schema_version": LOCALIZATION_MASTER_SCHEMA_VERSION,
        "created_at": _now_iso(),
        "templates_root": str(templates_root),
        "output_xlsx": str(output_xlsx),
        "source_paths": {locale: str(path) for locale, path in source_paths.items()},
        "locales": normalized_locales,
        "record_counts": {
            "headers": len(header_records),
            "messages": len(message_records),
            "names": len(name_records),
            "keywords": len(keyword_records),
            "audit_notes": len(import_notes),
        },
        "coverage": _build_coverage_summary(
            records=header_records + message_records + name_records + keyword_records,
            locales=normalized_locales,
        ),
        "audit_notes": import_notes,
    }

    report_json.parent.mkdir(parents=True, exist_ok=True)
    report_json.write_text(
        json.dumps(report, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    return report


def _normalize_locales(locales: Sequence[str]) -> List[str]:
    out: List[str] = []
    for locale in locales:
        normalized = normalize_step_solution_locale(locale)
        if normalized not in out:
            out.append(normalized)
    return out


def _load_template_workbook(path: Path):
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Template workbook not found: {path}")
    return load_workbook(path, data_only=False)


def _now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _write_metadata_sheet(
    ws,
    templates_root: Path,
    source_paths: Dict[str, Path],
    output_xlsx: Path,
    locales: Sequence[str],
) -> None:
    rows = [
        ("schema_version", LOCALIZATION_MASTER_SCHEMA_VERSION),
        ("created_at", _now_iso()),
        ("templates_root", str(templates_root)),
        ("output_xlsx", str(output_xlsx)),
        ("locales", ", ".join(locales)),
        ("note", "English Template_Messages.xlsx is the schema master."),
        ("note", "Generated localized Template_Messages.xlsx files should be built from this master."),
        ("note", "Machine keys must not be translated."),
        ("note", "Placeholders such as {cell}, {char}, {dim} must be preserved exactly."),
    ]

    for locale, path in source_paths.items():
        rows.append((f"source_{locale}", str(path)))

    ws.append(["key", "value"])
    for key, value in rows:
        ws.append([key, value])


def _build_headers_records(
    english_wb,
    source_workbooks: Dict[str, Any],
    locales: Sequence[str],
    import_notes: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    sheet_name = "Headers"
    if sheet_name not in english_wb.sheetnames:
        import_notes.append(_note("en", sheet_name, None, None, "missing_sheet", "English Headers sheet missing."))
        return []

    ws_en = english_wb[sheet_name]
    records: List[Dict[str, Any]] = []

    for row in ws_en.iter_rows():
        for cell in row:
            en_value = _cell_text(cell.value)
            if not en_value:
                continue

            row_index = int(cell.row)
            col_index = int(cell.column)
            record = _base_record(
                record_id=f"headers_r{row_index}_c{col_index}",
                sheet=sheet_name,
                source_row=row_index,
                source_column=col_index,
                role="translatable",
                key=f"r{row_index}_c{col_index}",
                field="header_cell",
                fragment_index=0,
                en=en_value,
                placeholders=extract_placeholders(en_value),
                notes="Header/template text. Translate naturally while preserving placeholders.",
                status="seeded",
            )

            _fill_locale_values_by_same_cell(
                record=record,
                source_workbooks=source_workbooks,
                locales=locales,
                sheet_name=sheet_name,
                row=row_index,
                column=col_index,
                import_notes=import_notes,
            )
            records.append(record)

    return records


def _build_messages_records(
    english_wb,
    source_workbooks: Dict[str, Any],
    locales: Sequence[str],
    import_notes: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    sheet_name = "Messages"
    if sheet_name not in english_wb.sheetnames:
        import_notes.append(_note("en", sheet_name, None, None, "missing_sheet", "English Messages sheet missing."))
        return []

    ws_en = english_wb[sheet_name]
    records: List[Dict[str, Any]] = []

    for row in ws_en.iter_rows():
        row_values = [_cell_text(cell.value) for cell in row]
        technique_keys = _technique_keys_from_row(row_values)
        if not technique_keys:
            continue

        for cell in row:
            en_cell_value = _cell_text(cell.value)
            if not en_cell_value:
                continue

            if _looks_like_machine_key_blob(en_cell_value):
                continue

            # Do not import explanatory/editorial remarks as translatable
            # solver narration records. These cells are comments for template
            # maintainers, not messages consumed by the legacy writer.
            if en_cell_value.strip().lower().startswith("remark:"):
                continue

            row_index = int(cell.row)
            col_index = int(cell.column)
            en_fragments = _split_message_fragments(en_cell_value)

            for fragment_index, en_fragment in enumerate(en_fragments, start=1):
                record = _base_record(
                    record_id=f"messages_r{row_index}_c{col_index}_f{fragment_index}",
                    sheet=sheet_name,
                    source_row=row_index,
                    source_column=col_index,
                    role="message_fragment",
                    key="|".join(technique_keys),
                    field="message",
                    fragment_index=fragment_index,
                    en=en_fragment,
                    placeholders=extract_placeholders(en_fragment),
                    notes="Solver narration fragment. Preserve placeholders exactly.",
                    status="seeded",
                )
                record["technique_keys"] = "\n".join(technique_keys)

                for locale in locales:
                    if locale == "en":
                        record["en"] = en_fragment
                        continue

                    loc_text = _cell_text_from_workbook(
                        source_workbooks=source_workbooks,
                        locale=locale,
                        sheet_name=sheet_name,
                        row=row_index,
                        column=col_index,
                        import_notes=import_notes,
                    )
                    loc_fragments = _split_message_fragments(loc_text)
                    loc_fragment = (
                        loc_fragments[fragment_index - 1]
                        if fragment_index - 1 < len(loc_fragments)
                        else ""
                    )
                    record[locale] = loc_fragment

                    _add_fragment_import_notes(
                        record=record,
                        locale=locale,
                        localized_value=loc_fragment,
                        import_notes=import_notes,
                    )

                records.append(record)

    return records


def _build_names_records(
    english_wb,
    source_workbooks: Dict[str, Any],
    locales: Sequence[str],
    import_notes: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    sheet_name = "Names"
    if sheet_name not in english_wb.sheetnames:
        import_notes.append(_note("en", sheet_name, None, None, "missing_sheet", "English Names sheet missing."))
        return []

    ws_en = english_wb[sheet_name]
    records: List[Dict[str, Any]] = []

    for row in ws_en.iter_rows():
        technique_key = _cell_text(row[0].value) if row else ""
        if not _looks_like_machine_key_blob(technique_key):
            continue

        for col_index, field_name in ((2, "technique_name"), (3, "rating")):
            cell = ws_en.cell(row=int(row[0].row), column=col_index)
            en_value = _cell_text(cell.value)
            if not en_value:
                continue

            row_index = int(cell.row)
            record = _base_record(
                record_id=f"names_{technique_key}_{field_name}",
                sheet=sheet_name,
                source_row=row_index,
                source_column=col_index,
                role="translatable",
                key=technique_key,
                field=field_name,
                fragment_index=0,
                en=en_value,
                placeholders=extract_placeholders(en_value),
                notes="Technique display name or difficulty rating.",
                status="seeded",
            )

            _fill_locale_values_by_same_cell(
                record=record,
                source_workbooks=source_workbooks,
                locales=locales,
                sheet_name=sheet_name,
                row=row_index,
                column=col_index,
                import_notes=import_notes,
            )
            _add_basic_import_notes_for_record(record, import_notes)
            records.append(record)

    return records


def _build_keywords_records(
    english_wb,
    source_workbooks: Dict[str, Any],
    locales: Sequence[str],
    import_notes: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    sheet_name = "Keywords"
    if sheet_name not in english_wb.sheetnames:
        import_notes.append(_note("en", sheet_name, None, None, "missing_sheet", "English Keywords sheet missing."))
        return []

    ws_en = english_wb[sheet_name]
    records: List[Dict[str, Any]] = []

    for row in ws_en.iter_rows():
        keyword_key = _cell_text(row[0].value) if row else ""
        if not keyword_key:
            continue

        for cell in row[1:]:
            en_value = _cell_text(cell.value)
            if not en_value:
                continue

            row_index = int(cell.row)
            col_index = int(cell.column)
            record = _base_record(
                record_id=f"keywords_r{row_index}_c{col_index}",
                sheet=sheet_name,
                source_row=row_index,
                source_column=col_index,
                role="keyword_value",
                key=keyword_key,
                field=f"value_{col_index}",
                fragment_index=0,
                en=en_value,
                placeholders=extract_placeholders(en_value),
                notes="Keyword vocabulary value. Keep grammar compatible with message templates.",
                status="seeded",
            )

            _fill_locale_values_by_same_cell(
                record=record,
                source_workbooks=source_workbooks,
                locales=locales,
                sheet_name=sheet_name,
                row=row_index,
                column=col_index,
                import_notes=import_notes,
            )
            _add_basic_import_notes_for_record(record, import_notes)
            records.append(record)

    return records


def _base_record(
    record_id: str,
    sheet: str,
    source_row: int,
    source_column: int,
    role: str,
    key: str,
    field: str,
    fragment_index: int,
    en: str,
    placeholders: Iterable[str],
    notes: str,
    status: str,
) -> Dict[str, Any]:
    return {
        "record_id": record_id,
        "sheet": sheet,
        "source_row": source_row,
        "source_column": source_column,
        "role": role,
        "key": key,
        "field": field,
        "fragment_index": fragment_index,
        "technique_keys": "",
        "en": en,
        "fr": "",
        "de": "",
        "it": "",
        "es": "",
        "placeholders": " ".join(sorted(placeholders)),
        "notes": notes,
        "status": status,
    }


def _fill_locale_values_by_same_cell(
    record: Dict[str, Any],
    source_workbooks: Dict[str, Any],
    locales: Sequence[str],
    sheet_name: str,
    row: int,
    column: int,
    import_notes: List[Dict[str, Any]],
) -> None:
    for locale in locales:
        if locale == "en":
            record["en"] = record["en"]
            continue

        value = _cell_text_from_workbook(
            source_workbooks=source_workbooks,
            locale=locale,
            sheet_name=sheet_name,
            row=row,
            column=column,
            import_notes=import_notes,
        )
        record[locale] = value
        _add_basic_import_notes_for_locale_value(record, locale, value, import_notes)


def _cell_text_from_workbook(
    source_workbooks: Dict[str, Any],
    locale: str,
    sheet_name: str,
    row: int,
    column: int,
    import_notes: List[Dict[str, Any]],
) -> str:
    wb = source_workbooks.get(locale)
    if wb is None:
        import_notes.append(
            _note(locale, sheet_name, row, column, "missing_workbook", "Locale workbook was not loaded.")
        )
        return ""

    if sheet_name not in wb.sheetnames:
        import_notes.append(
            _note(locale, sheet_name, row, column, "missing_sheet", f"Sheet {sheet_name!r} missing.")
        )
        return ""

    return _cell_text(wb[sheet_name].cell(row=row, column=column).value)


def _add_fragment_import_notes(
    record: Dict[str, Any],
    locale: str,
    localized_value: str,
    import_notes: List[Dict[str, Any]],
) -> None:
    _add_basic_import_notes_for_locale_value(record, locale, localized_value, import_notes)

    en_placeholders = extract_placeholders(record["en"])
    loc_placeholders = extract_placeholders(localized_value)

    if localized_value and en_placeholders != loc_placeholders:
        import_notes.append(
            _note(
                locale=locale,
                sheet=record["sheet"],
                row=record["source_row"],
                column=record["source_column"],
                issue_type="placeholder_mismatch_seed",
                message="Imported localized fragment has placeholder mismatch.",
                details={
                    "record_id": record["record_id"],
                    "missing": sorted(en_placeholders - loc_placeholders),
                    "extra": sorted(loc_placeholders - en_placeholders),
                    "en": record["en"],
                    "localized": localized_value,
                },
            )
        )


def _add_basic_import_notes_for_record(
    record: Dict[str, Any],
    import_notes: List[Dict[str, Any]],
) -> None:
    for locale in ("fr", "de", "it", "es"):
        _add_basic_import_notes_for_locale_value(
            record=record,
            locale=locale,
            value=str(record.get(locale) or ""),
            import_notes=import_notes,
        )


def _add_basic_import_notes_for_locale_value(
    record: Dict[str, Any],
    locale: str,
    value: str,
    import_notes: List[Dict[str, Any]],
) -> None:
    value = str(value or "")
    if not value:
        import_notes.append(
            _note(
                locale=locale,
                sheet=record["sheet"],
                row=record["source_row"],
                column=record["source_column"],
                issue_type="missing_seed_translation",
                message="No existing localized value found for this master record.",
                details={"record_id": record["record_id"], "en": record["en"]},
            )
        )
        return

    if value == record["en"] and locale != "en":
        import_notes.append(
            _note(
                locale=locale,
                sheet=record["sheet"],
                row=record["source_row"],
                column=record["source_column"],
                issue_type="same_as_english_seed",
                message="Localized seed value is identical to English.",
                details={"record_id": record["record_id"], "value": value},
            )
        )


def _note(
    locale: str,
    sheet: str,
    row: Optional[int],
    column: Optional[int],
    issue_type: str,
    message: str,
    details: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    return {
        "locale": locale,
        "sheet": sheet,
        "row": row,
        "column": column,
        "issue_type": issue_type,
        "message": message,
        "details": details or {},
    }


def _technique_keys_from_row(row_values: Sequence[str]) -> List[str]:
    keys: List[str] = []
    for value in row_values:
        if not value:
            continue
        if _looks_like_machine_key_blob(value):
            keys.extend([part.strip() for part in value.splitlines() if part.strip()])
    return keys


def _looks_like_machine_key_blob(value: str) -> bool:
    text = str(value or "").strip()
    if not text:
        return False

    parts = [part.strip() for part in text.splitlines() if part.strip()]
    if not parts:
        return False

    return all(_looks_like_machine_key(part) for part in parts)


def _looks_like_machine_key(value: str) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    if " " in text:
        return False
    if "{" in text or "}" in text:
        return False
    allowed = set("abcdefghijklmnopqrstuvwxyz0123456789-_")
    return text.lower() == text and all(char in allowed for char in text)


def _split_message_fragments(value: str) -> List[str]:
    """
    Split a Messages-sheet cell into editable message fragments.

    Most cells are:
        "Sentence with {placeholder}. "

    Some cells contain multiple quoted fragments:
        "First fragment. "
        "Second fragment. "

    The master stores those as separate rows so they can be translated and
    validated independently.
    """

    text = _cell_text(value)
    if not text:
        return []

    matches = MESSAGE_QUOTED_FRAGMENT_RE.findall(text)
    if matches:
        return [match.replace('""', '"') for match in matches]

    return [text]


def _cell_text(value: object) -> str:
    return normalize_template_text(value)


def _write_records_sheet(ws, records: List[Dict[str, Any]], locales: Sequence[str]) -> None:
    columns = [
        "record_id",
        "sheet",
        "source_row",
        "source_column",
        "role",
        "key",
        "field",
        "fragment_index",
        "technique_keys",
        "en",
    ]

    for locale in locales:
        if locale != "en":
            columns.append(locale)

    columns.extend(
        [
            "placeholders",
            "notes",
            "status",
        ]
    )

    ws.append(columns)

    for record in records:
        row = []
        for column in columns:
            row.append(record.get(column, ""))
        ws.append(row)


def _write_audit_notes_sheet(ws, notes: List[Dict[str, Any]]) -> None:
    columns = [
        "locale",
        "sheet",
        "row",
        "column",
        "issue_type",
        "message",
        "details_json",
    ]
    ws.append(columns)

    for note in notes:
        ws.append(
            [
                note.get("locale", ""),
                note.get("sheet", ""),
                note.get("row", ""),
                note.get("column", ""),
                note.get("issue_type", ""),
                note.get("message", ""),
                json.dumps(note.get("details") or {}, ensure_ascii=False),
            ]
        )


def _style_master_workbook(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        _auto_width(ws)


def _auto_width(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = str(cell.value or "")
            if len(value) > max_length:
                max_length = min(len(value), 70)

        ws.column_dimensions[column_letter].width = max(12, min(max_length + 2, 70))


def _build_coverage_summary(
    records: List[Dict[str, Any]],
    locales: Sequence[str],
) -> Dict[str, Any]:
    coverage: Dict[str, Any] = {}

    for locale in locales:
        if locale == "en":
            continue

        total = len(records)
        filled = sum(1 for record in records if str(record.get(locale) or "").strip())
        same_as_english = sum(
            1
            for record in records
            if str(record.get(locale) or "").strip()
            and str(record.get(locale) or "").strip() == str(record.get("en") or "").strip()
        )

        coverage[locale] = {
            "total_records": total,
            "filled_records": filled,
            "missing_records": total - filled,
            "same_as_english_records": same_as_english,
            "filled_percent": round((filled / total * 100.0), 2) if total else 0.0,
        }

    return coverage