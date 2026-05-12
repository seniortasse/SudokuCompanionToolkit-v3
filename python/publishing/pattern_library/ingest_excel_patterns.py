from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from python.publishing.ids.id_policy import build_pattern_id
from python.publishing.pattern_library.pattern_enricher import enrich_pattern_record
from python.publishing.pattern_library.pattern_identity import build_variant_code
from python.publishing.pattern_library.pattern_registry import PatternRegistry
from python.publishing.pattern_library.pattern_validator import validate_pattern_record_strict
from python.publishing.schemas.models import PatternRecord


TRUTHY_VALUES = {
    "1",
    "X",
    "TRUE",
    "T",
    "#",
    "■",
    "█",
    "YES",
    "Y",
}

FALSY_VALUES = {
    "0",
    ".",
    "",
    "FALSE",
    "F",
    "-",
    "_",
    "NO",
    "N",
}


@dataclass
class ExcelPatternSource:
    workbook_path: Path
    sheet_name: str
    top_row: int = 1
    left_col: int = 1
    pattern_name: Optional[str] = None
    slug: Optional[str] = None
    family_id: Optional[str] = None
    family_name: Optional[str] = None
    description: str = ""
    tags: Optional[List[str]] = None
    aliases: Optional[List[str]] = None
    status: str = "active"
    author: Optional[str] = None
    notes: Optional[str] = None
    source_ref: Optional[str] = None


@dataclass
class IngestResult:
    added_patterns: List[PatternRecord]
    skipped_duplicates: List[str]
    validation_failures: List[Tuple[str, List[str]]]


def _cell_to_binary(value: object) -> str:
    if value is None:
        return "0"

    if isinstance(value, bool):
        return "1" if value else "0"

    if isinstance(value, (int, float)):
        return "1" if int(value) != 0 else "0"

    s = str(value).strip().upper()
    if s in TRUTHY_VALUES:
        return "1"
    if s in FALSY_VALUES:
        return "0"

    return "1"


def _extract_mask81(sheet: Worksheet, top_row: int, left_col: int) -> str:
    bits: List[str] = []
    for r in range(top_row, top_row + 9):
        for c in range(left_col, left_col + 9):
            bits.append(_cell_to_binary(sheet.cell(row=r, column=c).value))
    return "".join(bits)


def _default_pattern_name(sheet_name: str, ordinal_hint: int) -> str:
    cleaned_sheet = str(sheet_name).strip() or "Sheet"
    return f"{cleaned_sheet} Pattern {ordinal_hint}"


def _parse_listish(value: object) -> List[str]:
    if value is None:
        return []

    raw = str(value).strip()
    if not raw:
        return []

    normalized = raw.replace("|", ",").replace(";", ",")
    parts = [part.strip() for part in normalized.split(",")]
    return [part for part in parts if part]


def _normalize_metadata_key(value: object) -> Optional[str]:
    if value is None:
        return None
    key = str(value).strip().lower()
    return key or None


def _read_metadata_panel(
    sheet: Worksheet,
    *,
    key_col: int = 11,   # K
    value_col: int = 12, # L
    start_row: int = 1,
    end_row: int = 20,
) -> Dict[str, str]:
    metadata: Dict[str, str] = {}

    for row in range(start_row, end_row + 1):
        key = _normalize_metadata_key(sheet.cell(row=row, column=key_col).value)
        if not key:
            continue

        value = sheet.cell(row=row, column=value_col).value
        metadata[key] = "" if value is None else str(value).strip()

    return metadata


def ingest_sources_into_registry(
    *,
    library_id: str,
    library_short: str,
    sources: Sequence[ExcelPatternSource],
    registry: PatternRegistry,
    created_at: Optional[str] = None,
    updated_at: Optional[str] = None,
) -> IngestResult:
    added_patterns: List[PatternRecord] = []
    skipped_duplicates: List[str] = []
    validation_failures: List[Tuple[str, List[str]]] = []

    next_ordinal = registry.get_next_ordinal()

    for source in sources:
        workbook = load_workbook(filename=str(source.workbook_path), data_only=True)
        try:
            if source.sheet_name not in workbook.sheetnames:
                validation_failures.append(
                    (
                        f"{source.workbook_path}::{source.sheet_name}",
                        [f"Sheet not found: {source.sheet_name}"],
                    )
                )
                continue

            sheet = workbook[source.sheet_name]
            mask81 = _extract_mask81(sheet, source.top_row, source.left_col)
            clue_count = sum(1 for ch in mask81 if ch == "1")

            family_name_for_variant = source.family_name or source.family_id or "base"

            pattern_name = source.pattern_name or _default_pattern_name(source.sheet_name, next_ordinal)
            pattern = PatternRecord(
                pattern_id=build_pattern_id(library_short, next_ordinal),
                library_id=library_id,
                name=pattern_name,
                slug=str(source.slug or ""),
                aliases=list(source.aliases or []),
                description=source.description,
                grid_size=9,
                layout_type="classic9x9",
                mask81=mask81,
                canonical_mask_signature="",
                clue_count=clue_count,
                symmetry_type="none",
                visual_family=str(source.family_id or "uncategorized"),
                family_id=source.family_id,
                family_name=source.family_name,
                variant_code=build_variant_code(family_name=family_name_for_variant, ordinal=1),
                tags=list(source.tags or []),
                status=str(source.status or "active"),
                source_type="excel_import",
                source_ref=source.source_ref or f"{source.workbook_path.name}::{source.sheet_name}",
                author=source.author,
                notes=source.notes,
                is_verified=False,
                created_at=created_at,
                updated_at=updated_at,
            )

            pattern = enrich_pattern_record(pattern)
            errors = validate_pattern_record_strict(pattern)

            if errors:
                validation_failures.append((pattern.name, errors))
                continue

            was_added = registry.add_or_skip_duplicate_mask(pattern)
            if not was_added:
                skipped_duplicates.append(pattern.name)
                continue

            added_patterns.append(pattern)
            next_ordinal += 1
        finally:
            workbook.close()

    return IngestResult(
        added_patterns=added_patterns,
        skipped_duplicates=skipped_duplicates,
        validation_failures=validation_failures,
    )


def build_sources_from_workbook(
    workbook_path: Path,
    *,
    sheet_names: Optional[Iterable[str]] = None,
    top_row: int = 1,
    left_col: int = 1,
    default_tags: Optional[List[str]] = None,
) -> List[ExcelPatternSource]:
    workbook = load_workbook(filename=str(workbook_path), read_only=False, data_only=True)
    try:
        selected_sheets = list(sheet_names) if sheet_names else list(workbook.sheetnames)

        sources: List[ExcelPatternSource] = []
        for sheet_name in selected_sheets:
            if sheet_name not in workbook.sheetnames:
                continue

            sheet = workbook[sheet_name]
            metadata = _read_metadata_panel(sheet)

            tags = _parse_listish(metadata.get("tags"))
            if default_tags:
                tags = list(dict.fromkeys(tags + list(default_tags)))

            aliases = _parse_listish(metadata.get("aliases"))

            source = ExcelPatternSource(
                workbook_path=workbook_path,
                sheet_name=sheet_name,
                top_row=top_row,
                left_col=left_col,
                pattern_name=metadata.get("name") or None,
                slug=metadata.get("slug") or None,
                family_id=metadata.get("family_id") or None,
                family_name=metadata.get("family_name") or None,
                description=metadata.get("description") or f"Ingested from {workbook_path.name}, sheet {sheet_name}",
                tags=tags,
                aliases=aliases,
                status=metadata.get("status") or "active",
                author=metadata.get("author") or None,
                notes=metadata.get("notes") or None,
                source_ref=metadata.get("source_ref") or f"{workbook_path.name}::{sheet_name}",
            )
            sources.append(source)

        return sources
    finally:
        workbook.close()