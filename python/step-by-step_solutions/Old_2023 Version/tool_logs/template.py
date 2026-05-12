
import re

from generator.algo_human import TECHNIQUES


SHEET_NAME_STEPS = "Solution"
SHEET_NAME_COLOR_CODES = "Color Code"
SHEET_NAME_FONT_SIZE = "Message Font Size"

ROW_START = 29
COL_START = 2

# STEP_NUM_ROWS = 1 + 9 + 3

# ROW_TO = 31
# COL_TO = 22

LENGTH = 3
SIZE = 9


"""

raises TemplateError
"""
def read_formatting(workbook):

    sheet_format = workbook.get_sheet_by_name(SHEET_NAME_STEPS)
    sheet_color_codes = workbook.get_sheet_by_name(SHEET_NAME_COLOR_CODES)
    sheet_font_sizes = workbook.get_sheet_by_name(SHEET_NAME_FONT_SIZE)

    ### READ FORMATTING

    # Read format
    style_header_page, style_header_step, style_cells, style_message, style_borders, colors_boxes = \
        _read_styles(sheet_format)

    # Read color codes
    colors_new_values, colors_bg, colors_key_cells, colors_bg_cleanup, font_cleanup = _read_colors(sheet_color_codes)

    # Read font size for messages (which depends on the number of new values found)
    font_sizes = _read_font_size(sheet_font_sizes)

    formatting = (
        style_header_page, style_header_step, style_cells, style_message, style_borders, colors_boxes,
        colors_new_values, colors_bg, colors_key_cells, colors_bg_cleanup, font_cleanup,
        font_sizes
    )

    return formatting


def _read_styles(sheet):

    cell_header_page = sheet.cell(ROW_START, COL_START)
    cell_header_step = sheet.cell(ROW_START + 2, COL_START)
    cells_grid = [
        [
            sheet.cell(ROW_START + 3 + i1, COL_START + i2)
            for i2 in range(SIZE)
        ]
        for i1 in range(SIZE)
    ]
    cell_message = sheet.cell(ROW_START + 3 + SIZE, COL_START)

    style_header_main = cell_header_page._style
    style_header_step = cell_header_step._style
    style_cells = [[cell._style for cell in row] for row in cells_grid]
    style_message = cell_message._style

    # Used for drawing borders for custom boxes layout
    style_borders = cells_grid[0][0].border

    # Used for coloring custom boxes
    cells_colors_boxes = [sheet.cell(4, 2), sheet.cell(5, 2)]
    colors_boxes = [cell.fill for cell in cells_colors_boxes]

    return style_header_main, style_header_step, style_cells, style_message, style_borders, colors_boxes


def _read_colors(sheet):

    # Process table
    # col_names = sheet_color_codes.get_row(2)
    # col_idx_technique_name =
    # technique_names = list(map(lambda x: x.value, sheet_color_codes["A"]))[2:]

    colors_new_values = {}
    colors_bg = {}
    colors_key_cells = {}
    colors_bg_cleanup = {}
    font_cleanup = {}
    for row in sheet.iter_rows(min_row=2):
        technique_name = row[0].value
        # TODO Temporary check which should be changed in the template (add "s" if not present)
        technique_name = re.sub(r"wing(s?)", r"wings", technique_name)
        colors_new_values[technique_name] = row[1].font.color
        colors_bg[technique_name] = row[2].fill.bgColor
        colors_key_cells[technique_name] = row[3].font
        colors_bg_cleanup[technique_name] = row[4].fill.bgColor
        font_cleanup[technique_name] = row[5].font

    technique_names_undefined = set(TECHNIQUES).difference(colors_new_values.keys())
    if technique_names_undefined:
        print(f"WARNING - Undefined colors for techniques: {technique_names_undefined}")

    return colors_new_values, colors_bg, colors_key_cells, colors_bg_cleanup, font_cleanup


def __read_font_size(workbook):
    font_sizes = {}
    for technique_name in TECHNIQUES:
        font_sizes[technique_name] = {}
        try:
            sheet = workbook.get_sheet_by_name(technique_name)
            # Unfortunately, cells are merged
            # print(technique_name)
            for merged_cell in sheet.merged_cells:
                cell = sheet[str(merged_cell).split(':')[0]]
                if cell.col_idx == 1:
                    if cell.value:
                        num_vals = cell.value.split()[0]
                        if num_vals.isdigit():
                            cell_text = cell.offset(0, 1)
                            font_size = cell_text.font.size
                            font_sizes[technique_name][int(num_vals)] = font_size
                            # print(num_vals, font_size)
                        else:
                            print(f"WARNING - Cell value {cell.value} cannot be parsed")
        except KeyError:
            print("WARNING - No font size found for technique:", technique_name)
        print(f"Font sizes for technique {technique_name}:", font_sizes[technique_name])

    return font_sizes


def _read_font_size(sheet):
    # Based on message length
    font_sizes = {}
    for row in sheet.iter_rows(min_row=2):
        message_length = row[0].value
        # We have to do this check as the cells are merged
        if message_length is not None:
            assert isinstance(message_length, int)
            font_size = row[1].font.size
            font_sizes[message_length] = font_size
    print("Font sizes for messages:", font_sizes)
    return font_sizes


if __name__ == "__main__":

    from openpyxl.reader.excel import load_workbook

    from tool_logs.writer import FILE_NAME_TEMPLATE

    workbook = load_workbook(FILE_NAME_TEMPLATE, data_only=True)

    formatting = read_formatting(workbook)

    # sheet = workbook[SHEET_NAME_COLOR_CODES]
    sheet = workbook[SHEET_NAME_FONT_SIZE]
