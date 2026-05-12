
from openpyxl import load_workbook
import re

from generator.model import DIMENSIONS
from generator.algo_human import TECHNIQUES, BASE_TECHNIQUES, ADVANCED_TECHNIQUES

# TODO Prettify
import tool_logs.messages_templates as messages_templates


def read_template(file_name_template):

    workbook = load_workbook(file_name_template)

    try:

        print("Reading templates for headers..")

        sheet = workbook["Headers"]
        templates_headers = _read_templates_headers(sheet)

        print("Reading templates for messages..")

        sheet = workbook["Messages"]
        templates_messages = _read_templates_messages(sheet)

        print("Reading technique names..")

        sheet = workbook["Names"]
        _read_names(sheet)

        print("Reading keywords..")

        sheet = workbook["Keywords"]
        _read_keywords(sheet)

    except AssertionError as e:
        error_message = '\n'.join(
            (
                "Something went wrong when reading templates for messages..",
                str(e)
            )
        )
        raise Exception(error_message)

    # template_messages = {
    #     "headers": templates_headers
    # }

    # return template_messages

    # raise Exception("Testing template")


def _read_templates_headers(sheet):

    row_no, col_no = 2, 1

    assert sheet.cell(row_no, col_no).value == "HEADERS"

    # TODO Should we make this adjustable?
    template_placeholders_required = {
        "main": ["instance_id"],
        "step": ["step_no"],
    }

    row_no += 1
    assert sheet.cell(row_no, col_no).value == "main"
    template_header_main = sheet.cell(row_no, col_no + 1).value

    _validate_template(template_header_main, template_placeholders_required["main"], "main header")

    messages_templates.template_header_main = template_header_main[1:-1]

    row_no += 1
    assert sheet.cell(row_no, col_no).value == "step"
    template_header_step = sheet.cell(row_no, col_no + 1).value

    _validate_template(template_header_step, template_placeholders_required["step"], "step header")

    messages_templates.template_header_step = template_header_step[1:-1]

    # TODO Add automatically, dry code
    # templates_headers = {
    #     "main": template_header_main.strip('"'),
    #     "step": template_header_step.strip('"'),
    # }

    # return templates_headers


def surround_with_quotes(s):
    return surround_with_char(s, '"')


def surround_with_char(s, char):
    return char + str(s) + char


# TODO Extract automatically from the templates still in the code
placeholders_required = {
    "singles-1": ["dim", "char", "cell"],
    "singles-2": ["char", "dim", "dims"],
    "singles-3": ["char", "dim", "dims"],
    "singles-naked-2": ["dims", "char", "cell"],
    "singles-naked-3": ["dims", "char", "cell"],
    "doubles": ["cells", "chars_and", "dim", "chars_or"],
    "triplets": ["cells", "chars_and", "dim", "chars_or"],
    "quads": ["cells", "chars_and", "dim", "chars_or"],
    "singles-pointing": ["dim", "char", "size", "box", "dir", "num", "_position_s_", "cells"],
    "singles-boxed": ["dim", "char", "size", "box"],
    "x-wings": ["dims", "vals", "char", "size", "dims_help", "vals_help"],
    "x-wings-3": ["dims", "vals", "char", "size", "dims_help", "vals_help"],
    "x-wings-4": ["dims", "vals", "char", "size", "dims_help", "vals_help"],
    "ab-chains": ["cells", "char", "cell"],
    "remote-pairs": ["cells", "char_1", "char_2", "cell"],
    "y-wings": ["cells", "cell", "chars", "cells_wings", "char", "_cell_s_", "cells_removed"],
    "doubles-naked": ["cells", "chars", "dim"],
    "triplets-naked": ["cells", "size", "chars", "dim"],
    "quads-naked": ["cells", "size", "chars", "dim"],
    "boxed-doubles": ["cell", "box", "char_1", "cell_target", "char_2", "chars_target"],
    "boxed-triplets": ["num", "cells", "box", "chars", "cell_target", "chars_target"],
    "boxed-quads": ["num", "cells", "box", "chars", "cell_target", "chars_target"],
    "boxed-wings": ["cell_row", "row", "cell_col", "col", "box", "char", "cell"],
    "boxed-rays": ["char", "box_ray", "cells_hor", "cells_ver", "cells_ver_target", "box_target", "cells_hor_target", "num", "_cell_s_", "cells", "_cell_s_remove_", "cells_remove"],
    "ab-rings": (["cells", "chars"], ["cells", "dim", "char"], ),
    "leftovers-1": (["dims", "num", "_in_s_", "cells_in", "_out_s_", "cells_out", "_contain_s_", "_candidate_s_in_", "chars_in", "_candidate_s_out_", "chars_out"], [""], ["chars", "num", "_in_s_out_s_"], ),
    "leftovers-2": (["dims", "num", "_in_s_", "cells_in", "_out_s_", "cells_out", "_contain_s_", "_candidate_s_in_", "chars_in", "_candidate_s_out_", "chars_out"], [""], ["chars", "num", "_in_s_out_s_"], ),
    "leftovers-3": (["dims", "num", "_in_s_", "cells_in", "_out_s_", "cells_out", "_contain_s_", "_candidate_s_in_", "chars_in", "_candidate_s_out_", "chars_out"], [""], ["chars", "num", "_in_s_out_s_"], ),
    "leftovers-4": (["dims", "num", "_in_s_", "cells_in", "_out_s_", "cells_out", "_contain_s_", "_candidate_s_in_", "chars_in", "_candidate_s_out_", "chars_out"], [""], ["chars", "num", "_in_s_out_s_"], ),
    "leftovers-5": (["dims", "num", "_in_s_", "cells_in", "_out_s_", "cells_out", "_contain_s_", "_candidate_s_in_", "chars_in", "_candidate_s_out_", "chars_out"], [""], ["chars", "num", "_in_s_out_s_"], ),
    "leftovers-6": (["dims", "num", "_in_s_", "cells_in", "_out_s_", "cells_out", "_contain_s_", "_candidate_s_in_", "chars_in", "_candidate_s_out_", "chars_out"], [""], ["chars", "num", "_in_s_out_s_"], ),
    "leftovers-7": (["dims", "num", "_in_s_", "cells_in", "_out_s_", "cells_out", "_contain_s_", "_candidate_s_in_", "chars_in", "_candidate_s_out_", "chars_out"], [""], ["chars", "num", "_in_s_out_s_"], ),
    "leftovers-8": (["dims", "num", "_in_s_", "cells_in", "_out_s_", "cells_out", "_contain_s_", "_candidate_s_in_", "chars_in", "_candidate_s_out_", "chars_out"], [""], ["chars", "num", "_in_s_out_s_"], ),
    "leftovers-9": (["dims", "num", "_in_s_", "cells_in", "_out_s_", "cells_out", "_contain_s_", "_candidate_s_in_", "chars_in", "_candidate_s_out_", "chars_out"], [""], ["chars", "num", "_in_s_out_s_"], ),
}
# TODO Enable when all techniques are added
assert not set(TECHNIQUES).symmetric_difference(placeholders_required.keys())


def _read_templates_messages(sheet):

    row_no, col_no = 2, 1

    required_header = "MESSAGES SINGLES"
    required_keys = BASE_TECHNIQUES
    row_no = _read_message_templates(sheet, row_no, col_no, required_header, required_keys, messages_templates.templates_messages, required_header.split()[-1].lower())

    row_no += 2

    required_header = "MESSAGES ADVANCED"
    required_keys = messages_templates.templates_messages_advanced.keys()
    row_no = _read_message_templates(sheet, row_no, col_no, required_header, required_keys, messages_templates.templates_messages_advanced, required_header.split()[-1].lower())

    row_no += 2

    required_header = "MESSAGES FINAL"
    required_keys = ["single-value", "single-position"]
    collection = dict.fromkeys(required_keys, None)
    _read_keys_values(sheet, row_no, col_no, required_header, required_keys, collection)

    placeholders_required_final = {
        "single-value": ["char", "cell"],
        "single-position": ["char", "dim"],
    }

    # TODO Merge logic
    template_message = collection["single-value"]
    _validate_template(template_message, placeholders_required_final["single-value"], template_id="single-value")
    messages_templates.template_message_final_cell = template_message[1:-1]

    template_message = collection["single-position"]
    _validate_template(template_message, placeholders_required_final["single-position"], template_id="single-position")
    messages_templates.template_message_final_rcb = template_message[1:-1]

    return messages_templates


def _read_message_templates(sheet, row_no, col_no, required_header, required_keys, collection, template_group):

    assert sheet.cell(row_no, col_no).value == required_header

    template_values = {}

    for idx_row, row in enumerate(sheet.rows):
        if idx_row < row_no:
            continue

        key = row[col_no - 1].value
        val = row[col_no].value

        if not key:
            break

        row_no += 1

        # Note: Some templates are used for multiple techniques
        names_techniques = key.split('\n')
        # Note: Some templates consist of multiple components
        template_message_components = val.split('\n')

        print(f"Message template for {' and '.join(map(surround_with_quotes, names_techniques))}:")
        for template_message_component in template_message_components:
            print(f"  {template_message_component}")

        for name_technique in names_techniques:
            assert name_technique in collection, \
                f"Message template for technique not recognised: {surround_with_quotes(name_technique)}"

            required_num_components = 1 if isinstance(placeholders_required[name_technique], list) else len(placeholders_required[name_technique])
            print(required_num_components)
            assert len(template_message_components) == required_num_components, \
                f"Invalid number of message template components provided for {surround_with_quotes(name_technique)}"

            # TODO This can be done just once
            for component_no, template_message_component in enumerate(template_message_components):
                _placeholders_required = placeholders_required[name_technique] if isinstance(placeholders_required[name_technique], list) else placeholders_required[name_technique][component_no]
                _validate_template(template_message_component, _placeholders_required, surround_with_quotes(name_technique) + f" (component {component_no + 1})" * (len(template_message_components) > 1))

            # TODO Preferably store in a dict or create a container object
            template_message = tuple(
                map(
                    lambda template_message_component: template_message_component[1:-1],
                    template_message_components
                )
            )
            template_values[name_technique] = template_message[0] if len(template_message) == 1 else template_message

    # Check: A message template is defined for all techniques
    keys_missing = set(required_keys).difference(template_values)
    assert not keys_missing, \
        f"No message templates defined for {template_group} techniques: {sorted(keys_missing)}"

    for key, val in template_values.items():
        collection[key] = val

    return row_no


def _read_names(sheet):

    row_no, col_no = 2, 1

    assert sheet.cell(row_no, col_no).value == "NAMES"

    names_techniques_requiring_placeholder = \
        ["singles-pointing", "singles-boxed"] + [f"leftovers-{i}" for i in range(1, 9 + 1)]

    placeholder_required = 'PLACEHOLDER'

    # First collect all values to perform some checks before overwriting
    template_values = {}
    target_collection = messages_templates.MAP_TECHNIQUE_NAMES

    for idx_row, row in enumerate(sheet.rows):
        if idx_row < row_no:
            continue

        name_technique = row[col_no - 1].value
        title = row[col_no].value
        difficulty = row[col_no + 1].value

        # assert name_technique in names_techniques_implemented, \
        #     f"This version of the tool is not yet able to handle custom name for technique \"{name_technique}\""

        # Process
        title = title.strip()
        # difficulty = difficulty.capitalize()

        # Checks
        assert name_technique in target_collection, \
            f"Name for technique not recognised: {surround_with_quotes(name_technique)}"

        placeholders = re.findall(r"{(.*?)}", title)
        if name_technique in names_techniques_requiring_placeholder:
            assert placeholders == [placeholder_required], \
                f"Name for technique {surround_with_quotes(name_technique)} should contain placeholder {{{placeholder_required}}}"
        else:
            assert not placeholders, \
                f"Name for technique {surround_with_quotes(name_technique)} should not contain any placeholders"

        assert difficulty in messages_templates.MAP_RATINGS, \
            f"Rating for technique {surround_with_quotes(name_technique)} not recognised: {difficulty}"

        print(f"Technique name for \"{name_technique}\":")
        print("  " + f"{title} ({difficulty})")

        assert name_technique not in template_values, \
            f"Name for technique {surround_with_quotes(name_technique)} defined multiple times"
        template_values[name_technique] = (title, difficulty)

    # Check: A name is defined for all techniques
    missing_names_techniques = set(TECHNIQUES).difference(template_values)
    assert not missing_names_techniques, \
        f"Names not defined for all techniques: {sorted(missing_names_techniques)}"

    for key, value in template_values.items():
        target_collection[key] = value


def _read_keywords(sheet):

    row_no, col_no = 2, 1

    assert sheet.cell(row_no, col_no).value == "DIMENSIONS"

    for dim in DIMENSIONS:

        row_no += 1

        cell_key = sheet.cell(row_no, col_no)
        key = cell_key.value
        val_1 = sheet.cell(row_no, col_no + 1).value
        val_2 = sheet.cell(row_no, col_no + 2).value

        assert key == (key_assert := dim), \
            f"Structure of template incorrect, expected {surround_with_quotes(key_assert)}" \
            f" in sheet {surround_with_quotes(sheet.title)} and cell {cell_key.coordinate}"
        assert val_1, f"No value provided for {dim}"
        assert val_2, f"No value provided for {dim} (plural)"

        messages_templates.MAP_DIMENSIONS[dim] = (val_1, val_2)

    # TODO Read cell somewhere
    # row_no += 1

    row_no += 2

    # TODONE This can easily be generalised, by parametrising further the collection where to write the value to
    required_header = "DIRECTIONS"
    required_keys = ["hor", "ver"]
    _read_keys_values(sheet, row_no, col_no, required_header, required_keys, messages_templates.MAP_DIRECTIONS)

    row_no += 2 + len(required_keys)

    required_header = "RATINGS"
    required_keys = ["Easy", "Medium", "Hard"]
    _read_keys_values(sheet, row_no, col_no, required_header, required_keys, messages_templates.MAP_RATINGS)

    row_no += 2 + len(required_keys)

    required_header = "COUNTS"
    required_keys = list(range(1, 9 + 1))
    _read_keys_values(sheet, row_no, col_no, required_header, required_keys, messages_templates.MAP_COUNTS)

    row_no += 2 + len(required_keys)

    required_header = "SIZES"
    required_keys = list(range(1, 9 + 1))
    _read_keys_values(sheet, row_no, col_no, required_header, required_keys, messages_templates.MAP_SIZE_TO_NAME)

    row_no += 2 + len(required_keys)

    required_header = "CONJUNCTIONS"
    required_keys = ["and", "or"]
    _read_keys_values(sheet, row_no, col_no, required_header, required_keys, messages_templates.MAP_CONJUNCTIONS)

    row_no += 2 + len(required_keys)

    required_header = "PLURALS"
    required_keys = ["cell", "position", "candidate", "contain", "in", "out"]
    _read_keys_values(sheet, row_no, col_no, required_header, required_keys, messages_templates.MAP_PLURALS, is_tuple=True)


def _read_keys_values(sheet, row_no, col_no, required_header, required_keys, collection, is_tuple=False):

    assert (cell := sheet.cell(row_no, col_no)).value == (val := required_header), \
        f"Structure of template incorrect, expected header {surround_with_quotes(val)}" \
        f" in sheet {surround_with_quotes(sheet.title)} and cell {cell.coordinate}"

    template_values = {}

    for required_key in required_keys:

        row_no += 1

        cell_key = sheet.cell(row_no, col_no)
        key = cell_key.value
        if not is_tuple:
            val = sheet.cell(row_no, col_no + 1).value
        else:
            val = tuple(sheet.cell(row_no, col_no + 1 + i).value for i in range(2))
            assert all(e for e in val), \
                f"Not all values defined in {surround_with_quotes(required_header)} for {surround_with_quotes(key)}"
        assert key == required_key, \
            f"Structure of template incorrect, expected {surround_with_quotes(required_key)}" \
            f" in sheet {surround_with_quotes(sheet.title)} and cell {cell_key.coordinate}"

        template_values[key] = val

    # Check all values are present, instead of relying on the defined list of required keys
    missing_keys = set(collection.keys()).difference(template_values)
    assert not missing_keys, \
        f"No values defined in {surround_with_quotes(required_header)} for required keys: {sorted(missing_keys)}"

    for key, val in template_values.items():
        collection[key] = val


def _validate_template(template, placeholders_required, template_id):

    try:

        # Surrounded by quotation marks
        assert template[0] == '"' and template[-1] == '"', \
            f"Template for {template_id} should be surrounded by quotation (\"\") marks"

        placeholders = re.findall(r"{(.*?)}", template)

        placeholders_missing = set(placeholders_required).difference(placeholders)
        assert not placeholders_missing, \
            f"Template for {template_id} is missing required placeholders: {sorted(placeholders_missing)}"

        placeholders_unused = set(placeholders).difference(placeholders_required)
        assert not placeholders_unused, \
            f"Template for {template_id} contains unused placeholders: {sorted(placeholders_unused)}"

    except AssertionError as e:
        raise e


if __name__ == "__main__":

    file_name_template = "Template_Messages.xlsx"

    template = read_template(file_name_template)

    print(template)
