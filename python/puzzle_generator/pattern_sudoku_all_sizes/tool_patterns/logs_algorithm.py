
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from generator.model import EMPTY_CHAR


# Similar to the logs in tool_create, these logs aim to outline the steps of the algorithm used to generate instances
#  based on a specified pattern;


IDX_ROW_START = 2
IDX_COL_START = 2


def _read_template(file_name_template, size):
    """
    Read template file and extract relevant formatting
    """

    # TODO Implement logic for different sizes
    assert size == 9, f"No template reading logic defined for grid size {size}"

    # Read template
    workbook = load_workbook(file_name_template)
    sheet = workbook.worksheets[0]

    # Extract formatting
    style_header = sheet.cell(IDX_ROW_START, IDX_COL_START)._style
    style_cells = [
        [
            sheet.cell(IDX_ROW_START + 2 + i1, IDX_COL_START + i2)._style
            for i2 in range(size)
        ]
        for i1 in range(size)
    ]
    style_message = sheet.cell(IDX_ROW_START + 2 + size + 1, IDX_COL_START)._style

    formatting_instance = (style_header, style_cells, style_message)

    return workbook, sheet, formatting_instance


# TODO Perhaps create a separate "formatting" package, which centralises all logic related to converting to user-
#  understandable output; For now this is the only function copied from tool_logs, but more might follow;
def convert_to_user_readable_value(i1, i2):
    """
    Converts internally used 0-based coords to user-understandable output
    """
    # TODO Implement logic for other dimensions when needed
    idx = "R{}C{}".format(i1 + 1, i2 + 1)
    return idx


def write_step_to_sheet(sheet, idx_row, idx_col, pattern, chars, step_details, header, message, formatting_instance):

    instance, idx_to_fill, idxs_propagated, cell_options, is_valid_propagation, highlight_header = step_details

    style_header, style_cells, style_message = formatting_instance

    size = len(instance)
    assert len(chars) == size
    sorted_chars = sorted(chars)

    # Write header
    idxs_header = (idx_row, idx_col)
    cell_header = sheet.cell(*idxs_header)
    cell_header.value = header
    cell_header._style = copy(style_header)
    if highlight_header:
        cell_header.fill = PatternFill("solid", start_color="FF5959")
    sheet.merge_cells(
        start_row=idxs_header[0], start_column=idxs_header[1],
        end_row=idxs_header[0], end_column=idxs_header[1] + size - 1
    )

    # Write cells
    # TODO Add color codes to template
    for i1 in range(size):
        for i2 in range(size):
            char = instance[i1][i2]

            value = char
            use_font_options = False

            if cell_options is not None:
                options_remaining = cell_options[(i1, i2)]
                if char != EMPTY_CHAR:
                    assert not options_remaining
                else:
                    # Show the possible remaining options instead
                    # Check: The number of remaining options for each cell should be larger than 2, or it should have
                    #  been found with propagation, unless the propagation was invalid
                    assert is_valid_propagation is not None  # Only to check the logs logic is consistent
                    assert not is_valid_propagation or len(options_remaining) >= 2
                    assert not set(options_remaining).difference(chars)
                    str_options_remaining = [_char if _char in options_remaining else ' ' for _char in sorted_chars]
                    value = '\n'.join(' '.join(str_options_remaining[3 * i:3 * (i + 1)]) for i in range(3))
                    use_font_options = True

            cell = sheet.cell(idx_row + 2 + i1, idx_col + i2)
            cell.value = value
            cell._style = copy(style_cells[i1][i2])

            if use_font_options:
                cell.font = cell.font.copy(size=cell.font.size * 0.412, color="BFBFBF", name="Consolas")

            # Highlight pattern (highlighted for every step)
            if pattern[i1][i2]:
                cell.fill = PatternFill("solid", start_color="EAEAEA")

            # Highlight propagated single-option idxs (note: this contains the idx targeted to be filled in, but this
            #  is overwritten with a custom color in the next check)
            if idxs_propagated is not None and (i1, i2) in idxs_propagated:
                cell.fill = PatternFill("solid", start_color="FFB465")

            # Highlight idx attempted to be filled in
            if (i1, i2) == idx_to_fill:
                cell.fill = PatternFill("solid", start_color="89B2FF")

            # Highlight invalid propagated values
            # Note: We make use of the logic in the algorithm which terminates after finding an invalid propagation,
            #  where the idx is still added to the list
            if idxs_propagated is not None and (not is_valid_propagation and (i1, i2) in idxs_propagated and idxs_propagated.index((i1, i2)) == len(idxs_propagated) - 1):
                cell.fill = PatternFill("solid", start_color="FF7878")

            # Highlight empty cells without remaining options (which also makes the propagation invalid)
            if char == EMPTY_CHAR and (cell_options is not None and not cell_options[(i1, i2)]):
                cell.fill = PatternFill("solid", start_color="FF7878")

    # Write message
    if message is not None:
        idxs_message = (idx_row + 2 + size + 1, idx_col)
        cell_message = sheet.cell(*idxs_message)
        cell_message.value = message
        cell_message._style = copy(style_message)
        sheet.merge_cells(
            start_row=idxs_message[0], start_column=idxs_message[1],
            end_row=idxs_message[0], end_column=idxs_message[1] + size - 1
        )


def write_line_to_file(sheet, idx_row, idx_col, idxs_written_instances, idx_col_start_steps, num_cols_per_step, size):

    # Extra: Add a line if there are gaps in the row
    idxs_same_row = [(_idx_row, _idx_col) for (_idx_row, _idx_col) in idxs_written_instances if _idx_row == idx_row]
    idxs_to_left = [(_idx_row, _idx_col) for (_idx_row, _idx_col) in idxs_same_row if _idx_col < idx_col]
    if idxs_to_left:
        idx_col_max = max(_idx_col for (_idx_row, _idx_col) in idxs_to_left)
        idxs_col_line = range(idx_col_max + num_cols_per_step, idx_col - 1)
    else:
        idxs_col_line = range(idx_col_start_steps, idx_col - 1)
    idx_row_line = idx_row + 2 + size // 2
    for idx_col_line in idxs_col_line:
        idx_line = (idx_row_line, idx_col_line)
        # print(f"Draw line segment at {idx_line}")
        cell = sheet.cell(*idx_line)
        cell.value = "---"


def write_logs_to_file(pattern, chars, final_instance_details, logs, file_name_template, file_name_output):

    # TODO The size could be read from the template in some future version
    size = len(pattern)
    assert size == 9, f"Writing algorithm logs to file not implemented for size {size}"  # TODO Implement for other sizes

    print("Reading template..")

    # Load template
    workbook, sheet, formatting_instance = _read_template(file_name_template, size)

    print("Writing to file..")

    # Unmerge cells
    # Note: This has to be done before removing rows, as otherwise they will not be present in the sheet and cannot be
    #  unmerged
    # TODO Only unmerge from the row where the sheet is cleaned
    for merged_cell_range in sheet.merged_cell_ranges:
        sheet.unmerge_cells(range_string=str(merged_cell_range))

    # Clean sheet
    idx_row_start_clean = 1
    sheet.delete_rows(idx_row_start_clean, sheet.max_row - idx_row_start_clean + 1)

    # Pre-processing
    empty_instance = [[EMPTY_CHAR] * size] * size
    pattern_number_hints = sum(map(sum, pattern))

    # Write pattern
    header = "PATTERN"
    message = None
    step_details = (empty_instance, None, None, None, None, False)
    write_step_to_sheet(
        sheet, IDX_ROW_START, IDX_COL_START, pattern, chars, step_details, header, message, formatting_instance
    )

    num_rows_per_step = 2 + size + 2 + 1
    num_cols_per_step = size + 1
    idx_row_start_steps = IDX_ROW_START + num_rows_per_step

    # Write steps for pattern
    idxs_pattern = [((i1, i2)) for i1 in range(size) for i2 in range(size) if pattern[i1][i2]]
    assert len(idxs_pattern) == pattern_number_hints
    for i in range(pattern_number_hints):
        idx_row = idx_row_start_steps + i * num_rows_per_step
        idx_col = 2

        idx_to_fill = idxs_pattern[i]

        header = "STEP {} / {}".format(i + 1, pattern_number_hints)
        message = None
        step_details = (empty_instance, idx_to_fill, None, None, None, False)
        write_step_to_sheet(
            sheet, idx_row, idx_col, pattern, chars, step_details, header, message, formatting_instance
        )

    # Preprocess logs
    logs_steps = logs["steps"]

    # TODO Include some check for too large logs

    # Write logs
    idx_col_start_steps = 2 + size + 3
    prev_step_no = -1

    # Options
    WRITE_SKIPPED_STEPS = True
    WRITE_LINES = True

    # For checking purposes
    idxs_written_instances = []
    prev_is_valid_propagation = True

    idx_col = idx_col_start_steps
    for idx_step, logs_step in enumerate(logs_steps):

        # Pre-process step logs
        idx_to_fill, idxs_propagated, filled_grid, is_valid_propagation, filled_value_details, cell_options = logs_step

        filled_value_counter, num_possible_values = filled_value_details

        assert idx_to_fill in idxs_pattern
        step_no = idxs_pattern.index(idx_to_fill)
        idx_row = idx_row_start_steps + step_no * num_rows_per_step
        if step_no <= prev_step_no:
            # We went a step back in the recursion, or repeated the same level
            idx_col += num_cols_per_step
            assert not prev_is_valid_propagation

        # Extra: Fill in placeholder steps when the value was already filled in using propagation
        # TODO This can also be added to the steps in the algorithm, and would not require this rather ugly custom logic
        if WRITE_SKIPPED_STEPS:

            if idx_step < len(logs_steps) - 1:
                next_idx_to_fill = logs_steps[idx_step + 1][0]
                next_step_no = idxs_pattern.index(next_idx_to_fill)
                skipped_number_steps = max(0, next_step_no - step_no - 1)
            else:
                skipped_number_steps = pattern_number_hints - step_no - 1

            for skip_no in range(skipped_number_steps):
                skipped_step_no = step_no + (skip_no + 1)
                skipped_idx_row = idx_row_start_steps + skipped_step_no * num_rows_per_step
                # Note: The col_idx is the same
                skipped_idx_to_fill = idxs_pattern[skipped_step_no]
                skipped_filled_grid = filled_grid
                skipped_cell_options = cell_options

                assert (skipped_idx_row, idx_col) not in idxs_written_instances
                idxs_written_instances.append((skipped_idx_row, idx_col))

                header = "ALREADY FILLED IN '{}' AT {}".format(  # WITH PROPAGATION
                    skipped_filled_grid[skipped_idx_to_fill[0]][skipped_idx_to_fill[1]], convert_to_user_readable_value(*skipped_idx_to_fill)
                )
                message = None

                step_details = (skipped_filled_grid, skipped_idx_to_fill, None, skipped_cell_options, True, False)
                write_step_to_sheet(
                    sheet, skipped_idx_row, idx_col, pattern, chars, step_details, header, message, formatting_instance
                )

                if WRITE_LINES:
                    write_line_to_file(
                        sheet, skipped_idx_row, idx_col, idxs_written_instances, idx_col_start_steps, num_cols_per_step, size
                    )

        prev_step_no = step_no
        prev_is_valid_propagation = is_valid_propagation

        assert (idx_row, idx_col) not in idxs_written_instances
        idxs_written_instances.append((idx_row, idx_col))

        # Generate header and message
        header = "ATTEMPT TO FILL IN '{}' at {} [{}/{}]".format(
            filled_grid[idx_to_fill[0]][idx_to_fill[1]], convert_to_user_readable_value(*idx_to_fill),
            filled_value_counter, num_possible_values,
        )
        message = None

        is_last_option = filled_value_counter == num_possible_values
        step_details = (filled_grid, idx_to_fill, idxs_propagated, cell_options, is_valid_propagation, not is_valid_propagation and is_last_option)
        write_step_to_sheet(
            sheet, idx_row, idx_col, pattern, chars, step_details, header, message, formatting_instance
        )

        if WRITE_LINES:
            write_line_to_file(
                sheet, idx_row, idx_col, idxs_written_instances, idx_col_start_steps, num_cols_per_step, size
            )

    # Write final instance
    idx_row_final_instance = idx_row_start_steps + pattern_number_hints * num_rows_per_step

    final_instance, is_solvable_using_human_techniques, counts_techniques, weight, number_non_empty_cells = final_instance_details

    header = "INSTANCE"
    message = '\n'.join([
        f"Solvable: {is_solvable_using_human_techniques}",
        f"Hints: {number_non_empty_cells}",
        f"Techniques: {counts_techniques}",
        f"Weight: {weight}",
    ])

    step_details = (final_instance, None, None, None, None, False)
    write_step_to_sheet(
        sheet, idx_row_final_instance, idx_col, pattern, chars, step_details, header, message, formatting_instance
    )

    # TODO Write final instance characteristics to file

    # Modify the sheet name
    sheet.title = "Steps"

    # Write output to file
    workbook.save(filename=file_name_output)
