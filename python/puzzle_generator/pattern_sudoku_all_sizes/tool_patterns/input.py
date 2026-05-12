
from openpyxl.reader.excel import load_workbook
import operator

from generator.boxes import preprocess_layout_boxes
from generator.model import Instance


def read_pattern_from_file(file_name_pattern, size):
    """
    Pattern data structure:
      Non-empty (1) - Keep value
      Empty - Remove value
    """

    workbook = load_workbook(file_name_pattern)
    sheet = workbook.active

    data = [
        [
            sheet.cell(i1 + 1, i2 + 1).value
            for i2 in range(size)
        ]
        for i1 in range(size)
    ]

    # Data validation check
    assert all(e in [1, None] for row in data for e in row), "Data structure of pattern incorrect, cells should be empty or 1"

    # Convert to internal data structure
    pattern = [
        [e == 1 for e in row]
        for row in data
    ]

    return pattern


def read_list_of_patterns_from_excel_file(file_name_pattern, check_for_duplicate_patterns=True):
    """
    Steps:
     - Read raw data
     - Group data by row ID
     - Validate and instantiate patterns
     - [optional] Screen for duplicates
    """

    # TODO Make this a variable
    SIZE = 9

    print("Reading list of patterns..")

    workbook = load_workbook(file_name_pattern)
    sheet = workbook.worksheets[0]

    SKIP_NUMBER_ROWS = 2

    SHOW_PROGRESS_INCREMENT = 10

    # Reading data
    print("Reading data..")
    max_row = sheet.max_row  # Make sure to do this once outside the loop as this is an expensive operation
    print(" Number of rows:", max_row)
    rows = []
    last_progress_shown = 0
    for idx_row in range(SKIP_NUMBER_ROWS, max_row):
        pattern_id = sheet.cell(1 + idx_row, 1).value
        line_chars = [sheet.cell(1 + idx_row, 1 + 1 + i2).value for i2 in range(SIZE)]
        line_layout = [sheet.cell(1 + idx_row, 1 + 1 + SIZE + i2).value for i2 in range(SIZE)]
        rows.append((pattern_id, (line_chars, line_layout)))
        progress = (idx_row + 1) / (max_row - SKIP_NUMBER_ROWS) * 100
        if progress - last_progress_shown >= SHOW_PROGRESS_INCREMENT:
            last_progress_shown = progress - (progress % SHOW_PROGRESS_INCREMENT)
            print("Progress: {} %".format(last_progress_shown))

    # Group lines
    print("Processing data..")
    grouped_rows = []
    cur_pattern_id = None
    cur_pattern_lines = []
    for pattern_id, line in rows:
        if pattern_id == cur_pattern_id:
            cur_pattern_lines.append(line)
        else:
            if cur_pattern_id is not None:
                grouped_rows.append((cur_pattern_id, cur_pattern_lines))
            cur_pattern_id = pattern_id
            cur_pattern_lines = [line]
    if cur_pattern_id is not None:
        grouped_rows.append((cur_pattern_id, cur_pattern_lines))

    def reconstruct_pattern(lines):
        assert len(lines) == SIZE, f"Number of rows incorrect: {len(lines)}"

        # Note: Optionally, a custom boxes layout can be specified
        lines_chars, lines_layout = (list(map(operator.itemgetter(i), lines)) for i in (0, 1))

        assert all(e in (None, 1) for row in lines_chars for e in row), \
            "Data structure of pattern incorrect, cells should be empty or 1"

        # Instantiate (convert to internal data structure)
        data = [
            [e == 1 for e in row]
            for row in lines_chars
        ]

        if all(e is None for row in lines_layout for e in row):
            # No layout specified
            pattern = Instance(data, is_chars=False)
        else:
            layout_boxes = preprocess_layout_boxes(lines_layout, SIZE)
            pattern = Instance(data, is_chars=False, layout_boxes=layout_boxes)

        return pattern

    # Reconstruct patterns
    patterns = []
    for pattern_id, lines in grouped_rows:
        try:
            pattern = reconstruct_pattern(lines)
            patterns.append((pattern_id, pattern))
        except Exception as e:
            print(f"Could not process pattern with ID '{pattern_id}': {e}")
            for line in lines:
                print(line)
            continue

    # Identify duplicate patterns
    if check_for_duplicate_patterns:
        _screen_for_duplicate_patterns(patterns)

    return patterns


def _screen_for_duplicate_patterns(patterns):

    print("Screening for duplicate patterns..")

    groups = {}
    for idx, (pattern_id, pattern) in enumerate(patterns):
        key = str(pattern)
        if key in groups:
            groups[key].append(pattern_id)
        else:
            groups[key] = [pattern_id]

        if idx > 0 and idx % 100 == 0:
            print(f"Progress: {idx} / {len(patterns)}")

    groups_with_duplicate_patterns = [
        pattern_ids
        for key, pattern_ids in groups.items()
        if len(pattern_ids) > 1
    ]

    if len(groups_with_duplicate_patterns) > 0:
        print("WARNING - Identified duplicate patterns:")
        for group in groups_with_duplicate_patterns:
            print(len(group), sorted(group))


def read_layout_boxes(file_name_layout_boxes, size):

    workbook = load_workbook(file_name_layout_boxes)
    sheet = workbook.active

    raw_data = [
        [
            sheet.cell(i1 + 1, i2 + 1).value
            for i2 in range(size)
        ]
        for i1 in range(size)
    ]

    layout_boxes = preprocess_layout_boxes(raw_data, size)

    return layout_boxes
