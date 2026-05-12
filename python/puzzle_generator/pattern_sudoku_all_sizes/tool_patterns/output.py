
import itertools
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN
import sys

from generator.model import count_non_empty_cells


# TODO The information in subwords_not_included can added to subwords_placements by using a None position
def write_to_excel_file(
        output_file_name, pattern, solution, instance,
        is_solvable_using_human_techniques, counts_techniques, weight, number_non_empty_cells):
    """
    The output logic is different for each tool, as is also the case for the tool using patterns to create instances.

    The output written to file for this tool:
      - The solution of the fitting phase
      - The pattern applied
      - The resulting instance

    As there is no mainword (and so the diagonal is not removed), the diagonal formatting logic can be skipped.
    The layout is also irrelevant and does not need to be written to file.
    """

    workbook = Workbook()
    sheet = workbook.active

    font_default = Font(name="calibri", size=12)
    font_highlighted = Font(name="calibri", size=12, color="FF7F4F")

    size = len(solution)

    # highlights = determine_highlights(size, subwords_placements)

    row = 1

    # Write the size to file, which is now necessary as we use the same code for all grid sizes
    sheet.cell(row, 1).value = "Size"
    sheet.cell(row, 2).value = str(size)

    row += 1

    # Optionally: Write default boxes layout to file
    if instance.layout is not None:

        sheet.cell(row, 1).value = "Layout"
        sheet.cell(row, 2).value = instance.layout

        row += 1

    row += 1

    # Write solution to file, where subwords are highlighted with a separate color
    sheet.cell(row, 1).value = "Solution"

    row += 1

    for i1 in range(size):
        for i2 in range(size):
            text = str(solution[i1][i2])
            # highlight = highlights[i1][i2]

            # if highlight:
            #     font = font_highlighted
            # else:
            font = font_default

            cell = sheet.cell(row + i1, i2 + 1)
            cell.value = text
            cell.font = font

    # Draw borders for solution
    draw_borders(sheet, (row, 1), instance)

    # Optionally: Write custom boxes layout to file
    if instance.uses_custom_boxes_layout:

        for (i1, i2) in itertools.product(range(size), repeat=2):
            idx_box = instance.get_idx_box(i1, i2)
            box_id = str(idx_box + 1)  # Convert internal 0-based to user-readable 1-based

            cell = sheet.cell(row + i1, size + 2 + i2)
            cell.value = box_id

    row += size + 2

    # # Specify the position of the subwords included in the solution
    # sheet[chr(ord("A") + 0) + str(row)] = "Subwords included"
    # for (subword, rotation, i1, i2) in subwords_placements:
    #     row += 1
    #     sheet[chr(ord("A") + 0) + str(row)] = subword
    #     sheet[chr(ord("A") + 1) + str(row)] = rotation
    #     sheet[chr(ord("A") + 2) + str(row)] = str((i1, i2))
    #
    # row += 2
    #
    # # Indicate which subwords could not be included in the solution
    # sheet[chr(ord("A") + 0) + str(row)] = "Subwords not included"
    # for subword in subwords_not_included:
    #     row += 1
    #     sheet[chr(ord("A") + 0) + str(row)] = subword
    #
    # row += 2

    # Show the pattern applied
    sheet.cell(row, 1).value = "Pattern"

    row += 1

    for i1 in range(size):
        for i2 in range(size):
            text = "x" if pattern[i1][i2] else ""

            cell = sheet.cell(row + i1, i2 + 1)
            cell.value = text

    row += size + 2

    # Write instance to file, which is the result of the solution after applying the pattern
    sheet.cell(row, 1).value = "Instance"

    row += 1

    for i1 in range(size):
        for i2 in range(size):
            text = str(instance[i1][i2])

            cell = sheet.cell(row + i1, i2 + 1)
            cell.value = text

    # Draw borders for instance
    draw_borders(sheet, (row, 1), instance)

    row += size + 2

    # Extra information on the instance, as reported in the tool_create_x output file
    sheet.cell(row, 1).value = "Solvable with human techniques"
    sheet.cell(row, 2).value = str(is_solvable_using_human_techniques)

    row += 1

    if is_solvable_using_human_techniques:

        sheet.cell(row, 1).value = "Hints"
        sheet.cell(row, 2).value = str(number_non_empty_cells)

        row += 1

        sheet.cell(row, 1).value = "Techniques"
        sheet.cell(row, 2).value = str(counts_techniques)

        row += 1

        sheet.cell(row, 1).value = "Weight"
        sheet.cell(row, 2).value = str(weight)

        row += 1

    workbook.save(filename=output_file_name)


def draw_borders(sheet, idx_start, instance):

    size = instance.size

    for (i1, i2) in itertools.product(range(size), repeat=2):
        idx_box = instance.get_idx_box(i1, i2)

        draw_left_border = (instance.get_idx_box(i1, i2 - 1) != idx_box) if i2 - 1 >= 0 else True
        draw_right_border = (instance.get_idx_box(i1, i2 + 1) != idx_box) if i2 + 1 < size else True
        draw_top_border = (instance.get_idx_box(i1 - 1, i2) != idx_box) if i1 - 1 >= 0 else True
        draw_bottom_border = (instance.get_idx_box(i1 + 1, i2) != idx_box) if i1 + 1 < size else True

        cell = sheet.cell(idx_start[0] + i1, idx_start[1] + i2)
        cell.border = Border(
            left=Side(border_style=BORDER_THIN) if draw_left_border else cell.border.left,
            right=Side(border_style=BORDER_THIN) if draw_right_border else cell.border.right,
            top=Side(border_style=BORDER_THIN) if draw_top_border else cell.border.top,
            bottom=Side(border_style=BORDER_THIN) if draw_bottom_border else cell.border.bottom,
        )


def _write_instance_to_sheet(sheet, idx_row, idx_col, instance, convert_to_text=True):
    # Note: Idxs are 1-based!

    # Note: Cannot access instance.size here as the pattern is a plain list of lists
    for i1, row in enumerate(instance):
        for i2, e in enumerate(row):
            text = str(e) if convert_to_text else e
            cell = sheet.cell(idx_row + i1, idx_col + i2)
            cell.value = text


def write_list_of_patterns_to_excel_file(file_name_pattern, file_name_output, patterns, instances_and_solutions):

    # TODO Make this a variable
    SIZE = 9

    # Reuse the formatting of the input file
    workbook = load_workbook(file_name_pattern)
    sheet = workbook.worksheets[0]

    idx_first_empty_row = 3

    max_row = sheet.max_row
    sheet.delete_rows(idx_first_empty_row, max_row)

    assert len(patterns) == len(instances_and_solutions)  # Also already done in main file

    idx_row = idx_first_empty_row
    for pattern_tuple, instance_and_solution_tuple in zip(patterns, instances_and_solutions):

        # Unpack pattern tuple
        (pattern_id, pattern) = pattern_tuple

        idx_col = 1

        # Write pattern ID
        sheet.cell(idx_row, idx_col).value = pattern_id

        idx_col += 1

        # Convert back to the data structure used by the input file to render the pattern
        pattern_converted = [
            [int(e) if e else "" for e in row]
            for row in pattern
        ]

        # Write pattern
        _write_instance_to_sheet(sheet, idx_row, idx_col, pattern_converted, convert_to_text=False)

        idx_col += SIZE + 1  # Insert one empty col

        if instance_and_solution_tuple is None:

            # Write apologising message
            sheet.cell(idx_row, idx_col).value = "NO INSTANCE"

        else:

            # Unpack instance and solution tuple
            (solution, instance, is_solvable_using_human_techniques, counts_techniques, weight, number_non_empty_cells) = instance_and_solution_tuple

            # Write header for solution
            for i2 in range(SIZE):
                cell = sheet.cell(idx_first_empty_row - 1, idx_col + i2)
                cell.value = "Col {}".format(i2 + 1)

            # Write solution
            _write_instance_to_sheet(sheet, idx_row, idx_col, solution)

            idx_col += SIZE + 1  # Insert one empty col

            # Write header for custom boxes layout
            for i2 in range(SIZE):
                cell = sheet.cell(idx_first_empty_row - 1, idx_col + i2)
                cell.value = "Col {}".format(i2 + 1)

            # Optionally: Write custom boxes layout
            if solution.uses_custom_boxes_layout:

                # TODO Reuse logic of function writing instance to file
                for (i1, i2) in itertools.product(range(solution.size), repeat=2):
                    idx_box = solution.get_idx_box(i1, i2)
                    box_id = idx_box + 1  # Convert internal 0-based to user 1-based

                    cell = sheet.cell(idx_row + i1, idx_col + i2)
                    cell.value = str(box_id)

            idx_col += SIZE + 1  # Insert empty col

            # Write header for instance
            for i2 in range(SIZE):
                cell = sheet.cell(idx_first_empty_row - 1, idx_col + i2)
                cell.value = "Col {}".format(i2 + 1)

            # Write instance
            _write_instance_to_sheet(sheet, idx_row, idx_col, instance)

            idx_col += SIZE + 1  # Insert one empty col

            # Write instance characteristics
            sheet.cell(idx_row + 0, idx_col + 0).value = "Solvable with human techniques"
            sheet.cell(idx_row + 0, idx_col + 1).value = str(is_solvable_using_human_techniques)

            # If the instance is not solvable using human techniques, the remaining characteristics do not make sense
            if is_solvable_using_human_techniques:

                sheet.cell(idx_row + 1, idx_col + 0).value = "Hints"
                sheet.cell(idx_row + 1, idx_col + 1).value = str(number_non_empty_cells)

                sheet.cell(idx_row + 2, idx_col + 0).value = "Techniques"
                sheet.cell(idx_row + 2, idx_col + 1).value = str(counts_techniques)

                sheet.cell(idx_row + 3, idx_col + 0).value = "Weight"
                sheet.cell(idx_row + 3, idx_col + 1).value = str(weight)

        # Update row idx
        idx_row += SIZE

    workbook.save(filename=file_name_output)


def write_instances_to_file(output_file_name, all_instances):

    # TODO Generalise this functionality, as it is also used in logs.py

    print(f"Write all instances to file:", output_file_name)
    output_file = open(output_file_name, "w")

    stdout_orig = sys.stdout
    sys.stdout = output_file

    for idx, _instance in enumerate(all_instances):
        instance, counts_techniques, weight = _instance

        # Recalculate stats
        number_non_empty_cells = count_non_empty_cells(instance)

        print(f"Instance {idx + 1}")

        print()
        print(instance)

        print(" - Number of hints:", number_non_empty_cells)
        print(" - Weight:", weight)
        print(" - Techniques used:", counts_techniques)
        print()
        print()

    output_file.close()
    sys.stdout = stdout_orig

    print("Finished")


def write_debug_info_to_file(debug_info, session_id):

    output_file_name = f"debug_info_{session_id}.txt"

    print("Writing debug info to file:", output_file_name)

    with open(output_file_name, "w") as file:
        output = json.dumps(debug_info, indent=2)
        file.write(output)

    print("Finished")
