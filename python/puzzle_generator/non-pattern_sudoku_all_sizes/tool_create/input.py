
import contextlib
from openpyxl.reader.excel import load_workbook
import operator
import re

from generator.model import Instance, SIZES_REQUIRING_STANDARD_LAYOUT
from generator.verification import verify_base_constraints
from generator.boxes import preprocess_layout_boxes

from tool_create.positioner import get_idxs_for_subword_placement


def read_solution_from_excel_file(file_name_input):
    """
    Read a single solution from file

    Note: Both the old format without custom boxes layout and the new format with custom boxes layout is supported
    """

    print("Reading input solution..")

    workbook = load_workbook(file_name_input)
    sheet = workbook.active

    print(" reading size/layout..")

    # Since we are now using the same code for all grid sizes, and we do not write the grid size to file, we have to
    #  derive it from the data - we assume the first information after the solution grid is always the same
    idx_row = 1
    while sheet.cell(idx_row, 1).value != "Subwords included":
        idx_row += 1

    size = idx_row - 2

    print(f"  size: {size}")

    print(" reading grid..")

    solution = [
        [sheet.cell(i1 + 1, i2 + 1).value for i2 in range(size)]
        for i1 in range(size)
    ]

    print(" reading custom layout..")

    # Optional for new format: Read custom boxes layout
    if sheet.cell(1, size + 2).value:

        raw_data = [
            [
                sheet.cell(i1 + 1, size + 1 + i2 + 1).value
                for i2 in range(size)
            ]
            for i1 in range(size)
        ]

        layout_boxes = preprocess_layout_boxes(raw_data, size)

    else:
        layout_boxes = None

        print("  not specified")

    idx_row = size + 2

    print(" reading subwords..")
    assert sheet.cell(idx_row, 1).value == "Subwords included", "Structure of input file is not correct"

    subwords_placements = []
    while True:
        idx_row += 1
        text = sheet.cell(idx_row, 1).value
        if not text:
            break
        else:
            subword = text
            orientation = sheet.cell(idx_row, 2).value
            idxs = sheet.cell(idx_row, 3).value
            subwords_placements.append((subword, orientation, tuple(int(e) for e in re.match(r"\((\d+), (\d+)\)", idxs).groups())))

    idx_row += 1

    assert sheet.cell(idx_row, 1).value == "Subwords not included", "Structure of input file is not correct"

    # We do not do anything with the subwords not included, but have to process those as other values are defined below
    while True:
        idx_row += 1
        text = sheet.cell(idx_row, 1).value
        if not text:
            break

    idx_row += 1

    print(" reading default layout..")

    if size in SIZES_REQUIRING_STANDARD_LAYOUT and layout_boxes is None:
        assert sheet.cell(idx_row, 1).value == "Layout", \
            f"Structure of input file is not correct: should specify a default layout for grid size {size}"

        layout = sheet.cell(idx_row, 2).value

        idx_row += 2
    else:
        assert sheet.cell(idx_row, 1).value != "Layout", \
            f"Structure of input file is not correct: a default layout should not be specified when a custom layout is specified"

        layout = None

    print(" ", layout)

    print(" reading option whether diagonal was fixed..")

    if sheet.cell(idx_row, 1).value == "Fixed diagonal":
        _value = sheet.cell(idx_row, 2).value
        assert _value in list(map(str, [True, False])), "Provided values in input file not correct"
        remove_diagonal = _value == "True"
    else:
        print("  WARNING: Could not read from file whether the main diagonal was fixed, asking for input..")
        # Could not read from file, ask the user for input
        while True:
            _input = input(f"  Remove the main diagonal [y/n]? ")
            print("Input:", _input)
            if _input == "y":
                remove_diagonal = True
                break
            elif _input == "n":
                remove_diagonal = False
                break
            else:
                print("Input not recognised, please try again..")
                continue

    print(f"  {remove_diagonal}")

    # Preprocess boxes layout
    assert (layout_boxes is not None) + (layout is not None) == 1 or size not in SIZES_REQUIRING_STANDARD_LAYOUT

    solution = Instance(solution, layout=layout, layout_boxes=layout_boxes)

    print("Verifying input solution..")
    _verify_inputs(solution, subwords_placements)

    return solution, subwords_placements, remove_diagonal


def _verify_inputs(solution, subwords_placements):
    """
    Verify that the solution & subwords placement in the file are valid
    """

    verify_base_constraints(solution)

    for subword, orientation, placement_idx in subwords_placements:
        idxs = get_idxs_for_subword_placement(subword, orientation, placement_idx)
        for i, (i1, i2) in enumerate(idxs):
            assert solution[i1][i2] == subword[i], f"Placement of subword \"{subword}\" not correctly specified in inputs"


def read_list_of_solutions_from_excel_file(file_name, check_for_duplicate_solutions=True):
    """
    Steps:
     - Read raw data
     - Group data
     - Reconstruct solutions
     - [optional] Screen for duplicate solutions
    """

    # For now we only enable reading solutions with grid size 9
    # TODO Make this a variable
    SIZE = 9

    # TODO Add functionality to read custom boxes layout from file with list of inputs

    print("Reading list of solutions..")

    workbook = load_workbook(file_name)
    sheet = workbook.worksheets[0]

    SKIP_NUMBER_ROWS = 2

    # Reading data
    print("Reading data..")
    max_row = sheet.max_row  # Make sure to do this once outside the loop as this is an expensive operation
    print(" Number of rows:", max_row)
    rows = []
    for idx_row in range(SKIP_NUMBER_ROWS, max_row):
        solution_id = sheet.cell(1 + idx_row, 1).value
        line_chars = list(map(str, (sheet.cell(1 + idx_row, 1 + 1 + i2).value for i2 in range(SIZE))))
        line_layout = [sheet.cell(1 + idx_row, 1 + 1 + SIZE + i2).value for i2 in range(SIZE)]
        rows.append((solution_id, (line_chars, line_layout)))
        if idx_row % 1_000 == 0:
            print(f"Progress: {idx_row} / {max_row}")

    # Group lines
    print("Processing data..")
    grouped_rows = []
    cur_solution_id = None
    cur_solution_lines = []
    for solution_id, line in rows:
        if solution_id == cur_solution_id:
            cur_solution_lines.append(line)
        else:
            if cur_solution_id is not None:
                grouped_rows.append((cur_solution_id, cur_solution_lines))
            cur_solution_id = solution_id
            cur_solution_lines = [line]
    if cur_solution_id is not None:
        grouped_rows.append((cur_solution_id, cur_solution_lines))

    def reconstruct_solution(lines):
        assert len(lines) == SIZE, f"Number of rows incorrect: {len(lines)}"
        # Note: Optionally, a custom boxes layout can be specified
        lines_chars, lines_layout = (list(map(operator.itemgetter(i), lines)) for i in (0, 1))
        try:
            if all(e is None for row in lines_layout for e in row):
                # No layout specified
                solution = Instance(lines_chars)
            else:
                layout_boxes = preprocess_layout_boxes(lines_layout, SIZE)
                solution = Instance(lines_chars, layout_boxes=layout_boxes)
            with contextlib.redirect_stdout(None):
                verify_base_constraints(solution)
        except AssertionError as e:
            raise AssertionError(f"Solution not valid: {e}")
        return solution

    # Reconstruct solutions
    solutions = []
    for solution_id, lines in grouped_rows:
        try:
            solution = reconstruct_solution(lines)
            solutions.append((solution_id, solution))
        except Exception as e:
            print(f"Could not process solution with ID '{solution_id}': {e}")
            for line in lines:
                print(line)
            continue

    # Identify duplicate solutions
    if check_for_duplicate_solutions:
        _screen_for_duplicate_solutions(solutions)

    return solutions


def _screen_for_duplicate_solutions(solutions):

    print("Screening for duplicate solutions..")

    groups = {}
    for idx, (solution_id, solution) in enumerate(solutions):
        key = str(solution)
        if key in groups:
            groups[key].append(solution_id)
        else:
            groups[key] = [solution_id]

        if idx > 0 and idx % 100 == 0:
            print(f"Progress: {idx} / {len(solutions)}")

    groups_with_duplicate_solutions = [
        instance_ids
        for key, instance_ids in groups.items()
        if len(instance_ids) > 1
    ]

    if len(groups_with_duplicate_solutions) > 0:
        print("WARNING - Identified duplicate solutions:")
        for group in groups_with_duplicate_solutions:
            print(len(group), sorted(group))
