
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from generator.model import EMPTY_CHAR


# An extra logs file was requested, which visualises the value-removal process step-by-step;
# This is similar to the step-by-step solution logs of 'tool_logs', but instead the process is reversed, and
#  unsuccessful removals are included;


IDX_ROW_START = 2
IDX_COL_START = 2


def _read_template(file_name_template, size):
    """
    Read template file and extract relevant formatting
    """

    # TODO Implement logic for different sizes
    assert size == 9, f"No template reading logic defined for grid size {size}"

    # Read template
    workbook = load_workbook(file_name_template)
    sheet = workbook.worksheets[0]

    # Extract formatting
    style_header = sheet.cell(IDX_ROW_START, IDX_COL_START)._style
    style_cells = [
        [
            sheet.cell(IDX_ROW_START + 2 + i1, IDX_COL_START + i2)._style
            for i2 in range(size)
        ]
        for i1 in range(size)
    ]
    style_message = sheet.cell(IDX_ROW_START + 2 + size + 1, IDX_COL_START)._style

    formatting_instance = (style_header, style_cells, style_message)

    return workbook, sheet, formatting_instance


def write_step_to_sheet(sheet, idx_row, idx_col, header, instance, idxs_to_remove, idxs_to_remove_failed, message, formatting_instance):

    style_header, style_cells, style_message = formatting_instance

    size = instance.size

    # Write header
    idxs_header = (idx_row, idx_col)
    cell_header = sheet.cell(*idxs_header)
    cell_header.value = header
    cell_header._style = copy(style_header)
    sheet.merge_cells(
        start_row=idxs_header[0], start_column=idxs_header[1],
        end_row=idxs_header[0], end_column=idxs_header[1] + size - 1
    )

    # Write cells
    # TODO Add color codes to template
    for i1 in range(size):
        for i2 in range(size):
            cell = sheet.cell(idx_row + 2 + i1, idx_col + i2)
            cell.value = instance[i1][i2]
            cell._style = copy(style_cells[i1][i2])

            # Highlight non-empty cells
            if instance[i1][i2] != EMPTY_CHAR:
                cell.fill = PatternFill("solid", start_color="EAEAEA")

            # Highlight idxs earlier failed to be removed
            if (i1, i2) in idxs_to_remove_failed:
                cell.fill = PatternFill("solid", start_color="FF8888")

            # Highlight removed idxs
            if (i1, i2) in idxs_to_remove:
                cell.fill = PatternFill("solid", start_color="89B2FF")

    # Write message
    # Note: For the final instance no message is displayed
    if message is not None:
        idxs_message = (idx_row + 2 + size + 1, idx_col)
        cell_message = sheet.cell(*idxs_message)
        cell_message.value = message
        cell_message._style = copy(style_message)
        sheet.merge_cells(
            start_row=idxs_message[0], start_column=idxs_message[1],
            end_row=idxs_message[0], end_column=idxs_message[1] + size - 1
        )


# TODO Perhaps create a separate "formatting" package, which centralises all logic related to converting to user-
#  understandable output; For now this is the only function copied from tool_logs, but more might follow;
def convert_to_user_readable_value(i1, i2):
    """
    Converts internally used 0-based coords to user-understandable output
    """
    # TODO Implement logic for other dimensions when needed
    idx = "R{}C{}".format(i1 + 1, i2 + 1)
    return idx


def write_grouped_steps_to_file(sheet, idx_row, idx_col, grouped_logs_step, idxs_to_remove_failed, formatting_instance):

    assert len(grouped_logs_step) >= 1
    instance = grouped_logs_step[0][0]

    for logs_step in grouped_logs_step:
        instance_step, idxs_to_remove_step, result_step, counts_techniques_step = logs_step
        assert instance_step == instance
        assert len(set(idxs_to_remove_step).intersection(idxs_to_remove_failed)) == 0

        # TODO This should probably be done in write_step() or a separate processing function
        header = "ATTEMPT TO REMOVE " + ' AND '.join(convert_to_user_readable_value(*idx) for idx in idxs_to_remove_step)

        # TODO Standardise/parametrise result messages
        MAP_RESULT_TO_MESSAGE = {
            "no:single_occurrence": "No, not all characters are present after removing the new value(s)",
            "no:no_unique_solution": "No, the instance does not have a unique solution after removing the new value(s)",
            "no:not_solvable_with_techniques": "No, the resulting instance is can not be solved with the techniques",
        }

        if result_step == "yes":
            message = "Yes, using techniques: " + str(counts_techniques_step)
        else:
            message = MAP_RESULT_TO_MESSAGE[result_step]

        write_step_to_sheet(sheet, idx_row, idx_col, header, instance, idxs_to_remove_step, idxs_to_remove_failed, message, formatting_instance)

        if result_step != "yes":
            idxs_to_remove_failed.extend(idxs_to_remove_step)

        # Update idx
        idx_col += instance.size + 1


def write_logs_to_file(solution_id, solution, instance, logs, file_name_template, file_name_output):

    # TODO The size could be read from the template in some future version
    size = solution.size

    print("Reading template..")

    # Load template
    workbook, sheet, formatting_instance = _read_template(file_name_template, size)

    print("Writing to file..")

    # Unmerge cells
    # Note: This has to be done before removing rows, as otherwise they will not be present in the sheet and cannot be
    #  unmerged
    # TODO Only unmerge from the row where the sheet is cleaned
    for merged_cell_range in sheet.merged_cell_ranges:
        sheet.unmerge_cells(range_string=str(merged_cell_range))

    # Clean sheet
    idx_row_start_clean = 1
    sheet.delete_rows(idx_row_start_clean, sheet.max_row - idx_row_start_clean + 1)

    # Write solution
    header = "SOLUTION"
    message = None
    write_step_to_sheet(sheet, IDX_ROW_START, IDX_COL_START, header, solution, [], [], message, formatting_instance)

    # Write logs
    num_rows_per_step = 2 + size + 2 + 1  # TODO For symmetry possibly extra rows are added

    # Preprocess logs: Aggregate steps for the same instance, which are shown on the same row
    logs_steps = logs["steps"]
    grouped_logs_steps = [logs_steps[:1]]
    for logs_step in logs_steps[1:]:
        instance_step = logs_step[0]
        if instance_step == grouped_logs_steps[-1][0][0]:
            grouped_logs_steps[-1].append(logs_step)
        else:
            grouped_logs_steps.append([logs_step])

    # Keep track of this to highlight in next steps
    idxs_to_remove_failed = []

    idx_row = IDX_ROW_START + num_rows_per_step
    for grouped_logs_step in grouped_logs_steps:
        write_grouped_steps_to_file(sheet, idx_row, 2, grouped_logs_step, idxs_to_remove_failed, formatting_instance)
        idx_row += num_rows_per_step

    # Write instance
    header = "GENERATED INSTANCE"
    message = None
    write_step_to_sheet(sheet, idx_row, 2, header, instance, [], [], message, formatting_instance)

    # Modify the sheet name
    sheet.title = "Steps"

    # Write output to file
    workbook.save(filename=file_name_output)
