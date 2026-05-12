
from openpyxl.reader.excel import load_workbook
import os
import sys

sys.path.append(os.getcwd())

from generator.algo_human import solve_using_human_techniques, TECHNIQUES, TECHNIQUES_REQUIRING_STANDARD_BOXES_LAYOUT
from generator.model import Instance, EMPTY_CHAR, SIZES_REQUIRING_STANDARD_LAYOUT
from generator.boxes import preprocess_layout_boxes


def generate_logs(instance, input_file_path):

    # Check whether a folder structure was specified in the input file name
    if '/' in input_file_path:
        _idx = input_file_path.rindex('/') + 1
        data_folder = input_file_path[:_idx]
        input_file_name = input_file_path[_idx:]
        print("Detected data folder:", data_folder)
    else:
        data_folder = ""
        input_file_name = input_file_path

    output_file_name = input_file_name.split('.')[0] + "_logs" + ".rtf"
    output_file_path = data_folder + output_file_name

    print(f"Write logs to file:", output_file_path)
    output_file = open(output_file_path, "w")

    print(f"Generating logs..")

    stdout_orig = sys.stdout
    sys.stdout = output_file

    # When generating an instance for a solution with a non-standard boxes layout, do not include the boxes-x
    #  techniques, as their implementation requires a standard boxes layout
    techniques_disabled = TECHNIQUES_REQUIRING_STANDARD_BOXES_LAYOUT if instance.uses_custom_boxes_layout else []
    if techniques_disabled:
        print(f"Disabled techniques {techniques_disabled} as the solution uses a custom boxes layout")
    use_techniques = [technique for technique in TECHNIQUES if technique not in techniques_disabled]

    solved_instance, _ = solve_using_human_techniques(instance, use_techniques=use_techniques, show_logs=True)

    output_file.close()
    sys.stdout = stdout_orig

    print("Finished")


def _read_instance(file_name_input):
    """
    Read a single instance from file
    """

    print("Reading instance..")

    workbook = load_workbook(file_name_input)
    sheet = workbook.active

    print(" reading size/layout..")

    # Since we are now using the same code for all grid sizes, and we do not write the grid size to file, we have to
    #  derive it from the data - we assume the first information after the solution grid is always the same
    # TODO Reuse this logic from input.py
    idx_row = 1
    while sheet.cell(idx_row, 1).value not in ["Layout", "Rating"]:
        idx_row += 1

    size = idx_row - 2

    print(f"  size: {size}")

    print(" reading grid..")

    instance = [
        [
            (lambda value: str(value) if value else EMPTY_CHAR)(sheet.cell(i1 + 1, i2 + 1).value)
            for i2 in range(size)
        ]
        for i1 in range(size)
    ]

    print(" reading custom layout..")

    # Optionally: Read custom boxes layout
    if sheet.cell(1, size + 2).value:

        raw_data = [
            [
                sheet.cell(i1 + 1, size + 1 + i2 + 1).value
                for i2 in range(size)
            ]
            for i1 in range(size)
        ]

        layout_boxes = preprocess_layout_boxes(raw_data, size)

    else:
        layout_boxes = None

        print("  not specified")

    idx_row = size + 2

    print(" reading default layout..")

    if size in SIZES_REQUIRING_STANDARD_LAYOUT and layout_boxes is None:
        assert sheet.cell(idx_row, 1).value == "Layout", \
            f"Structure of input file is not correct: should specify a default layout for grid size {size}"

        layout = sheet.cell(idx_row, 2).value
    else:
        assert sheet.cell(idx_row, 1).value != "Layout", \
            f"Structure of input file is not correct: a default layout should not be specified when a custom layout is specified"

        layout = None

    print(" ", layout)

    # Preprocess boxes layout
    assert (layout_boxes is not None) + (layout is not None) == 1 or size not in SIZES_REQUIRING_STANDARD_LAYOUT

    instance = Instance(instance, layout=layout, layout_boxes=layout_boxes)

    # TODO Verify instance

    return instance


if __name__ == "__main__":
    """
    Steps:
     0 Process command
     1 Read data from file
        - TODO Validate
        - TODO Find out which files are used as input, and which structure should be checked
     2 Generate logs and write to file
    """

    # Constants
    # Removed as the tool can now be used for all grid sizes

    # 0 Process command
    args = sys.argv[1:]

    tool_name = f"tool_create"
    syntax = "\n ".join([
        f"  python {tool_name}/logs.py" + " {file_name}",
        "",
        "  file_name:  .xlsx"
    ])

    print()

    # Check for help command
    if len(args) == 1 and args[0] in ("-h", "--help"):
        print("The following syntax is used:")
        print(syntax)
        quit()

    # Verify syntax/inputs
    try:

        assert len(args) == 1, "Only one argument required"

        input_file_path = args[0]
        assert input_file_path.endswith(".xlsx"), "Input file should be a .xlsx file"

    except AssertionError as e:
        print("Inputs not valid:", str(e))
        print()
        message = (
            "Please provide inputs using the following syntax:\n"
            f"{syntax}"
        )
        print(message)
        quit()

    # 1 Read data
    try:
        # Note: We need a custom function for this, there is no existing function which reads generated instances, as
        #  they are only outputted but never read into another tool; We cannot reuse the read_inputs_from_excel_file()
        #  function from inputs as it checks for the full structure of solution files
        instance = _read_instance(input_file_path)
    except FileNotFoundError:
        print(f"Could not find input file {input_file_path}..")
        quit()
    except AssertionError as e:
        print("Inputs invalid:", str(e))
        quit()

    print("Generating logs for instance:")
    print(instance)

    # 2 Generate logs
    generate_logs(instance, input_file_path)
