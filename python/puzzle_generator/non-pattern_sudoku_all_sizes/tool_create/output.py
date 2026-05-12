
import colorama
import itertools
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN

from generator.model import count_non_empty_cells

from tool_create.positioner import subword_overlaps_idx


colorama.init(autoreset=True)


def write_to_excel_file(instance, rating, counts_techniques, weight, number_non_empty_cells, output_file_name):

    workbook = Workbook()
    sheet = workbook.active

    font_default = Font(name="calibri", size=12)

    size = instance.size

    for i1 in range(size):
        for i2 in range(size):
            text = str(instance[i1][i2])

            cell = sheet.cell(i1 + 1, i2 + 1)
            cell.value = str(text)
            cell.font = font_default

    # Draw borders
    for (i1, i2) in itertools.product(range(size), repeat=2):
        idx_box = instance.get_idx_box(i1, i2)

        draw_left_border = (instance.get_idx_box(i1, i2 - 1) != idx_box) if i2 - 1 >= 0 else True
        draw_right_border = (instance.get_idx_box(i1, i2 + 1) != idx_box) if i2 + 1 < size else True
        draw_top_border = (instance.get_idx_box(i1 - 1, i2) != idx_box) if i1 - 1 >= 0 else True
        draw_bottom_border = (instance.get_idx_box(i1 + 1, i2) != idx_box) if i1 + 1 < size else True

        cell = sheet.cell(i1 + 1, i2 + 1)
        cell.border = Border(
            left=Side(border_style=BORDER_THIN) if draw_left_border else cell.border.left,
            right=Side(border_style=BORDER_THIN) if draw_right_border else cell.border.right,
            top=Side(border_style=BORDER_THIN) if draw_top_border else cell.border.top,
            bottom=Side(border_style=BORDER_THIN) if draw_bottom_border else cell.border.bottom,
        )

    # Optionally: Write custom boxes layout to file
    if instance.uses_custom_boxes_layout:

        for (i1, i2) in itertools.product(range(size), repeat=2):
            idx_box = instance.get_idx_box(i1, i2)
            box_id = idx_box + 1  # Convert internal 0-based to user 1-based

            cell = sheet.cell(i1 + 1, size + 1 + i2 + 1)
            cell.value = str(box_id)

    idx_row = size + 2

    # Add extra information to the output file: Layout, rating, techniques used, weight
    if instance.layout is not None:

        sheet[chr(ord("A") + 0) + str(idx_row)] = "Layout"
        sheet[chr(ord("A") + 1) + str(idx_row)] = instance.layout

        idx_row += 1

    sheet[chr(ord("A") + 0) + str(idx_row)] = "Rating"
    sheet[chr(ord("A") + 1) + str(idx_row)] = str(rating)

    idx_row += 1

    sheet[chr(ord("A") + 0) + str(idx_row)] = "Hints"
    sheet[chr(ord("A") + 1) + str(idx_row)] = str(number_non_empty_cells)

    idx_row += 1

    sheet[chr(ord("A") + 0) + str(idx_row)] = "Techniques"
    sheet[chr(ord("A") + 1) + str(idx_row)] = str(counts_techniques)

    idx_row += 1

    sheet[chr(ord("A") + 0) + str(idx_row)] = "Weight"
    sheet[chr(ord("A") + 1) + str(idx_row)] = str(weight)

    workbook.save(filename=output_file_name)


def _write_instance_to_sheet(sheet, idx_row, idx_col, instance):
    # Note: Idxs are 1-based!

    size = instance.size

    for i1 in range(size):
        for i2 in range(size):
            e = instance[i1][i2]
            text = str(e)
            # cell = sheet[idx_row + i1][idx_col + i2]
            # Have to construct the cell ID, as if it is empty it is not included in the tuples
            # cell_id = chr(ord("A") - 1 + idx_col + i2) + str(idx_row + i1)
            # cell_id = _idxs_to_cell_id(idx_row + i1, idx_col + i2)
            cell = sheet.cell(idx_row + i1, idx_col + i2)
            cell.value = text


def write_list_of_solutions_to_excel_file(solutions, instances, file_name_input, output_file_name):

    # For now we only enable writing a list of solutions for grid size 9
    # TODO Make this a variable
    SIZE = 9

    # Reuse the formatting of the input file
    workbook = load_workbook(file_name_input)
    sheet = workbook.worksheets[0]

    idx_first_empty_row = 3

    max_row = sheet.max_row
    sheet.delete_rows(idx_first_empty_row, max_row)

    assert len(solutions) == len(instances)  # Also already done in main file

    idx_row = idx_first_empty_row
    for solution_tuple, instance_tuple in zip(solutions, instances):

        # Unpack solution tuple
        (solution_id, solution, _, _) = solution_tuple

        idx_col = 1

        # Write solution ID
        sheet.cell(idx_row, idx_col).value = solution_id

        idx_col += 1

        # Write solution
        _write_instance_to_sheet(sheet, idx_row, idx_col, solution)

        idx_col += SIZE + 1  # Insert empty col

        # Write header for custom boxes layout
        for i2 in range(SIZE):
            cell = sheet.cell(idx_first_empty_row - 1, idx_col + i2)
            cell.value = "Col {}".format(i2 + 1)

        # Optionally: Write custom boxes layout
        if solution.uses_custom_boxes_layout:

            # TODO Reuse logic of function writing instance to file
            for (i1, i2) in itertools.product(range(solution.size), repeat=2):
                idx_box = solution.get_idx_box(i1, i2)
                box_id = idx_box + 1  # Convert internal 0-based to user 1-based

                cell = sheet.cell(idx_row + i1, idx_col + i2)
                cell.value = str(box_id)

        idx_col += SIZE + 1  # Insert empty col

        # Add header for instance
        for i2 in range(SIZE):
            # idx_col = 1 + 1 + SIZE + 1 + i2
            cell = sheet.cell(idx_first_empty_row - 1, idx_col + i2)
            cell.value = "Col {}".format(i2 + 1)

        if instance_tuple is None:

            # Write apologising message
            sheet.cell(idx_row, idx_col).value = "NO INSTANCE"

        else:

            # Unpack instance tuple
            (instance, counts_techniques, weight, number_non_empty_cells) = instance_tuple

            # Write instance
            _write_instance_to_sheet(sheet, idx_row, idx_col, instance)

            idx_col += SIZE + 1  # Insert one empty col

            # Write instance characteristics
            sheet.cell(idx_row + 0, idx_col + 0).value = "Hints"
            sheet.cell(idx_row + 0, idx_col + 1).value = str(number_non_empty_cells)

            sheet.cell(idx_row + 1, idx_col + 0).value = "Techniques"
            sheet.cell(idx_row + 1, idx_col + 1).value = str(counts_techniques)

            sheet.cell(idx_row + 2, idx_col + 0).value = "Weight"
            sheet.cell(idx_row + 2, idx_col + 1).value = str(weight)

        # Update row idx
        idx_row += SIZE

    workbook.save(filename=output_file_name)


def _determine_highlights(size, subwords_placements):

    highlights = [[False for _ in range(size)] for _ in range(size)]

    for i1 in range(size):
        for i2 in range(size):
            # We can reuse the overlap functionality here
            any_subword_overlaps = any(
                subword_overlaps_idx(subword, placement_i1, placement_i2, orientation, i1, i2)
                for subword, orientation, placement_i1, placement_i2 in subwords_placements
            )
            if any_subword_overlaps:
                highlights[i1][i2] = True

    return highlights


def pretty_print(solution, subwords_placements, highlight_diagonal):

    highlights = _determine_highlights(solution.size, [(subword, orientation, i1, i2) for subword, orientation, (i1, i2) in subwords_placements])

    for i1, row in enumerate(solution):
        print('[', end='')
        for i2, e in enumerate(row):
            if highlights[i1][i2]:
                text = colorama.Fore.RED + e
            elif i1 == i2 and highlight_diagonal:
                text = colorama.Fore.LIGHTYELLOW_EX + e
            else:
                text = e
            print('\'', end='')
            print(text, end='')
            print('\'', end='')
            if i2 < len(row) - 1:
                print(', ', end='')
        print(']')


def write_instances_to_file(file_name_input, all_instances):

    # TODO Generalise this functionality, as it is also used in logs.py

    output_file_name = file_name_input.split('.')[0] + "_all_tries" + ".rtf"

    print(f"Write all instances to file:", output_file_name)
    output_file = open(output_file_name, "w")

    stdout_orig = sys.stdout
    sys.stdout = output_file

    for idx, _instance in enumerate(all_instances):
        instance, counts_techniques, weight = _instance

        # Recalculate stats
        number_non_empty_cells = count_non_empty_cells(instance)

        print(f"Try {idx + 1}")

        print()
        print(instance)

        print(" - Number of hints:", number_non_empty_cells)
        print(" - Weight:", weight)
        print(" - Techniques used:", counts_techniques)
        print()
        print()

    output_file.close()
    sys.stdout = stdout_orig

    print("Finished")
